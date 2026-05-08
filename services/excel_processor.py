from __future__ import annotations

import csv
import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


class RowCountMismatchError(Exception):
    def __init__(self, *, excel_rows: int, source_rows: int) -> None:
        self.excel_rows = excel_rows
        self.source_rows = source_rows
        super().__init__(
            f"Количество строк не совпадает: Excel={excel_rows}, Таблица={source_rows}"
        )


class ExcelProcessor:
    PRICE_COLUMN = "Предлагаемая цена за ед. (без учета НДС)"
    MANUFACTURER_COLUMN = "Производитель"
    TECH_COLUMN = "Технические характеристики"

    _ZERO_PLACEHOLDER_FIELDS = {"price", "total"}

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = "" if value is None else str(value)
        text = text.strip().lower().replace("ё", "е")
        return re.sub(r"[^0-9a-zа-я]+", "", text)

    @classmethod
    def _header_candidates(cls, header: Any) -> list[tuple[str, int]]:
        normalized = cls._normalize_header(header)
        if not normalized:
            return []

        candidates: list[tuple[str, int]] = []
        if "альтернативноенаимен" in normalized:
            candidates.append(("name", 100))
        elif "наименован" in normalized:
            candidates.append(("name", 40))

        if "едизм" in normalized or "единицаизмер" in normalized:
            candidates.append(("unit", 100))

        if "колво" in normalized or "количество" in normalized:
            candidates.append(("qty", 100))

        if "предлагаемаяценазаед" in normalized:
            candidates.append(("price", 100))
        elif (
            "ценазаед" in normalized
            and "лучш" not in normalized
            and "началь" not in normalized
            and "старт" not in normalized
        ):
            candidates.append(("price", 70))

        if "сумма" in normalized and "лучш" not in normalized:
            candidates.append(("total", 90))

        if "срок" in normalized and "постав" in normalized:
            candidates.append(("delivery_time", 100))

        if "производ" in normalized:
            candidates.append(("manufacturer", 100))

        if "техничес" in normalized and "характерист" in normalized:
            candidates.append(("technical", 100))

        if "статус" in normalized and "постав" in normalized:
            candidates.append(("supplier_status", 100))

        if "гарант" in normalized:
            candidates.append(("warranty", 100))

        return candidates

    @classmethod
    def _resolve_target_columns(cls, headers: list[Any]) -> dict[str, int]:
        resolved: dict[str, tuple[int, int]] = {}
        for index, header in enumerate(headers, start=1):
            for field, priority in cls._header_candidates(header):
                current = resolved.get(field)
                if current is None or priority > current[1]:
                    resolved[field] = (index, priority)
        return {field: column for field, (column, _priority) in resolved.items()}

    @classmethod
    def _find_header_row(cls, worksheet: Any) -> tuple[int, dict[str, int]]:
        best_row = 1
        best_columns: dict[str, int] = {}
        best_score = 0
        scan_limit = min(int(worksheet.max_row or 1), 20)

        for row_index in range(1, scan_limit + 1):
            headers = [
                worksheet.cell(row=row_index, column=column_index).value
                for column_index in range(1, int(worksheet.max_column or 1) + 1)
            ]
            columns = cls._resolve_target_columns(headers)
            score = len(columns)
            if score > best_score:
                best_row = row_index
                best_columns = columns
                best_score = score

        return best_row, best_columns

    @staticmethod
    def _is_empty_value(value: Any, *, zero_is_empty: bool = False) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return True
            return zero_is_empty and float(value) == 0.0
        try:
            if pd.isna(value):
                return True
        except Exception:
            return False
        if zero_is_empty:
            try:
                return float(value) == 0.0
            except Exception:
                return False
        return False

    @classmethod
    def _worksheet_data_rows_count(cls, worksheet: Any, header_row: int) -> int:
        for row_index in range(int(worksheet.max_row or header_row), header_row, -1):
            for column_index in range(1, int(worksheet.max_column or 1) + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if not cls._is_empty_value(value):
                    return row_index - header_row
        return 0

    @staticmethod
    def _source_value_from_keys(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
        for key in keys:
            value = row.get(key)
            if not ExcelProcessor._is_empty_value(value):
                return value
        return None

    @staticmethod
    def _number_or_none(value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return None
            return float(value)
        text = (
            str(value)
            .strip()
            .replace("\xa0", " ")
            .replace(" ", "")
            .replace("₽", "")
            .replace("руб.", "")
            .replace("руб", "")
            .replace("RUB", "")
            .replace("rub", "")
            .replace(",", ".")
        )
        if not text or re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text) is None:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @classmethod
    def _source_value(cls, row: dict[str, Any], field: str) -> Any:
        if field == "name":
            return cls._source_value_from_keys(row, ("name", "title", "Наименование"))
        if field == "unit":
            return cls._source_value_from_keys(row, ("unit", "unit_name", "Ед. изм."))
        if field == "qty":
            return cls._source_value_from_keys(row, ("qty", "quantity", "Кол-во"))
        if field == "price":
            return cls._source_value_from_keys(
                row,
                ("price", "unit_price", "proposal_price", "Цена за ед."),
            )
        if field == "total":
            explicit_total = cls._source_value_from_keys(
                row,
                ("total", "sum", "amount", "Сумма"),
            )
            if not cls._is_empty_value(explicit_total):
                return explicit_total
            qty = cls._number_or_none(cls._source_value(row, "qty"))
            price = cls._number_or_none(cls._source_value(row, "price"))
            if qty is not None and price is not None:
                return round(qty * price, 2)
            return None
        if field == "delivery_time":
            return cls._source_value_from_keys(
                row,
                ("delivery_time", "supplier_delivery_time", "term", "Срок поставки"),
            )
        if field == "manufacturer":
            return cls._source_value_from_keys(
                row,
                ("manufacturer", "producer", "Производитель"),
            )
        if field == "technical":
            return cls._source_value(row, "manufacturer")
        if field == "supplier_status":
            return cls._source_value_from_keys(
                row,
                ("supplier_status", "status", "Статус поставщика"),
            )
        if field == "warranty":
            return cls._source_value_from_keys(row, ("warranty", "guarantee", "Гарантия"))
        return None

    def can_fill_exported_excel(self, file_path: str) -> bool:
        suffix = Path(file_path).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            workbook = load_workbook(file_path, read_only=True, data_only=False)
            try:
                _header_row, columns = self._find_header_row(workbook.active)
                return bool(columns)
            finally:
                workbook.close()

        if suffix == ".csv":
            df, _delimiter, _encoding = self._read_csv(file_path, nrows=0)
            return bool(self._resolve_target_columns(list(df.columns)))

        df = pd.read_excel(file_path, nrows=0)
        return bool(self._resolve_target_columns(list(df.columns)))

    def fill_exported_excel(
        self,
        file_path: str,
        source_rows: list,
        *,
        strict_row_count: bool = True,
        overwrite_existing: bool = False,
    ) -> None:
        suffix = Path(file_path).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            self._fill_openpyxl(
                file_path,
                source_rows,
                strict_row_count=strict_row_count,
                overwrite_existing=overwrite_existing,
            )
            return

        if suffix == ".csv":
            self._fill_csv(
                file_path,
                source_rows,
                strict_row_count=strict_row_count,
                overwrite_existing=overwrite_existing,
            )
            return

        self._fill_pandas(
            file_path,
            source_rows,
            strict_row_count=strict_row_count,
            overwrite_existing=overwrite_existing,
        )

    def _fill_openpyxl(
        self,
        file_path: str,
        source_rows: list,
        *,
        strict_row_count: bool,
        overwrite_existing: bool,
    ) -> None:
        workbook = load_workbook(file_path)
        try:
            worksheet = workbook.active
            header_row, target_columns = self._find_header_row(worksheet)
            if not target_columns:
                raise Exception("Не найдена колонка для заполнения")

            excel_rows = self._worksheet_data_rows_count(worksheet, header_row)
            if strict_row_count and excel_rows != len(source_rows):
                raise RowCountMismatchError(
                    excel_rows=excel_rows,
                    source_rows=len(source_rows),
                )

            rows_to_copy = excel_rows if strict_row_count else min(excel_rows, len(source_rows))
            for index in range(rows_to_copy):
                row = source_rows[index] if isinstance(source_rows[index], dict) else {}
                excel_row = header_row + 1 + index
                for field, column in target_columns.items():
                    value = self._source_value(row, field)
                    if self._is_empty_value(value):
                        continue
                    cell = worksheet.cell(row=excel_row, column=column)
                    zero_is_empty = field in self._ZERO_PLACEHOLDER_FIELDS
                    if overwrite_existing or self._is_empty_value(
                        cell.value,
                        zero_is_empty=zero_is_empty,
                    ):
                        cell.value = value

            workbook.save(file_path)
        finally:
            workbook.close()

    def _fill_pandas(
        self,
        file_path: str,
        source_rows: list,
        *,
        strict_row_count: bool,
        overwrite_existing: bool,
    ) -> None:
        dataframe = pd.read_excel(file_path)
        self._fill_dataframe(
            dataframe,
            source_rows,
            strict_row_count=strict_row_count,
            overwrite_existing=overwrite_existing,
        )
        dataframe.to_excel(file_path, index=False)

    @staticmethod
    def _read_csv(file_path: str, *, nrows: int | None = None) -> tuple[pd.DataFrame, str, str]:
        path = Path(file_path)
        encoding = "utf-8-sig"
        try:
            sample = path.read_text(encoding=encoding)[:4096]
        except UnicodeDecodeError:
            encoding = "cp1251"
            sample = path.read_text(encoding=encoding)[:4096]

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","

        dataframe = pd.read_csv(
            file_path,
            sep=delimiter,
            encoding=encoding,
            nrows=nrows,
        )
        return dataframe, delimiter, encoding

    def _fill_csv(
        self,
        file_path: str,
        source_rows: list,
        *,
        strict_row_count: bool,
        overwrite_existing: bool,
    ) -> None:
        dataframe, delimiter, encoding = self._read_csv(file_path)
        self._fill_dataframe(
            dataframe,
            source_rows,
            strict_row_count=strict_row_count,
            overwrite_existing=overwrite_existing,
        )
        dataframe.to_csv(file_path, sep=delimiter, index=False, encoding=encoding)

    def _fill_dataframe(
        self,
        dataframe: pd.DataFrame,
        source_rows: list,
        *,
        strict_row_count: bool,
        overwrite_existing: bool,
    ) -> None:
        target_columns = self._resolve_target_columns(list(dataframe.columns))
        if not target_columns:
            raise Exception("Не найдена колонка для заполнения")

        if strict_row_count and len(dataframe) != len(source_rows):
            raise RowCountMismatchError(
                excel_rows=len(dataframe),
                source_rows=len(source_rows),
            )

        rows_to_copy = len(dataframe) if strict_row_count else min(len(dataframe), len(source_rows))
        headers = list(dataframe.columns)
        for field, column_number in target_columns.items():
            if field in {"price", "total", "qty"}:
                continue
            column_name = headers[column_number - 1]
            dataframe[column_name] = dataframe[column_name].astype("object")

        for index in range(rows_to_copy):
            row = source_rows[index] if isinstance(source_rows[index], dict) else {}
            for field, column_number in target_columns.items():
                value = self._source_value(row, field)
                if self._is_empty_value(value):
                    continue
                column_name = headers[column_number - 1]
                zero_is_empty = field in self._ZERO_PLACEHOLDER_FIELDS
                if overwrite_existing or self._is_empty_value(
                    dataframe.at[index, column_name],
                    zero_is_empty=zero_is_empty,
                ):
                    dataframe.at[index, column_name] = value
