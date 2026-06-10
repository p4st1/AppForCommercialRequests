from __future__ import annotations

import ast
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from PySide6.QtCore import QSettings, QThread, Signal, QTimer, Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDoubleSpinBox,
    QFileDialog,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QWidget,
)

from app.ui.table_autosize import configure_table_autosize, resize_table_to_contents
from config import Config
from retrade.retrade_service import RetradeService
from services.currency_service import CurrencyService
from services.excel_processor import ExcelProcessor, RowCountMismatchError
from services.excel_recalc import force_excel_recalc
from services.google_drive_service import GoogleDriveService
from services.trade_exporter import TradeExporter
from tools import DatabaseTools as Tool


class ExportTradeWorker(QThread):
    finished = Signal(str)
    error = Signal(str)

    def __init__(
        self,
        *,
        trade_id: int | None = None,
        lot_id: int | None = None,
        bid_id: int | None = None,
        is_retrade: bool = False,
        is_submission_acceptance: bool = False,
        submission_search_text: str = "",
        download_path: str,
        parent: Any = None,
    ) -> None:
        super().__init__(parent)
        self._trade_id = int(trade_id) if trade_id is not None else None
        self._lot_id = int(lot_id) if lot_id is not None else None
        self._bid_id = int(bid_id) if bid_id is not None else None
        self._is_retrade = bool(is_retrade)
        self._is_submission_acceptance = bool(is_submission_acceptance)
        self._submission_search_text = str(submission_search_text or "")
        if self._trade_id is None and self._lot_id is None:
            raise ValueError("Не указан trade_id или lot_id для экспорта")
        self._download_path = str(download_path)

    def run(self) -> None:
        try:
            exporter = TradeExporter()
            if self._lot_id is not None and self._is_retrade:
                saved_path = exporter.export_retrade_lot_data(
                    lot_id=self._lot_id,
                    download_path=self._download_path,
                    trade_id=self._trade_id,
                    bid_id=self._bid_id,
                )
            elif self._lot_id is not None and self._is_submission_acceptance:
                saved_path = exporter.export_submission_lot_data(
                    lot_id=self._lot_id,
                    download_path=self._download_path,
                    trade_id=self._trade_id,
                    trade_search_text=self._submission_search_text,
                )
            elif self._lot_id is not None:
                saved_path = exporter.export_lot_data(
                    lot_id=self._lot_id,
                    download_path=self._download_path,
                )
            else:
                saved_path = exporter.export_trade_data(
                    trade_id=int(self._trade_id),
                    download_path=self._download_path,
                )
            self.finished.emit(saved_path)
        except Exception as exc:
            self.error.emit(str(exc))


class ImportRetradeWorker(QThread):
    finished = Signal(str)
    error = Signal(str)

    def __init__(
        self,
        *,
        bid_id: int,
        file_path: str,
        parent: Any = None,
    ) -> None:
        super().__init__(parent)
        self._bid_id = int(bid_id)
        self._file_path = str(file_path)

    def run(self) -> None:
        try:
            imported_path = RetradeService.import_excel(
                bid_id=self._bid_id,
                file_path=self._file_path,
            )
            self.finished.emit(imported_path)
        except Exception as exc:
            self.error.emit(str(exc))


class ExportMixin:
    AUTO_TRADE_TIMER_MIN_MINUTES = 1
    AUTO_TRADE_TIMER_MAX_MINUTES = 1440
    EXPORT_MISMATCH_OPEN_WITHOUT_COPY = "open_without_copy"
    EXPORT_MISMATCH_OPEN_AND_COPY = "open_and_copy"
    EXPORT_MISMATCH_CANCEL = "cancel"
    RATING_COLUMN_INDEX = 10
    BEST_PRICE_COLUMN_INDEX = 11
    RETRADE_INNER_TAB_MAIN = 0
    RETRADE_INNER_TAB_CALCULATIONS = 1
    RETRADE_INNER_TAB_HISTORY = 2
    TABLE_SETTINGS_ORG = "MyApp"
    TABLE_SETTINGS_APP = "TableSettings"
    TABLE_SETTINGS_FONT_MIN = 8
    TABLE_SETTINGS_FONT_MAX = 24
    RETRADE_MIN_MARGIN_DEFAULT = 1.1
    RETRADE_DELTA_PERCENT_DEFAULT = 2.0
    RETRADE_ROW_CHECKBOX_COLUMN_MARKER = "retrade_row_checkbox_column"
    RETRADE_MIN_MARGIN_HIGHLIGHT_ROLE_OFFSET = 101
    RETRADE_MIN_MARGIN_PREVIOUS_BACKGROUND_ROLE_OFFSET = 102
    RETRADE_MIN_MARGIN_HIGHLIGHT_COLOR = QColor(198, 239, 206)
    NO_FORMAT_COLUMNS = {
        "каталожный номер",
        "№",
        "артикул",
        "код",
    }
    TABLE_SETTINGS_DEFAULTS = {
        "font_size": 13,
        "bg_color": "#ffffff",
        "text_color": "#000000",
        "header_color": "#f3f3f3",
        "selection_color": "#cce5ff",
        "alternating": True,
    }
    RETRADE_UI_FILE = "retrade.ui"

    def init_export_mixin(self) -> None:
        self._export_trade_worker: ExportTradeWorker | None = None
        self._retrade_import_worker: ImportRetradeWorker | None = None
        self.excel_processor = ExcelProcessor()
        self._auto_trade_timer: QTimer | None = None
        self.current_retrade_excel_path = ""
        self.current_retrade = ""
        self.current_retrade_context: dict[str, Any] = {}
        self.current_retrade_bid_id: int | None = None
        self.current_retrade_trade_id: int | None = None
        self.current_retrade_lot_id: int | None = None
        self.current_retrade_last_export_at = ""
        self.current_retrade_calculations_drive_file_id = ""
        self.current_retrade_calculations_drive_link = ""
        self.current_retrade_calculations_drive_name = ""
        self._pending_retrade_bid_id: int | None = None
        self._pending_retrade_context: dict[str, Any] = {}
        self._active_export_workflow = ""
        self._generate_retrade_after_export = False
        self._pending_submission_export_metadata: dict[str, str] = {}
        self.current_submission_acceptance_excel_path = ""
        self._updating_retrade_main_table = False
        self.calculations_file_path = ""
        self.workbook = None
        self.main_sheet_name = ""
        self.current_calculations_sheet_name = ""
        self.retrade_calculations_loaded = False
        self.retrade_calculations_data: dict[str, Any] = {
            "headers": [],
            "rows": [],
            "total_without_vat": None,
            "total_without_vat_currency": None,
            "totals": {
                "price": 0.0,
                "logistic": 0.0,
                "customs": 0.0,
            },
            "totals_currency": {
                "price": None,
                "logistic": None,
                "customs": None,
            },
        }
        self._ensure_auto_trade_timer()
        self._apply_main_table_font_settings()
        self._ensure_retrade_tab()
        self._ensure_auto_resize_columns_action()
        self._ensure_export_button()
        self.btn_export_trade.clicked.connect(self.export_selected_trade)
        self.btn_export_retrade.clicked.connect(self.export_selected_retrade)

    def _ensure_auto_resize_columns_action(self) -> None:
        if hasattr(self, "action_auto_resize_columns"):
            return

        action = QAction("Подогнать ширину столбцов", self)
        action.setObjectName("action_auto_resize_columns")
        action.setShortcut("Ctrl+Alt+A")
        action.triggered.connect(self._handle_auto_resize_columns)

        edit_menu = getattr(getattr(self, "ui", None), "EditMenu", None)
        if edit_menu is not None:
            edit_menu.addAction(action)

        self.action_auto_resize_columns = action
        self.ui.action_auto_resize_columns = action

    @staticmethod
    def _auto_resize_columns(table: QTableWidget) -> None:
        configure_table_autosize(table)
        table.resizeColumnsToContents()
        max_width = 400
        for column_index in range(table.columnCount()):
            width = table.columnWidth(column_index)
            if width > max_width:
                table.setColumnWidth(column_index, max_width)
        if table.columnCount() > 1:
            table.horizontalHeader().setSectionResizeMode(
                1,
                QHeaderView.ResizeMode.Interactive,
            )
            table.setColumnWidth(1, 300)
        table.resizeRowsToContents()
        table.viewport().update()

    @classmethod
    def _normalize_table_bool_setting(cls, raw_value: Any, default: bool) -> bool:
        if isinstance(raw_value, bool):
            return raw_value
        if isinstance(raw_value, (int, float)):
            return bool(raw_value)
        if isinstance(raw_value, str):
            text = raw_value.strip().lower()
            if text in {"1", "true", "yes", "on"}:
                return True
            if text in {"0", "false", "no", "off"}:
                return False
        return bool(default)

    @classmethod
    def _normalize_table_color_setting(cls, raw_value: Any, default: str) -> str:
        text = str(raw_value or "").strip()
        color = QColor(text)
        if color.isValid():
            return color.name()
        fallback = QColor(default)
        if fallback.isValid():
            return fallback.name()
        return "#ffffff"

    @classmethod
    def _load_table_settings(cls) -> dict[str, Any]:
        defaults = cls.TABLE_SETTINGS_DEFAULTS
        settings = QSettings(cls.TABLE_SETTINGS_ORG, cls.TABLE_SETTINGS_APP)

        font_size_raw = settings.value("font_size", defaults["font_size"])
        try:
            font_size = int(font_size_raw)
        except (TypeError, ValueError):
            font_size = int(defaults["font_size"])
        font_size = max(cls.TABLE_SETTINGS_FONT_MIN, min(cls.TABLE_SETTINGS_FONT_MAX, font_size))

        bg_color = cls._normalize_table_color_setting(
            settings.value("bg_color", defaults["bg_color"]),
            str(defaults["bg_color"]),
        )
        text_color = cls._normalize_table_color_setting(
            settings.value("text_color", defaults["text_color"]),
            str(defaults["text_color"]),
        )
        header_color = cls._normalize_table_color_setting(
            settings.value("header_color", defaults["header_color"]),
            str(defaults["header_color"]),
        )
        selection_color = cls._normalize_table_color_setting(
            settings.value("selection_color", defaults["selection_color"]),
            str(defaults["selection_color"]),
        )
        alternating = cls._normalize_table_bool_setting(
            settings.value("alternating", defaults["alternating"]),
            bool(defaults["alternating"]),
        )

        return {
            "font_size": font_size,
            "bg_color": bg_color,
            "text_color": text_color,
            "header_color": header_color,
            "selection_color": selection_color,
            "alternating": alternating,
        }

    @classmethod
    def _apply_font_size_and_geometry(
        cls,
        table: QTableWidget,
        font_size: int,
        *,
        bold_header: bool = True,
    ) -> None:
        table_font = table.font()
        table_font.setPixelSize(font_size)
        table.setFont(table_font)

        row_height = 24
        vertical_header = table.verticalHeader()
        vertical_header.setMinimumSectionSize(row_height)
        vertical_header.setDefaultSectionSize(row_height)

        horizontal_header = table.horizontalHeader()
        header_font = horizontal_header.font()
        header_font.setPixelSize(font_size)
        header_font.setBold(bold_header)
        horizontal_header.setFont(header_font)

        header_height = max(28, horizontal_header.fontMetrics().height() + 12)
        horizontal_header.setFixedHeight(header_height)

        configure_table_autosize(table, min_row_height=row_height)

    @classmethod
    def apply_table_settings(cls, table: QTableWidget) -> None:
        table_settings = cls._load_table_settings()
        font_size = int(table_settings["font_size"])
        bg_color = str(table_settings["bg_color"])
        text_color = str(table_settings["text_color"])
        header_color = str(table_settings["header_color"])
        selection_color = str(table_settings["selection_color"])

        cls._apply_font_size_and_geometry(table, font_size, bold_header=True)

        table.setShowGrid(True)
        table.setGridStyle(Qt.PenStyle.SolidLine)
        table.setStyleSheet(
            f"""
QTableWidget {{
    background-color: {bg_color};
    color: {text_color};
    gridline-color: #d0d0d0;
    font-size: {font_size}px;
}}

QHeaderView::section {{
    background-color: {header_color};
    font-weight: bold;
    border: 1px solid #d0d0d0;
    padding: 4px;
}}

QTableWidget::item {{
    border: 1px solid #e0e0e0;
    padding: 4px;
}}

QTableWidget::item:selected {{
    background-color: {selection_color};
    color: black;
}}

QTableWidget::item:hover {{
    background-color: #f5f5f5;
}}

QTableWidget::indicator {{
    width: 18px;
    height: 18px;
}}
"""
        )
        table.setAlternatingRowColors(bool(table_settings["alternating"]))
        table.setMouseTracking(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        vertical_header = table.verticalHeader()
        vertical_header.setVisible(False)

        horizontal_header = table.horizontalHeader()
        horizontal_header.setStretchLastSection(False)
        horizontal_header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )
        horizontal_header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        if table.columnCount() > 1:
            horizontal_header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
            table.setColumnWidth(1, 300)
        if table.rowCount() > 0:
            resize_table_to_contents(table)

    def _apply_main_table_font_settings(self) -> None:
        table = getattr(getattr(self, "ui", None), "KpTable", None)
        if not isinstance(table, QTableWidget):
            return

        table_settings = self._load_table_settings()
        font_size = int(table_settings["font_size"])
        self._apply_font_size_and_geometry(table, font_size, bold_header=True)

    @classmethod
    def _configure_excel_like_table(cls, table: QTableWidget) -> None:
        cls.apply_table_settings(table)

    def _get_active_retrade_table(self) -> QTableWidget | None:
        tabs = getattr(self, "tabWidget", None)
        if tabs is None:
            tabs = getattr(getattr(self, "ui", None), "tabWidget", None)
        retrade_tab = getattr(self, "retrade_tab", None)
        if not isinstance(tabs, QTabWidget) or retrade_tab is None:
            return None
        if tabs.currentWidget() is not retrade_tab:
            return None

        inner_tabs = getattr(self, "retrade_inner_tabs", None)
        if not isinstance(inner_tabs, QTabWidget):
            return None

        current_index = inner_tabs.currentIndex()
        if current_index == self.RETRADE_INNER_TAB_MAIN:
            table = getattr(self, "retrade_table", None)
            return table if isinstance(table, QTableWidget) else None
        if current_index == self.RETRADE_INNER_TAB_CALCULATIONS:
            table = getattr(self, "retrade_calculations_table", None)
            return table if isinstance(table, QTableWidget) else None
        return None

    def refresh_retrade_table_settings(self) -> None:
        self._apply_main_table_font_settings()
        for table_name in ("retrade_table", "retrade_calculations_table"):
            table = getattr(self, table_name, None)
            if isinstance(table, QTableWidget):
                self.apply_table_settings(table)

    def _handle_auto_resize_columns(self) -> None:
        table = self._get_active_retrade_table()
        if table is None:
            return
        self._auto_resize_columns(table)

    def _get_retrade_ui_path(self) -> Path:
        resource_path = getattr(self, "resourcePath", None)
        candidate_paths = []
        if callable(resource_path):
            candidate_paths.append(Path(resource_path(self.RETRADE_UI_FILE)))
        candidate_paths.append(Path(Tool.resourcePath(self.RETRADE_UI_FILE)))
        candidate_paths.append(Tool.app_dir() / self.RETRADE_UI_FILE)

        for candidate_path in candidate_paths:
            if candidate_path.exists():
                return candidate_path

        searched_paths = ", ".join(str(path) for path in candidate_paths)
        raise RuntimeError(f"Не найден {self.RETRADE_UI_FILE}; проверенные пути: {searched_paths}")

    def _copy_retrade_ui_attrs(self, retrade_tab: QWidget, attr_names: tuple[str, ...]) -> None:
        for attr_name in attr_names:
            widget = retrade_tab.findChild(QWidget, attr_name)
            if widget is None:
                widget = retrade_tab.findChild(QHBoxLayout, attr_name)
            if widget is None:
                raise RuntimeError(f"В {self.RETRADE_UI_FILE} не найден элемент {attr_name}")
            setattr(self, attr_name, widget)
            setattr(self.ui, attr_name, widget)

    def _load_retrade_ui(self, tabs: QTabWidget) -> QWidget:
        ui_path = self._get_retrade_ui_path()
        loader = QUiLoader()
        retrade_tab = loader.load(str(ui_path), tabs)
        if not isinstance(retrade_tab, QWidget):
            raise RuntimeError(f"Не удалось загрузить интерфейс {ui_path}")
        return retrade_tab

    def _ensure_retrade_tab(self) -> None:
        if hasattr(self, "table_retrade") and hasattr(self, "retrade_tab"):
            return

        tabs = getattr(self, "tabWidget", None)
        if tabs is None:
            tabs = getattr(getattr(self, "ui", None), "tabWidget", None)
        if not isinstance(tabs, QTabWidget):
            raise RuntimeError("Не найден tabWidget для вкладки Переторжка")

        retrade_tab = self._load_retrade_ui(tabs)

        self._copy_retrade_ui_attrs(
            retrade_tab,
            (
                "table_retrade",
                "retrade_inner_tabs",
                "retrade_main_table_tab",
                "retrade_calculations_tab",
                "retrade_history_tab",
                "retrade_top_controls_layout",
                "retrade_calculations_container",
                "retrade_calculations_totals",
                "btn_auto_trade",
                "btn_open_retrade_calculations",
                "btnGenerate",
                "btn_load_retrade_excel",
                "retrade_controls_layout",
                "label_auto_trade_status",
                "label_retrade_calculations_status",
                "retrade_calculations_table",
                "retrade_total_without_vat_label",
                "retrade_price_total_label",
                "sum_label",
                "total_label",
                "profit_label",
            ),
        )

        self.retrade_tab = retrade_tab
        self.retrade_table = self.table_retrade
        self.retradingTable = self.retrade_table
        self.calculationsTable = self.retrade_calculations_table
        self.retrade_calculations_container_layout = (
            self.retrade_calculations_container.layout()
        )
        if self.retrade_calculations_container_layout is None:
            raise RuntimeError(
                f"В {self.RETRADE_UI_FILE} не найден layout контейнера расчетов"
            )
        self.total_without_vat_label = self.retrade_total_without_vat_label
        self.price_total_label = self.retrade_price_total_label
        self.ui.retradeTab = retrade_tab
        self.ui.retrade_tab = retrade_tab
        self.ui.retrade_table = self.retrade_table
        self.ui.retradingTable = self.retradingTable
        self.ui.calculationsTable = self.calculationsTable
        self.ui.retrade_calculations_container_layout = (
            self.retrade_calculations_container_layout
        )
        self.ui.total_without_vat_label = self.total_without_vat_label
        self.ui.price_total_label = self.price_total_label
        self.ui.update_retrade_table = self.update_retrade_table

        self._configure_excel_like_table(self.retrade_table)
        self._configure_excel_like_table(self.retrade_calculations_table)
        self._ensure_retrade_context_labels()
        self._ensure_retrade_main_table_controls()
        self._ensure_retrade_calculations_controls()
        self._ensure_retrade_calculations_sheet_selector()
        self._refresh_retrade_context_labels()
        self._set_auto_trade_status(False)
        self._set_retrade_calculations_loaded_status(False)

        self.btn_auto_trade.clicked.connect(self._toggle_auto_trade_status)
        self.btn_open_retrade_calculations.clicked.connect(
            self._open_retrade_calculations
        )
        self.btnGenerate.clicked.connect(self._on_generate_retrade_calculation_clicked)
        self.btn_load_retrade_excel.clicked.connect(self.load_retrade_excel)
        self.retrade_table.itemChanged.connect(
            self._on_retrade_main_table_item_changed
        )

        tab_index = tabs.insertTab(0, retrade_tab, "Переторжка")
        self.retrade_tab_index = tab_index
        self.retrade_inner_tabs.setCurrentIndex(self.RETRADE_INNER_TAB_MAIN)
        self._clear_retrade_calculations_view()

    def _ensure_retrade_context_labels(self) -> None:
        if isinstance(getattr(self, "label_current_retrade", None), QLabel):
            return

        controls_layout = getattr(self, "retrade_top_controls_layout", None)
        if not isinstance(controls_layout, QHBoxLayout):
            return

        parent = getattr(self, "retrade_tab", None)
        self.label_current_retrade = QLabel(parent)
        self.label_current_retrade.setObjectName("label_current_retrade")
        self.label_current_retrade.setText("Текущая переторжка: не прикреплена")

        self.label_retrade_last_export = QLabel(parent)
        self.label_retrade_last_export.setObjectName("label_retrade_last_export")
        self.label_retrade_last_export.setText("Последний экспорт: -")

        insert_index = controls_layout.indexOf(self.btn_load_retrade_excel)
        if insert_index < 0:
            insert_index = 0
        controls_layout.insertWidget(insert_index + 1, self.label_current_retrade)
        controls_layout.insertWidget(insert_index + 2, self.label_retrade_last_export)

        self.ui.label_current_retrade = self.label_current_retrade
        self.ui.label_retrade_last_export = self.label_retrade_last_export

    @staticmethod
    def _format_retrade_datetime(value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y %H:%M:%S")
        text = str(value or "").strip()
        if not text:
            return ""
        return text

    @classmethod
    def _retrade_context_identity(cls, context: dict[str, Any] | None) -> tuple[str, str, str]:
        source = context if isinstance(context, dict) else {}
        return (
            str(source.get("trade_id") or "").strip(),
            str(source.get("lot_id") or "").strip(),
            str(source.get("bid_id") or "").strip(),
        )

    def _format_current_retrade_label(self) -> str:
        context = getattr(self, "current_retrade_context", {})
        if not isinstance(context, dict) or not context:
            return "Текущая переторжка: не прикреплена"

        retrade_number = str(
            context.get("retrade_number")
            or context.get("title")
            or ""
        ).strip()
        bid_number = str(
            context.get("number")
            or context.get("bid_number")
            or ""
        ).strip()
        bid_id = str(context.get("bid_id") or "").strip()

        parts = []
        if retrade_number:
            parts.append(retrade_number)
        if bid_number:
            parts.append(f"заявка {bid_number}")
        if bid_id:
            parts.append(f"bid_id {bid_id}")
        return "Текущая переторжка: " + (", ".join(parts) if parts else "прикреплена")

    def _refresh_retrade_context_labels(self) -> None:
        current_label = getattr(self, "label_current_retrade", None)
        if isinstance(current_label, QLabel):
            current_label.setText(self._format_current_retrade_label())

        last_export_label = getattr(self, "label_retrade_last_export", None)
        if isinstance(last_export_label, QLabel):
            last_export_at = self._format_retrade_datetime(
                getattr(self, "current_retrade_last_export_at", "")
            )
            last_export_label.setText(
                f"Последний экспорт: {last_export_at}" if last_export_at else "Последний экспорт: -"
            )

    def _mark_current_retrade_table_exported_now(self) -> None:
        timestamp = self._format_retrade_datetime(datetime.now())
        self.current_retrade_last_export_at = timestamp
        context = getattr(self, "current_retrade_context", {})
        if isinstance(context, dict):
            context["last_export_at"] = timestamp
            self.current_retrade_context = context
        self._refresh_retrade_context_labels()

    def _ensure_retrade_main_table_controls(self) -> None:
        if isinstance(getattr(self, "import_button", None), QPushButton):
            return

        controls_layout = getattr(self, "retrade_controls_layout", None)
        if not isinstance(controls_layout, QHBoxLayout):
            return

        parent = getattr(self, "retrade_tab", None)
        self.import_button = QPushButton("Обновить предложение", parent)
        self.import_button.setObjectName("import_button")
        self.import_button.setStyleSheet(
            """
QPushButton {
    background-color: #007aff;
    color: #ffffff;
    border: 1px solid #0071eb;
    padding: 6px 12px;
    border-radius: 8px;
    font-weight: 600;
    font-size: 13px;
    min-height: 32px;
    min-width: 120px;
}
QPushButton:hover {
    background-color: #0a84ff;
}
QPushButton:pressed {
    background-color: #006bd6;
}
QPushButton:disabled {
    background-color: #d2d2d7;
    color: #ffffff;
    border: 1px solid #c7c7cc;
}
"""
        )

        insert_index = controls_layout.indexOf(self.label_auto_trade_status)
        if insert_index < 0:
            insert_index = controls_layout.count()
        controls_layout.insertWidget(insert_index, self.import_button)

        self.ui.import_button = self.import_button
        self.import_button.clicked.connect(self.on_import_clicked)

    @staticmethod
    def _load_retrade_excel_rows(file_path: str) -> list[list[Any]]:
        workbook = load_workbook(file_path, data_only=True)
        try:
            worksheet = workbook.active
            return [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    @classmethod
    def format_money(cls, value: Any) -> str:
        numeric_value = cls._parse_retrade_number_or_none(value)
        if numeric_value is None:
            return "" if value is None else str(value)
        return f"{numeric_value:,.2f}".replace(",", " ").replace(".", ",")

    @staticmethod
    def _currency_format(currency: Any) -> str:
        if currency:
            return f'"{currency}"#,##0.00'
        return "#,##0.00"

    @staticmethod
    def _currency_symbol(value: Any) -> str:
        text = str(value or "").strip()
        upper_text = text.upper()
        if "₽" in text or "RUB" in upper_text or "РУБ" in upper_text:
            return "₽"
        if "$" in text or "USD" in upper_text:
            return "$"
        if "€" in text or "EUR" in upper_text:
            return "€"
        if "¥" in text or "CNY" in upper_text or "JPY" in upper_text:
            return "¥"
        if "₸" in text or "KZT" in upper_text:
            return "₸"
        return ""

    @classmethod
    def _currency_display_symbol(cls, value: Any) -> str:
        code = CurrencyService.normalize_currency_code(value)
        if code == "RUB":
            return "₽"
        if code == "USD":
            return "$"
        if code == "EUR":
            return "€"
        if code in {"CNY", "JPY"}:
            return "¥"
        if code == "KZT":
            return "₸"
        return cls._currency_symbol(value) or str(value or "").strip()

    @classmethod
    def _parse_retrade_number_or_none(cls, value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            try:
                if pd.isna(value):
                    return None
            except Exception:
                pass
            return float(value)
        if not isinstance(value, str):
            return None

        text = (
            value.strip()
            .replace("\xa0", " ")
            .replace(" ", "")
            .replace("₽", "")
            .replace("руб", "")
            .replace("RUB", "")
            .replace("rub", "")
            .replace("USD", "")
            .replace("usd", "")
            .replace("EUR", "")
            .replace("eur", "")
            .replace("CNY", "")
            .replace("cny", "")
            .replace("JPY", "")
            .replace("jpy", "")
            .replace("KZT", "")
            .replace("kzt", "")
            .replace("$", "")
            .replace("€", "")
            .replace("¥", "")
            .replace("₸", "")
            .replace(",", ".")
        )
        if not text:
            return None
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text) is None:
            return None
        try:
            return float(text)
        except Exception:
            return None

    @classmethod
    def _find_retrade_main_column(cls, headers: list[Any], kind: str) -> int | None:
        if kind == "proposal_price":
            return cls._find_update_positions_column(headers, "proposal_price")

        fallback_qty_col: int | None = None
        for index, header in enumerate(headers):
            normalized = cls._normalize_table_header(header)
            if not normalized:
                continue

            if kind == "qty":
                if "колво" in normalized or "количество" in normalized:
                    if "предлага" not in normalized:
                        return index
                    if fallback_qty_col is None:
                        fallback_qty_col = index
            elif kind == "total":
                if "сумма" in normalized:
                    return index
            elif kind == "delivery_time":
                if "срок" in normalized and "постав" in normalized:
                    return index
            elif kind == "manufacturer":
                if "производ" in normalized:
                    return index
            elif kind == "technical":
                if "характерист" in normalized:
                    return index

        if kind == "qty":
            return fallback_qty_col
        return None

    @classmethod
    def _get_retrade_main_columns(cls, headers: list[Any]) -> dict[str, int | None]:
        return {
            "proposal_price": cls._find_retrade_main_column(headers, "proposal_price"),
            "qty": cls._find_retrade_main_column(headers, "qty"),
            "total": cls._find_retrade_main_column(headers, "total"),
            "delivery_time": cls._find_retrade_main_column(headers, "delivery_time"),
            "manufacturer": cls._find_retrade_main_column(headers, "manufacturer"),
            "technical": cls._find_retrade_main_column(headers, "technical"),
        }

    @staticmethod
    def _retrade_main_editable_tooltips() -> dict[str, str]:
        return {
            "delivery_time": "Например: 30 дней",
            "manufacturer": "Например: Atlas Copco",
            "technical": "Свободный текст",
        }

    @staticmethod
    def _retrade_main_required_field_labels() -> dict[str, str]:
        return {
            "delivery_time": "Срок поставки",
            "manufacturer": "Производитель",
            "technical": "Технические характеристики",
        }

    @classmethod
    def _retrade_main_editable_columns(
        cls,
        columns: dict[str, int | None],
    ) -> dict[int, str]:
        tooltips = cls._retrade_main_editable_tooltips()
        editable: dict[int, str] = {}
        for kind, tooltip in tooltips.items():
            column = columns.get(kind)
            if column is not None:
                editable[int(column)] = tooltip
        return editable

    @classmethod
    def _recalculate_retrade_main_row_values(
        cls,
        row_values: list[Any],
        columns: dict[str, int | None],
    ) -> list[Any]:
        updated_row = list(row_values)
        price_col = columns.get("proposal_price")
        qty_col = columns.get("qty")
        total_col = columns.get("total")
        if price_col is None or qty_col is None or total_col is None:
            return updated_row

        max_col = max(int(price_col), int(qty_col), int(total_col))
        if len(updated_row) <= max_col:
            updated_row.extend(None for _ in range(max_col + 1 - len(updated_row)))

        price = cls._parse_retrade_number_or_none(updated_row[int(price_col)])
        qty = cls._parse_retrade_number_or_none(updated_row[int(qty_col)])
        if price is None or qty is None:
            return updated_row

        updated_row[int(total_col)] = round(price * qty, 2)
        return updated_row

    @classmethod
    def _retrade_main_display_text(
        cls,
        value: Any,
        col_index: int,
        columns: dict[str, int | None],
    ) -> str:
        if col_index in {columns.get("proposal_price"), columns.get("total")}:
            return cls.format_money(value)
        return "" if value is None else str(value)

    @classmethod
    def _configure_retrade_main_item(
        cls,
        item: QTableWidgetItem,
        value: Any,
        col_index: int,
        columns: dict[str, int | None],
        editable_columns: dict[int, str],
    ) -> None:
        item.setText(cls._retrade_main_display_text(value, col_index, columns))
        item.setData(Qt.ItemDataRole.UserRole, value)
        if col_index in {columns.get("proposal_price"), columns.get("total")}:
            item.setTextAlignment(
                int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            )
        elif cls._is_numeric_table_value(value):
            item.setTextAlignment(
                int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            )
        else:
            item.setTextAlignment(
                int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            )

        if col_index in editable_columns:
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(editable_columns[col_index])

    def _fill_retrade_table_from_excel_rows(self, data: list[list[Any]]) -> None:
        table = getattr(self, "retrade_table", None)
        if table is None:
            raise RuntimeError("Таблица Переторжка не найдена")

        rows = [list(row or []) for row in data]
        headers = rows[0] if rows else []
        data_rows = rows[1:] if rows else []
        cols_count = max(
            [len(headers), *(len(row) for row in data_rows)],
            default=0,
        )

        self._configure_excel_like_table(table)
        table.clear()
        table.setColumnCount(cols_count)
        table.setRowCount(len(data_rows))

        if cols_count > 0:
            header_labels = [
                "" if value is None else str(value)
                for value in headers
            ]
            if len(header_labels) < cols_count:
                header_labels.extend("" for _ in range(cols_count - len(header_labels)))
            table.setHorizontalHeaderLabels(header_labels[:cols_count])
        else:
            header_labels = []

        columns = self._get_retrade_main_columns(header_labels)
        editable_columns = self._retrade_main_editable_columns(columns)
        previous_block_state = table.blockSignals(True)
        self._updating_retrade_main_table = True
        try:
            for row_index, row_values in enumerate(data_rows):
                normalized_row = list(row_values[:cols_count])
                if len(normalized_row) < cols_count:
                    normalized_row.extend(None for _ in range(cols_count - len(normalized_row)))
                normalized_row = self._recalculate_retrade_main_row_values(
                    normalized_row,
                    columns,
                )
                for col_index, cell_value in enumerate(normalized_row[:cols_count]):
                    item = QTableWidgetItem()
                    self._configure_retrade_main_item(
                        item,
                        cell_value,
                        col_index,
                        columns,
                        editable_columns,
                    )
                    table.setItem(row_index, col_index, item)
        finally:
            self._updating_retrade_main_table = False
            table.blockSignals(previous_block_state)

        resize_table_to_contents(table)

    def _recalculate_retrade_main_table_row(
        self,
        table: QTableWidget,
        row: int,
        columns: dict[str, int | None] | None = None,
    ) -> None:
        columns = columns or self._get_retrade_main_columns(self._table_headers(table))
        total_col = columns.get("total")
        price_col = columns.get("proposal_price")
        qty_col = columns.get("qty")
        if total_col is None or price_col is None or qty_col is None:
            return

        price_item = table.item(row, int(price_col))
        qty_item = table.item(row, int(qty_col))
        price = self._parse_retrade_number_or_none(self._table_item_edit_value(price_item))
        qty = self._parse_retrade_number_or_none(self._table_item_edit_value(qty_item))
        if price is None or qty is None:
            return

        total_value = round(price * qty, 2)
        total_item = table.item(row, int(total_col))
        if total_item is None:
            total_item = QTableWidgetItem()
            table.setItem(row, int(total_col), total_item)
        self._configure_retrade_main_item(
            total_item,
            total_value,
            int(total_col),
            columns,
            self._retrade_main_editable_columns(columns),
        )

    def _on_retrade_main_table_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_retrade_main_table or item is None:
            return

        table_getter = getattr(item, "tableWidget", None)
        table = table_getter() if callable(table_getter) else None
        if not isinstance(table, QTableWidget):
            table = getattr(self, "retrade_table", None)
        if not isinstance(table, QTableWidget):
            return

        columns = self._get_retrade_main_columns(self._table_headers(table))
        changed_col = item.column()
        price_col = columns.get("proposal_price")
        qty_col = columns.get("qty")
        if changed_col not in {price_col, qty_col}:
            return

        previous_block_state = table.blockSignals(True)
        self._updating_retrade_main_table = True
        try:
            if changed_col == price_col:
                price_value = self._parse_retrade_number_or_none(
                    self._table_item_edit_value(item)
                )
                if price_value is not None:
                    self._configure_retrade_main_item(
                        item,
                        price_value,
                        changed_col,
                        columns,
                        self._retrade_main_editable_columns(columns),
                    )
            self._recalculate_retrade_main_table_row(table, item.row(), columns)
        finally:
            self._updating_retrade_main_table = False
            table.blockSignals(previous_block_state)

    @staticmethod
    def _extract_delivery_time_text(text: Any) -> str:
        if text is None:
            return ""
        normalized_text = " ".join(str(text).replace("\xa0", " ").split())
        if not normalized_text:
            return ""

        match = re.search(
            r"\b\d+(?:[,.]\d+)?\s*(?:рабочих\s+|календарных\s+)?дн(?:ей|я|ь|\.)?\b",
            normalized_text,
            flags=re.IGNORECASE,
        )
        if match:
            return match.group(0).strip(" ;,.")
        return ""

    @classmethod
    def _extract_kp_delivery_times_from_document(cls, document: Any) -> list[str]:
        delivery_times: list[str] = []
        for table in getattr(document, "tables", []):
            rows = list(getattr(table, "rows", []) or [])
            header_row_index: int | None = None
            delivery_column: int | None = None

            for row_index, row in enumerate(rows):
                cells = [
                    str(getattr(cell, "text", "") or "").strip()
                    for cell in getattr(row, "cells", [])
                ]
                for column_index, cell_text in enumerate(cells):
                    normalized = cls._normalize_table_header(cell_text)
                    if "срок" in normalized and "постав" in normalized:
                        header_row_index = row_index
                        delivery_column = column_index
                        break
                if delivery_column is not None:
                    break

            if delivery_column is not None and header_row_index is not None:
                for row in rows[header_row_index + 1 :]:
                    cells = [
                        str(getattr(cell, "text", "") or "").strip()
                        for cell in getattr(row, "cells", [])
                    ]
                    if delivery_column >= len(cells):
                        continue
                    raw_value = cells[delivery_column]
                    value = cls._extract_delivery_time_text(raw_value) or raw_value
                    if value:
                        delivery_times.append(value)
                continue

            for row in rows:
                for cell in getattr(row, "cells", []):
                    value = cls._extract_delivery_time_text(
                        getattr(cell, "text", "")
                    )
                    if value:
                        delivery_times.append(value)
                        break

        return delivery_times

    @staticmethod
    def _extract_kp_manufacturer_from_document(document: Any) -> str:
        for paragraph in getattr(document, "paragraphs", []):
            text = " ".join(str(getattr(paragraph, "text", "") or "").split())
            if "производитель" not in text.lower():
                continue

            match = re.search(
                r"производитель\s*[-–—:]\s*([^;\n\r]+)",
                text,
                flags=re.IGNORECASE,
            )
            if match:
                return match.group(1).strip(" ;,.")

            parts = re.split(r"[-–—:]", text, maxsplit=1)
            if len(parts) == 2:
                return parts[1].strip(" ;,.")
        return ""

    @classmethod
    def _extract_kp_data_from_document(cls, document: Any) -> dict[str, Any]:
        return {
            "delivery_times": cls._extract_kp_delivery_times_from_document(document),
            "manufacturer": cls._extract_kp_manufacturer_from_document(document),
        }

    @classmethod
    def _load_kp_docx_data(cls, file_path: str) -> dict[str, Any]:
        document = Document(file_path)
        return cls._extract_kp_data_from_document(document)

    @staticmethod
    def _table_item_edit_value(item: Any) -> Any:
        if item is None:
            return None

        data_getter = getattr(item, "data", None)
        if callable(data_getter):
            try:
                value = data_getter(Qt.ItemDataRole.EditRole)
            except Exception:
                value = None
            if value is not None:
                return value

        text_getter = getattr(item, "text", None)
        if callable(text_getter):
            return text_getter()
        return None

    @staticmethod
    def _excel_cell_contains_formula(cell: Any) -> bool:
        value = getattr(cell, "value", None)
        return isinstance(value, str) and value.startswith("=")

    @staticmethod
    def _is_blank_retrade_table_value(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        try:
            return bool(pd.isna(value))
        except Exception:
            return False

    @classmethod
    def _retrade_table_row_has_value(cls, table: Any, row: int) -> bool:
        for col in range(table.columnCount()):
            item = table.item(row, col)
            value = cls._table_item_edit_value(item)
            if not cls._is_blank_retrade_table_value(value):
                return True
        return False

    @classmethod
    def _retrade_main_position_columns(
        cls,
        headers: list[Any],
        columns: dict[str, int | None],
    ) -> list[int]:
        ignored_columns = {
            column
            for column in (
                columns.get("delivery_time"),
                columns.get("manufacturer"),
                columns.get("technical"),
            )
            if column is not None
        }
        position_columns: list[int] = []
        for index, header in enumerate(headers):
            if index in ignored_columns:
                continue
            normalized = cls._normalize_table_header(header)
            if (
                "наимен" in normalized
                or "каталож" in normalized
                or "артикул" in normalized
            ):
                position_columns.append(index)
        return position_columns

    @classmethod
    def _retrade_main_row_has_position(
        cls,
        table: Any,
        row: int,
        position_columns: list[int],
    ) -> bool:
        if not position_columns:
            return cls._retrade_table_row_has_value(table, row)

        for col in position_columns:
            item = table.item(row, col)
            value = cls._table_item_edit_value(item)
            if not cls._is_blank_retrade_table_value(value):
                return True
        return False

    @classmethod
    def _collect_missing_retrade_required_cells(cls, table: Any) -> list[dict[str, Any]]:
        headers = cls._table_headers(table)
        columns = cls._get_retrade_main_columns(headers)
        required_labels = cls._retrade_main_required_field_labels()
        required_columns = [
            (kind, int(column))
            for kind, column in columns.items()
            if kind in required_labels and column is not None
        ]
        if not required_columns:
            return []

        missing: list[dict[str, Any]] = []
        for row in range(table.rowCount()):
            if not cls._retrade_table_row_has_value(table, row):
                continue
            for kind, col in required_columns:
                item = table.item(row, col)
                value = cls._table_item_edit_value(item)
                if cls._is_blank_retrade_table_value(value):
                    missing.append(
                        {
                            "row": row,
                            "column": col,
                            "label": required_labels[kind],
                        }
                    )
        return missing

    @staticmethod
    def _format_missing_retrade_required_cells_message(
        missing_cells: list[dict[str, Any]],
    ) -> str:
        preview_limit = 20
        lines = [
            f"- строка {int(cell['row']) + 1}: {cell['label']}"
            for cell in missing_cells[:preview_limit]
        ]
        if len(missing_cells) > preview_limit:
            lines.append(f"... и ещё {len(missing_cells) - preview_limit}")

        return (
            "Заполните обязательные ячейки перед сохранением:\n"
            + "\n".join(lines)
            + "\n\nМожно вернуться к заполнению или нажать «Пропустить», "
            "чтобы сохранить файл как есть."
        )

    def _confirm_skip_missing_retrade_required_cells(
        self,
        missing_cells: list[dict[str, Any]],
    ) -> bool:
        message_box = QMessageBox(self)
        icon_enum = getattr(QMessageBox, "Icon", None)
        warning_icon = (
            getattr(icon_enum, "Warning", None)
            if icon_enum is not None
            else getattr(QMessageBox, "Warning", None)
        )
        if warning_icon is not None:
            message_box.setIcon(warning_icon)

        message_box.setWindowTitle("Не заполнены ячейки")
        message_box.setText(
            self._format_missing_retrade_required_cells_message(missing_cells)
        )

        role_enum = getattr(QMessageBox, "ButtonRole", QMessageBox)
        action_role = getattr(role_enum, "ActionRole", 0)
        reject_role = getattr(role_enum, "RejectRole", action_role)
        fill_button = message_box.addButton("Заполнить", reject_role)
        skip_button = message_box.addButton("Пропустить", action_role)
        message_box.setDefaultButton(fill_button)
        message_box.setEscapeButton(fill_button)
        message_box.exec()

        return message_box.clickedButton() == skip_button

    @staticmethod
    def _coerce_retrade_save_value(
        value: Any,
        *,
        numeric_text: bool = False,
    ) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            text = value.strip()
            if not text or text.startswith("="):
                return None
            if numeric_text:
                numeric_value = ExportMixin._parse_retrade_number_or_none(value)
                if numeric_value is not None:
                    return float(numeric_value)
            return value
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            try:
                if pd.isna(value):
                    return None
            except Exception:
                pass
            return float(value)
        return value

    def _write_retrade_table_to_worksheet(self, worksheet: Any, table: Any) -> int:
        headers = self._table_headers(table)
        columns = self._get_retrade_main_columns(headers)
        money_columns = {
            column
            for column in (
                columns.get("proposal_price"),
                columns.get("total"),
            )
            if column is not None
        }
        numeric_columns = {
            column
            for column in (
                columns.get("proposal_price"),
                columns.get("qty"),
                columns.get("total"),
            )
            if column is not None
        }
        default_currency = ""
        for header in headers:
            default_currency = self._currency_symbol(header)
            if default_currency:
                break

        written_count = 0
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                raw_value = self._table_item_edit_value(item)
                value = self._coerce_retrade_save_value(
                    raw_value,
                    numeric_text=col in numeric_columns,
                )
                if value is None:
                    continue

                cell = worksheet.cell(row=row + 2, column=col + 1)
                currency = ""
                if col in money_columns:
                    header = headers[col] if col < len(headers) else ""
                    item_text = item.text() if item is not None else ""
                    currency = (
                        self._currency_symbol(raw_value)
                        or self._currency_symbol(item_text)
                        or self._currency_symbol(header)
                        or self._currency_symbol(cell.number_format)
                        or default_currency
                    )

                if self._excel_cell_contains_formula(cell):
                    if col in money_columns:
                        cell.number_format = self._currency_format(currency)
                    continue

                cell.value = value
                if col in money_columns:
                    cell.number_format = self._currency_format(currency)
                written_count += 1
        return written_count

    def _save_current_retrade_excel(self, *, show_success_status: bool = True) -> bool:
        file_path = str(getattr(self, "current_retrade_excel_path", "") or "").strip()
        if not file_path:
            QMessageBox.warning(self, "Ошибка", "Файл не загружен")
            return False

        table = getattr(self, "retrade_table", None)
        if not isinstance(table, QTableWidget):
            QMessageBox.warning(self, "Ошибка", "Основная таблица не найдена")
            return False

        missing_cells = self._collect_missing_retrade_required_cells(table)
        if missing_cells:
            if not self._confirm_skip_missing_retrade_required_cells(missing_cells):
                return False

        workbook = None
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            self._write_retrade_table_to_worksheet(worksheet, table)
            workbook.save(file_path)
        except Exception as exc:
            error_text = f"Не удалось сохранить Excel файл: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return False
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

        if show_success_status:
            self._show_export_status("Файл сохранён", 3_000)
        return True

    def on_save_clicked(self) -> None:
        self._save_current_retrade_excel(show_success_status=True)

    def on_import_clicked(self) -> None:
        file_path = str(getattr(self, "current_retrade_excel_path", "") or "").strip()
        if not file_path or not os.path.exists(file_path):
            QMessageBox.warning(self, "Ошибка", "Текущий Excel файл переторжки не найден")
            return

        if not self._save_current_retrade_excel(show_success_status=False):
            return

        try:
            bid_id = self._get_retrade_bid_id_for_import()
            self._start_retrade_import_worker(bid_id=bid_id, file_path=file_path)
        except Exception as exc:
            self._on_retrade_import_error(str(exc))

    def _get_retrade_bid_id_for_import(self) -> int:
        stored_bid_id = self._get_current_retrade_bid_id()
        if stored_bid_id is not None:
            try:
                return self._parse_positive_bid_id(stored_bid_id)
            except Exception:
                Tool.write_log(f"Некорректный сохранённый bid_id переторжки: {stored_bid_id}")

        return self._get_selected_retrade_bid_id_for_export()

    def _get_current_retrade_bid_id(self) -> int | None:
        context = getattr(self, "current_retrade_context", {})
        if isinstance(context, dict):
            raw_bid_id = context.get("bid_id")
            if raw_bid_id is not None:
                try:
                    return self._parse_positive_bid_id(raw_bid_id)
                except Exception:
                    Tool.write_log(
                        f"Некорректный bid_id текущей переторжки: {raw_bid_id}"
                    )

        stored_bid_id = getattr(self, "current_retrade_bid_id", None)
        if stored_bid_id is not None:
            try:
                return self._parse_positive_bid_id(stored_bid_id)
            except Exception:
                Tool.write_log(
                    f"Некорректный сохранённый bid_id переторжки: {stored_bid_id}"
                )
        return None

    def _start_retrade_import_worker(self, *, bid_id: int, file_path: str) -> None:
        worker = getattr(self, "_retrade_import_worker", None)
        is_running = getattr(worker, "isRunning", None)
        if worker is not None and callable(is_running) and is_running():
            raise RuntimeError("Обновление предложения уже выполняется")

        self._set_retrade_import_loading_state(is_loading=True)
        worker = ImportRetradeWorker(
            bid_id=bid_id,
            file_path=file_path,
            parent=self,
        )
        worker.finished.connect(self._on_retrade_import_finished)
        worker.error.connect(self._on_retrade_import_error)
        self._retrade_import_worker = worker
        worker.start()

    def _set_retrade_import_loading_state(self, *, is_loading: bool) -> None:
        import_button = getattr(self, "import_button", None)
        if isinstance(import_button, QPushButton):
            import_button.setEnabled(not is_loading)
            import_button.setText(
                "Обновление..." if is_loading else "Обновить предложение"
            )

        if is_loading:
            self._show_export_status("Обновление предложения...")

    def _finish_retrade_import(self, status_message: str) -> None:
        self._set_retrade_import_loading_state(is_loading=False)
        worker = getattr(self, "_retrade_import_worker", None)
        self._retrade_import_worker = None
        delete_later = getattr(worker, "deleteLater", None)
        if callable(delete_later):
            delete_later()

        self._show_export_status(status_message, 5_000)

    def _on_retrade_import_finished(self, file_path: str) -> None:
        Tool.write_log(f"Обновление предложения завершено: {file_path}")
        self._finish_retrade_import("Предложение обновлено")

    def _on_retrade_import_error(self, message: str) -> None:
        error_text = str(message or "Неизвестная ошибка")
        Tool.write_log(f"Ошибка обновления предложения: {error_text}")
        QMessageBox.warning(self, "Ошибка обновления предложения", error_text)
        self._finish_retrade_import("Ошибка обновления предложения")

    def load_retrade_excel(self) -> None:
        default_dir_raw = str(Config.config.get("pathToSaveExcel", "")).strip()
        default_dir = (
            str(Path(default_dir_raw).expanduser())
            if default_dir_raw
            else str(Path.home())
        )
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel файл",
            default_dir,
            "Excel Files (*.xlsx *.xls)",
        )
        if not file_path:
            return

        try:
            data = self._load_retrade_excel_rows(file_path)
            self._fill_retrade_table_from_excel_rows(data)
        except Exception as exc:
            error_text = f"Не удалось загрузить Excel файл: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return

        self.current_retrade_excel_path = file_path
        self._clear_current_retrade_context()
        self._activate_retrade_tab()
        self._log_ui(f"Excel спецификации загружен: {file_path}")
        status_bar_getter = getattr(self, "statusBar", None)
        status_bar = status_bar_getter() if callable(status_bar_getter) else None
        if status_bar is not None:
            status_bar.showMessage("Excel спецификации загружен в Переторжку", 5_000)

    def _open_retrade_calculations_tab(self) -> None:
        inner_tabs = getattr(self, "retrade_inner_tabs", None)
        if isinstance(inner_tabs, QTabWidget):
            inner_tabs.setCurrentIndex(self.RETRADE_INNER_TAB_CALCULATIONS)

    def _reload_retrade_calculations_view(
        self,
        selected_sheet_name: str = "",
    ) -> bool:
        file_path = str(getattr(self, "calculations_file_path", "") or "").strip()
        if not file_path:
            return False

        workbook = None
        try:
            workbook = self._load_retrade_calculations_workbook(file_path)
            self._replace_retrade_calculations_workbook(workbook)
            workbook = None
            self._populate_retrade_sheets_list(selected_sheet_name=selected_sheet_name)
            self._open_retrade_calculations_tab()
            self._set_retrade_calculations_loaded_status(True)
            return True
        except Exception as exc:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            Tool.write_log(f"Не удалось обновить отображение расчетов: {exc}")
            return False

    @classmethod
    def _load_retrade_calculation_setting(
        cls,
        key: str,
        default: float,
    ) -> float:
        settings = QSettings(cls.TABLE_SETTINGS_ORG, cls.TABLE_SETTINGS_APP)
        raw_value = settings.value(f"retrade/{key}", default)
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            value = default
        return max(0.0, min(100.0, value))

    def _save_retrade_calculation_settings(self, *_args: Any) -> None:
        settings = QSettings(self.TABLE_SETTINGS_ORG, self.TABLE_SETTINGS_APP)
        settings.setValue("retrade/min_margin", self.get_min_margin())
        settings.setValue("retrade/delta_percent", self.get_delta_percent())

    def _ensure_retrade_calculations_controls(self) -> None:
        if isinstance(getattr(self, "minMarginInput", None), QDoubleSpinBox):
            return

        container = getattr(self, "retrade_calculations_container", None)
        container_layout = getattr(self, "retrade_calculations_container_layout", None)
        if container is None or container_layout is None:
            return

        controls_widget = QWidget(container)
        controls_widget.setObjectName("retradeCalculationsControls")
        controlsLayout = QHBoxLayout(controls_widget)
        controlsLayout.setContentsMargins(0, 0, 0, 8)
        controlsLayout.setSpacing(8)

        self.minMarginLabel = QLabel("Минимальная наценка:", controls_widget)
        self.minMarginInput = QDoubleSpinBox(controls_widget)
        self.minMarginInput.setObjectName("minMarginInput")
        self.minMarginInput.setRange(0, 100)
        self.minMarginInput.setSingleStep(0.1)
        self.minMarginInput.setDecimals(2)
        self.minMarginInput.setValue(
            self._load_retrade_calculation_setting(
                "min_margin",
                self.RETRADE_MIN_MARGIN_DEFAULT,
            )
        )

        self.deltaLabel = QLabel("Дельта, %:", controls_widget)
        self.deltaInput = QDoubleSpinBox(controls_widget)
        self.deltaInput.setObjectName("deltaInput")
        self.deltaInput.setRange(0, 100)
        self.deltaInput.setSingleStep(0.5)
        self.deltaInput.setDecimals(2)
        self.deltaInput.setValue(
            self._load_retrade_calculation_setting(
                "delta_percent",
                self.RETRADE_DELTA_PERCENT_DEFAULT,
            )
        )

        self.allPositionsCheckbox = QCheckBox(
            "Формировать по всем позициям",
            controls_widget,
        )
        self.allPositionsCheckbox.setObjectName("allPositionsCheckbox")
        self.allPositionsCheckbox.setChecked(True)

        self.roundingLabel = QLabel("Знаков после запятой:", controls_widget)
        self.roundingInput = QSpinBox(controls_widget)
        self.roundingInput.setObjectName("roundingInput")
        self.roundingInput.setRange(0, 6)
        self.roundingInput.setValue(2)

        self.btn_update_retrade_positions = QPushButton(
            "Обновить цены",
            controls_widget,
        )
        self.btn_update_retrade_positions.setObjectName("btn_update_retrade_positions")

        controlsLayout.addWidget(self.minMarginLabel)
        controlsLayout.addWidget(self.minMarginInput)
        controlsLayout.addSpacing(20)
        controlsLayout.addWidget(self.deltaLabel)
        controlsLayout.addWidget(self.deltaInput)
        controlsLayout.addSpacing(20)
        controlsLayout.addWidget(self.allPositionsCheckbox)
        controlsLayout.addSpacing(20)
        controlsLayout.addWidget(self.roundingLabel)
        controlsLayout.addWidget(self.roundingInput)
        controlsLayout.addWidget(self.btn_update_retrade_positions)
        controlsLayout.addStretch()

        container_layout.insertWidget(0, controls_widget)

        self.retrade_calculations_controls = controls_widget
        self.retrade_calculations_controls_layout = controlsLayout
        self.ui.minMarginLabel = self.minMarginLabel
        self.ui.minMarginInput = self.minMarginInput
        self.ui.deltaLabel = self.deltaLabel
        self.ui.deltaInput = self.deltaInput
        self.ui.allPositionsCheckbox = self.allPositionsCheckbox
        self.ui.roundingLabel = self.roundingLabel
        self.ui.roundingInput = self.roundingInput
        self.ui.btn_update_retrade_positions = self.btn_update_retrade_positions

        self.minMarginInput.valueChanged.connect(
            self._save_retrade_calculation_settings
        )
        self.minMarginInput.valueChanged.connect(
            self._apply_min_margin_highlighting
        )
        self.deltaInput.valueChanged.connect(
            self._save_retrade_calculation_settings
        )
        self.allPositionsCheckbox.stateChanged.connect(
            self.on_positions_mode_changed
        )
        self.roundingInput.valueChanged.connect(self.refresh_table)
        self.btn_update_retrade_positions.clicked.connect(
            self.update_retrade_positions
        )

    def get_min_margin(self) -> float:
        input_widget = getattr(self, "minMarginInput", None)
        if isinstance(input_widget, QDoubleSpinBox):
            return float(input_widget.value())
        return self.RETRADE_MIN_MARGIN_DEFAULT

    def get_delta_percent(self) -> float:
        input_widget = getattr(self, "deltaInput", None)
        if isinstance(input_widget, QDoubleSpinBox):
            return float(input_widget.value())
        return self.RETRADE_DELTA_PERCENT_DEFAULT

    def get_rounding(self) -> int:
        input_widget = getattr(self, "roundingInput", None)
        if isinstance(input_widget, QSpinBox):
            return int(input_widget.value())
        return 2

    @staticmethod
    def _qt_user_role_offset(offset: int) -> int:
        user_role = Qt.ItemDataRole.UserRole
        return int(getattr(user_role, "value", user_role)) + offset

    @classmethod
    def _min_margin_highlight_role(cls) -> int:
        return cls._qt_user_role_offset(cls.RETRADE_MIN_MARGIN_HIGHLIGHT_ROLE_OFFSET)

    @classmethod
    def _min_margin_previous_background_role(cls) -> int:
        return cls._qt_user_role_offset(
            cls.RETRADE_MIN_MARGIN_PREVIOUS_BACKGROUND_ROLE_OFFSET
        )

    @staticmethod
    def _background_role() -> Any:
        return getattr(Qt.ItemDataRole, "BackgroundRole", None)

    def _find_calculations_rating_column_index(self, table: Any) -> int | None:
        try:
            column_count = int(table.columnCount())
        except Exception:
            return None

        offset = 0
        try:
            offset = 1 if self._has_row_checkbox_column(table) else 0
        except Exception:
            offset = 0

        headers: list[str] = []
        for column in range(offset, column_count):
            header = table.horizontalHeaderItem(column)
            headers.append(self._text_from_table_item(header).strip())

        try:
            return offset + headers.index("Рейтинг")
        except ValueError:
            return None

    @staticmethod
    def _parse_min_margin_rating_value(text: Any) -> float | None:
        normalized = str(text or "").strip().replace("\xa0", "").replace(" ", "")
        if not normalized:
            return None
        normalized = normalized.replace(",", ".")
        try:
            return float(normalized)
        except Exception:
            return None

    @classmethod
    def _clear_min_margin_highlighting(cls, table: Any) -> None:
        highlight_role = cls._min_margin_highlight_role()
        previous_background_role = cls._min_margin_previous_background_role()
        background_role = cls._background_role()

        for row in range(table.rowCount()):
            for column in range(table.columnCount()):
                item = table.item(row, column)
                if item is None:
                    continue
                data_getter = getattr(item, "data", None)
                if not callable(data_getter) or not item.data(highlight_role):
                    continue

                previous_background = item.data(previous_background_role)
                if background_role is not None:
                    item.setData(background_role, previous_background)
                else:
                    item.setBackground(QColor())
                item.setData(highlight_role, None)
                item.setData(previous_background_role, None)

    @classmethod
    def _set_min_margin_row_highlighted(cls, table: Any, row: int) -> None:
        highlight_role = cls._min_margin_highlight_role()
        previous_background_role = cls._min_margin_previous_background_role()
        background_role = cls._background_role()

        for column in range(table.columnCount()):
            item = table.item(row, column)
            if item is None:
                continue

            if not item.data(highlight_role):
                previous_background = (
                    item.data(background_role) if background_role is not None else None
                )
                item.setData(previous_background_role, previous_background)
            item.setBackground(cls.RETRADE_MIN_MARGIN_HIGHLIGHT_COLOR)
            item.setData(highlight_role, True)

    def _apply_min_margin_highlighting(self, *_args: Any) -> None:
        table = self._get_calculations_table()
        if table is None:
            return

        set_updates_enabled = getattr(table, "setUpdatesEnabled", None)
        if callable(set_updates_enabled):
            table.setUpdatesEnabled(False)
        try:
            self._clear_min_margin_highlighting(table)
            rating_col_index = self._find_calculations_rating_column_index(table)
            if rating_col_index is None:
                return

            threshold = self.get_min_margin() - 1
            for row in range(table.rowCount()):
                item = table.item(row, rating_col_index)
                if item is None:
                    continue

                value = self._parse_min_margin_rating_value(item.text())
                if value is None:
                    continue

                if value + 1e-12 >= threshold:
                    self._set_min_margin_row_highlighted(table, row)
        finally:
            if callable(set_updates_enabled):
                table.setUpdatesEnabled(True)

    def _get_calculations_table(self) -> QTableWidget | None:
        table = getattr(self, "calculationsTable", None)
        if isinstance(table, QTableWidget):
            return table

        table = getattr(self, "retrade_calculations_table", None)
        return table if isinstance(table, QTableWidget) else None

    def _has_row_checkbox_column(self, table: QTableWidget) -> bool:
        if table.columnCount() == 0:
            return False

        header_item = table.horizontalHeaderItem(0)
        if header_item is None:
            return False

        return (
            header_item.data(Qt.ItemDataRole.UserRole)
            == self.RETRADE_ROW_CHECKBOX_COLUMN_MARKER
        )

    def format_rubles(self, value: Any) -> str:
        numeric_value = self._parse_retrade_numeric_value(value)
        if numeric_value is None:
            return "" if value is None else str(value)
        precision = self.get_rounding()
        formatted = f"{numeric_value:,.{precision}f}"
        return formatted.replace(",", " ").replace(".", ",") + " ₽"

    def format_number(self, value: Any) -> str:
        numeric_value = self._parse_retrade_numeric_value(value)
        if numeric_value is None:
            return "" if value is None else str(value)
        precision = self.get_rounding()
        formatted = f"{numeric_value:,.{precision}f}"
        return formatted.replace(",", " ").replace(".", ",")

    def format_rating(self, value: Any) -> str:
        numeric_value = self._parse_retrade_numeric_value(value)
        if numeric_value is None:
            return "" if value is None else str(value)
        return f"{numeric_value:.2f}"

    @staticmethod
    def _is_excel_date_like_value(value: Any) -> bool:
        return (
            hasattr(value, "year")
            and hasattr(value, "month")
            and hasattr(value, "day")
        )

    def _format_calculations_display_value(
        self,
        value: Any,
        *,
        col_index: int,
        header: str,
        price_columns: set[int],
        rating_columns: set[int],
        currency: str | None = None,
    ) -> str:
        if value is None:
            return ""
        header_text = str(header or "").strip()
        header_lower = header_text.lower()
        if header_lower in self.NO_FORMAT_COLUMNS:
            return str(value)
        if col_index in rating_columns:
            return self.format_rating(value)
        if isinstance(value, str):
            stripped_value = value.strip()
            if (
                stripped_value
                and self._parse_retrade_number_or_none(stripped_value) is None
            ):
                return value
        if self._is_excel_date_like_value(value):
            return str(value)
        detected_currency = (
            CurrencyService.normalize_currency_code(currency)
            or self._detect_currency(value, header_text)
        )
        if detected_currency:
            return self._format_retrade_calculations_cell_for_display(
                {
                    "value": value,
                    "currency": detected_currency,
                }
            )
        if "₽" in str(value) or "руб" in header_lower:
            return self.format_rubles(value)
        if col_index in price_columns:
            return self.format_number(value)
        if self._parse_retrade_numeric_value(value) is not None:
            return self.format_number(value)
        return str(value)

    def _is_main_calculations_sheet_selected(self) -> bool:
        main_sheet_name = str(getattr(self, "main_sheet_name", "") or "")
        current_sheet_name = str(
            getattr(self, "current_calculations_sheet_name", "") or ""
        )
        return bool(main_sheet_name) and current_sheet_name == main_sheet_name

    def find_number_column(self) -> int | None:
        table = self._get_calculations_table()
        if table is None:
            return None

        def is_number_header(text: str) -> bool:
            normalized = text.strip()
            return normalized == "№" or normalized.casefold() in {"no", "n"}

        start_col = self.get_calculations_column_offset()
        for col in range(start_col, table.columnCount()):
            header = table.horizontalHeaderItem(col)
            if header is not None and is_number_header(header.text()):
                return col

        for row in range(table.rowCount()):
            for col in range(start_col, table.columnCount()):
                item = table.item(row, col)
                if item is not None and is_number_header(item.text()):
                    return col

        return None

    def has_position_number(self, row: int) -> bool:
        table = self._get_calculations_table()
        if table is None:
            return False

        col = self.find_number_column()
        if col is None:
            return False

        item = table.item(row, col)
        if item is None:
            return False

        return item.text().strip().isdigit()

    def _apply_row_checkbox_mode(self) -> None:
        checkbox = getattr(self, "allPositionsCheckbox", None)
        if (
            self._is_main_calculations_sheet_selected()
            and isinstance(checkbox, QCheckBox)
            and not checkbox.isChecked()
        ):
            self.add_row_checkboxes()
        else:
            self.remove_row_checkboxes()

    def get_calculations_column_offset(self) -> int:
        table = self._get_calculations_table()
        if table is None:
            return 0
        return 1 if self._has_row_checkbox_column(table) else 0

    def on_positions_mode_changed(self, *_args: Any) -> None:
        self._apply_row_checkbox_mode()
        self._apply_min_margin_highlighting()

    def add_row_checkboxes(self) -> None:
        table = self._get_calculations_table()
        if table is None or table.columnCount() == 0:
            return

        if not self._is_main_calculations_sheet_selected():
            self.remove_row_checkboxes()
            return

        if self._has_row_checkbox_column(table):
            table.setColumnWidth(0, 40)
            return

        if self.find_number_column() is None:
            return

        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.insertColumn(0)
        header_item = QTableWidgetItem("")
        header_item.setData(
            Qt.ItemDataRole.UserRole,
            self.RETRADE_ROW_CHECKBOX_COLUMN_MARKER,
        )
        table.setHorizontalHeaderItem(0, header_item)

        for row in range(table.rowCount()):
            if not self.has_position_number(row):
                continue

            item = QTableWidgetItem()
            item.setFlags(
                Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled
            )
            item.setCheckState(Qt.CheckState.Checked)
            item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            table.setItem(row, 0, item)

        table.setColumnWidth(0, 40)

    def remove_row_checkboxes(self) -> None:
        table = self._get_calculations_table()
        if table is None or table.columnCount() == 0:
            return

        if self._has_row_checkbox_column(table):
            table.removeColumn(0)

    def get_selected_rows(self) -> list[int]:
        table = self._get_calculations_table()
        if table is None:
            return []

        if not self._is_main_calculations_sheet_selected():
            return []

        checkbox = getattr(self, "allPositionsCheckbox", None)
        if not isinstance(checkbox, QCheckBox) or checkbox.isChecked():
            return [
                row
                for row in range(table.rowCount())
                if self.has_position_number(row)
            ]

        selected_rows: list[int] = []
        if not self._has_row_checkbox_column(table):
            return selected_rows

        for row in range(table.rowCount()):
            if not self.has_position_number(row):
                continue

            item = table.item(row, 0)
            if item is not None and item.checkState() == Qt.CheckState.Checked:
                selected_rows.append(row)

        return selected_rows

    def refresh_table(self, *_args: Any) -> None:
        sheets_list = getattr(self, "sheetsList", None)
        workbook = getattr(self, "workbook", None)
        if not isinstance(sheets_list, QListWidget) or workbook is None:
            return

        current_item = sheets_list.currentItem()
        if current_item is None:
            return

        sheet_name = current_item.text()
        if not sheet_name or sheet_name not in workbook.sheetnames:
            return

        self.current_calculations_sheet_name = sheet_name
        self.display_sheet(workbook[sheet_name])
        self._apply_row_checkbox_mode()
        self._apply_min_margin_highlighting()

    def _ensure_retrade_calculations_sheet_selector(self) -> None:
        if isinstance(getattr(self, "sheetsList", None), QListWidget):
            return

        table = getattr(self, "retrade_calculations_table", None)
        container = getattr(self, "retrade_calculations_container", None)
        container_layout = getattr(self, "retrade_calculations_container_layout", None)
        if (
            not isinstance(table, QTableWidget)
            or container is None
            or container_layout is None
        ):
            return

        self.sheetsList = QListWidget(container)
        self.sheetsList.setObjectName("sheetsList")
        self.sheetsList.setMinimumWidth(160)
        self.sheetsList.setMaximumWidth(260)
        self.sheetsList.setUniformItemSizes(True)

        sheets_view = QWidget(container)
        sheets_view.setObjectName("retradeCalculationsSheetsView")
        layout = QHBoxLayout(sheets_view)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        table_index = container_layout.indexOf(table)
        if table_index < 0:
            table_index = 0
        else:
            container_layout.takeAt(table_index)

        table.setParent(sheets_view)
        layout.addWidget(self.sheetsList, 1)
        layout.addWidget(table, 4)
        container_layout.insertWidget(table_index, sheets_view, 1)

        self.retrade_calculations_sheets_view = sheets_view
        self.retrade_calculations_sheets_layout = layout
        self.ui.sheetsList = self.sheetsList
        self.sheetsList.currentTextChanged.connect(self.on_sheet_selected)

    def _close_retrade_calculations_workbook(self) -> None:
        workbook = getattr(self, "workbook", None)
        if workbook is not None and hasattr(workbook, "close"):
            try:
                workbook.close()
            except Exception:
                pass
        self.workbook = None
        self.main_sheet_name = ""
        self.current_calculations_sheet_name = ""

    def _replace_retrade_calculations_workbook(self, workbook: Any) -> None:
        self._close_retrade_calculations_workbook()
        self.workbook = workbook
        sheet_names = list(getattr(workbook, "sheetnames", []) or [])
        self.main_sheet_name = sheet_names[0] if sheet_names else ""

    @staticmethod
    def _normalize_formula_reference(reference: Any) -> str:
        return str(reference or "").replace("$", "").upper()

    @classmethod
    def _formula_numeric_value(cls, value: Any) -> Any:
        if value is None:
            return 0
        parsed = cls._parse_retrade_number_or_none(value)
        if parsed is not None:
            return parsed
        return value

    @classmethod
    def _flatten_formula_values(cls, values: Any) -> list[Any]:
        if isinstance(values, (list, tuple)):
            flattened: list[Any] = []
            for value in values:
                flattened.extend(cls._flatten_formula_values(value))
            return flattened
        return [values]

    @classmethod
    def _excel_sum(cls, *values: Any) -> float:
        total = 0.0
        for value in cls._flatten_formula_values(values):
            parsed = cls._parse_retrade_number_or_none(value)
            if parsed is not None:
                total += parsed
        return total

    @staticmethod
    def _excel_round(value: Any, digits: Any = 0) -> float:
        try:
            parsed_value = float(value)
        except Exception:
            parsed_value = 0.0
        try:
            parsed_digits = int(float(digits))
        except Exception:
            parsed_digits = 0
        return round(parsed_value, parsed_digits)

    @classmethod
    def _evaluate_retrade_formula_cell(
        cls,
        formula_sheet: Any,
        values_sheet: Any,
        row: int,
        column: int,
        cache: dict[tuple[str, int, int], Any],
        resolving: set[tuple[str, int, int]],
    ) -> Any:
        sheet_title = str(getattr(formula_sheet, "title", ""))
        key = (sheet_title, int(row), int(column))
        if key in cache:
            return cache[key]

        formula_cell = formula_sheet.cell(row=row, column=column)
        values_cell = values_sheet.cell(row=row, column=column)
        formula_value = formula_cell.value

        if not (isinstance(formula_value, str) and formula_value.startswith("=")):
            value = values_cell.value if values_cell.value is not None else formula_value
            cache[key] = value
            return value

        if values_cell.value is not None:
            cache[key] = values_cell.value
            return values_cell.value

        if key in resolving:
            return None

        resolving.add(key)
        try:
            value = cls._evaluate_retrade_formula_expression(
                formula_value,
                formula_sheet,
                values_sheet,
                cache,
                resolving,
            )
        finally:
            resolving.discard(key)

        cache[key] = value
        return value

    @classmethod
    def _formula_cell_value(
        cls,
        reference: Any,
        formula_sheet: Any,
        values_sheet: Any,
        cache: dict[tuple[str, int, int], Any],
        resolving: set[tuple[str, int, int]],
    ) -> Any:
        normalized = cls._normalize_formula_reference(reference)
        match = re.fullmatch(r"([A-Z]{1,3})([1-9]\d*)", normalized)
        if match is None:
            return 0
        column = column_index_from_string(match.group(1))
        row = int(match.group(2))
        value = cls._evaluate_retrade_formula_cell(
            formula_sheet,
            values_sheet,
            row,
            column,
            cache,
            resolving,
        )
        return cls._formula_numeric_value(value)

    @classmethod
    def _formula_range_values(
        cls,
        reference: Any,
        formula_sheet: Any,
        values_sheet: Any,
        cache: dict[tuple[str, int, int], Any],
        resolving: set[tuple[str, int, int]],
    ) -> list[Any]:
        normalized = cls._normalize_formula_reference(reference)
        try:
            min_col, min_row, max_col, max_row = range_boundaries(normalized)
        except Exception:
            return []

        values: list[Any] = []
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                value = cls._evaluate_retrade_formula_cell(
                    formula_sheet,
                    values_sheet,
                    row,
                    column,
                    cache,
                    resolving,
                )
                values.append(cls._formula_numeric_value(value))
        return values

    @classmethod
    def _pythonize_retrade_formula(
        cls,
        formula: Any,
    ) -> str:
        expression = str(formula or "").strip()
        if expression.startswith("="):
            expression = expression[1:].strip()
        if not expression:
            return ""

        expression = expression.replace(";", ",").replace("^", "**")
        expression = re.sub(
            r"\b(round|sum|if|min|max|abs)\s*\(",
            lambda match: f"{match.group(1).upper()}(",
            expression,
            flags=re.IGNORECASE,
        )
        expression = re.sub(r"\bTRUE\b", "True", expression, flags=re.IGNORECASE)
        expression = re.sub(r"\bFALSE\b", "False", expression, flags=re.IGNORECASE)
        expression = expression.replace("<>", "!=")
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)

        range_references: list[str] = []

        def _replace_range(match: re.Match[str]) -> str:
            range_references.append(cls._normalize_formula_reference(match.group(0)))
            return f"__RANGE_{len(range_references) - 1}__"

        expression = re.sub(
            r"\$?[A-Za-z]{1,3}\$?[1-9]\d*:\$?[A-Za-z]{1,3}\$?[1-9]\d*",
            _replace_range,
            expression,
        )

        def _replace_cell(match: re.Match[str]) -> str:
            reference = cls._normalize_formula_reference(match.group(0))
            return f'CELL("{reference}")'

        expression = re.sub(
            r"(?<![A-Za-z0-9_\"'])\$?[A-Za-z]{1,3}\$?[1-9]\d*(?![A-Za-z0-9_\"'])",
            _replace_cell,
            expression,
        )

        for index, reference in enumerate(range_references):
            expression = expression.replace(
                f"__RANGE_{index}__",
                f'RANGE("{reference}")',
            )
        return expression

    @staticmethod
    def _is_safe_retrade_formula_ast(
        parsed_expression: ast.Expression,
        allowed_names: set[str],
    ) -> bool:
        allowed_nodes = (
            ast.Expression,
            ast.BinOp,
            ast.UnaryOp,
            ast.Call,
            ast.Name,
            ast.Load,
            ast.Constant,
            ast.Compare,
            ast.BoolOp,
            ast.Add,
            ast.Sub,
            ast.Mult,
            ast.Div,
            ast.Pow,
            ast.Mod,
            ast.USub,
            ast.UAdd,
            ast.Eq,
            ast.NotEq,
            ast.Lt,
            ast.LtE,
            ast.Gt,
            ast.GtE,
            ast.And,
            ast.Or,
        )
        reference_pattern = re.compile(
            r"^[A-Z]{1,3}[1-9]\d*(?::[A-Z]{1,3}[1-9]\d*)?$"
        )

        for node in ast.walk(parsed_expression):
            if not isinstance(node, allowed_nodes):
                return False
            if isinstance(node, ast.Call):
                if not isinstance(node.func, ast.Name):
                    return False
                if node.func.id not in allowed_names:
                    return False
                if node.keywords:
                    return False
            elif isinstance(node, ast.Name) and node.id not in allowed_names:
                return False
            elif isinstance(node, ast.Constant):
                value = node.value
                if isinstance(value, str):
                    if reference_pattern.fullmatch(value) is None:
                        return False
                elif not isinstance(value, (int, float, bool, type(None))):
                    return False
        return True

    @classmethod
    def _evaluate_retrade_formula_expression(
        cls,
        formula: Any,
        formula_sheet: Any,
        values_sheet: Any,
        cache: dict[tuple[str, int, int], Any],
        resolving: set[tuple[str, int, int]],
    ) -> Any:
        expression = cls._pythonize_retrade_formula(formula)
        if not expression:
            return None

        def _cell(reference: Any) -> Any:
            return cls._formula_cell_value(
                reference,
                formula_sheet,
                values_sheet,
                cache,
                resolving,
            )

        def _range(reference: Any) -> list[Any]:
            return cls._formula_range_values(
                reference,
                formula_sheet,
                values_sheet,
                cache,
                resolving,
            )

        names = {
            "ABS": abs,
            "CELL": _cell,
            "IF": lambda condition, yes=0, no=0: yes if condition else no,
            "MAX": max,
            "MIN": min,
            "RANGE": _range,
            "ROUND": cls._excel_round,
            "SUM": cls._excel_sum,
        }
        try:
            parsed_expression = ast.parse(expression, mode="eval")
            if not cls._is_safe_retrade_formula_ast(
                parsed_expression,
                set(names),
            ):
                return None
            return eval(
                compile(parsed_expression, "<retrade-formula>", "eval"),
                {"__builtins__": {}},
                names,
            )
        except Exception:
            return None

    @classmethod
    def _apply_retrade_formula_display_values(
        cls,
        file_path: str,
        workbook_values: Any,
    ) -> None:
        workbook_formulas = load_workbook(file_path, data_only=False)
        try:
            for sheet_name in list(getattr(workbook_values, "sheetnames", []) or []):
                if sheet_name not in workbook_formulas.sheetnames:
                    continue

                values_sheet = workbook_values[sheet_name]
                formula_sheet = workbook_formulas[sheet_name]
                cache: dict[tuple[str, int, int], Any] = {}
                resolving: set[tuple[str, int, int]] = set()
                for formula_row in formula_sheet.iter_rows(values_only=False):
                    for formula_cell in formula_row:
                        formula_value = formula_cell.value
                        if not (
                            isinstance(formula_value, str)
                            and formula_value.startswith("=")
                        ):
                            continue
                        evaluated = cls._evaluate_retrade_formula_cell(
                            formula_sheet,
                            values_sheet,
                            formula_cell.row,
                            formula_cell.column,
                            cache,
                            resolving,
                        )
                        if evaluated is not None:
                            values_sheet.cell(
                                row=formula_cell.row,
                                column=formula_cell.column,
                            ).value = evaluated
        finally:
            workbook_formulas.close()

    @classmethod
    def _load_retrade_calculations_workbook(cls, file_path: str) -> Any:
        workbook = load_workbook(file_path, data_only=True)
        try:
            cls._apply_retrade_formula_display_values(file_path, workbook)
        except Exception as exc:
            cls._log_calc(f"мини-пересчет формул пропущен: {exc}")
        return workbook

    def _populate_retrade_sheets_list(self, selected_sheet_name: str = "") -> None:
        sheets_list = getattr(self, "sheetsList", None)
        workbook = getattr(self, "workbook", None)
        if not isinstance(sheets_list, QListWidget) or workbook is None:
            return

        sheet_names = list(getattr(workbook, "sheetnames", []) or [])
        self.main_sheet_name = sheet_names[0] if sheet_names else ""
        if sheet_names:
            main_sheet = sheet_names[0]
            if main_sheet != "Рассчеты":
                Tool.write_log("Первый лист должен называться 'Рассчеты'")

        requested_sheet_name = str(selected_sheet_name or "").strip()
        selected_sheet_name = ""
        sheets_list.blockSignals(True)
        try:
            sheets_list.clear()
            for sheet_name in sheet_names:
                item = QListWidgetItem(sheet_name)
                if sheet_name == "Рассчеты":
                    item.setBackground(QColor(200, 255, 200))
                sheets_list.addItem(item)
            if sheets_list.count() > 0:
                target_row = 0
                if requested_sheet_name in sheet_names:
                    target_row = sheet_names.index(requested_sheet_name)
                sheets_list.setCurrentRow(target_row)
                selected_item = sheets_list.currentItem()
                selected_sheet_name = (
                    selected_item.text()
                    if selected_item is not None
                    else ""
                )
        finally:
            sheets_list.blockSignals(False)

        if selected_sheet_name:
            self.on_sheet_selected(selected_sheet_name)

    @staticmethod
    def _worksheet_to_retrade_calculations_cells_data(
        worksheet: Any,
    ) -> list[list[dict[str, Any]]]:
        data: list[list[dict[str, Any]]] = []
        for row in worksheet.iter_rows(values_only=False):
            parsed_row: list[dict[str, Any]] = []
            for cell in row:
                value = cell.value
                currency = ExportMixin._detect_currency(value, cell.number_format)
                parsed_row.append(
                    {
                        "value": value,
                        "currency": currency,
                    }
                )
            data.append(parsed_row)
        return data

    @classmethod
    def _worksheet_to_visible_retrade_rows(
        cls,
        worksheet: Any,
    ) -> tuple[
        list[dict[str, Any]],
        list[tuple[int, list[dict[str, Any]]]],
        int,
    ]:
        visible_rows: list[tuple[int, list[dict[str, Any]]]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=False),
            start=1,
        ):
            parsed_row: list[dict[str, Any]] = []
            for cell in row:
                value = cell.value
                parsed_row.append(
                    {
                        "value": value,
                        "currency": cls._detect_currency(value, cell.number_format),
                    }
                )
            if not cls._is_retrade_calculations_row_present(parsed_row):
                continue
            visible_rows.append((row_number, parsed_row))

        if not visible_rows:
            return [], [], 0

        row_cells_for_columns = [row_cells for _row_number, row_cells in visible_rows]
        visible_column_indices = cls._non_empty_retrade_column_indices(
            row_cells_for_columns
        )
        header_cells = cls._filter_retrade_row_by_indices(
            visible_rows[0][1],
            visible_column_indices,
        )
        data_rows = [
            (
                row_number,
                cls._filter_retrade_row_by_indices(row_cells, visible_column_indices),
            )
            for row_number, row_cells in visible_rows[1:]
        ]
        return header_cells, data_rows, len(visible_column_indices)

    @staticmethod
    def _warn_if_retrade_calculations_formulas_unresolved(
        file_path: str,
        worksheet_values: Any,
    ) -> None:
        warning_message = (
            "WARNING: Some formula cells returned None. Excel file may need recalculation."
        )
        workbook_formulas = None
        try:
            workbook_formulas = load_workbook(file_path, data_only=False)
            worksheet_title = getattr(worksheet_values, "title", "")
            if worksheet_title in workbook_formulas.sheetnames:
                worksheet_formulas = workbook_formulas[worksheet_title]
            else:
                worksheet_formulas = workbook_formulas.active

            has_unresolved_formula = False
            for formula_row in worksheet_formulas.iter_rows(values_only=False):
                for formula_cell in formula_row:
                    formula_value = formula_cell.value
                    if not (
                        isinstance(formula_value, str)
                        and formula_value.startswith("=")
                    ):
                        continue

                    calculated_value = worksheet_values.cell(
                        row=formula_cell.row,
                        column=formula_cell.column,
                    ).value
                    if calculated_value is None:
                        has_unresolved_formula = True
                        break
                if has_unresolved_formula:
                    break

            if has_unresolved_formula:
                Tool.write_log(warning_message)
        finally:
            try:
                workbook_formulas.close()
            except Exception:
                pass

    def on_sheet_selected(self, sheet_name: str) -> None:
        if not sheet_name:
            return

        workbook = getattr(self, "workbook", None)
        if workbook is None or sheet_name not in workbook.sheetnames:
            return

        sheet = workbook[sheet_name]
        self.current_calculations_sheet_name = sheet_name
        self.display_sheet(sheet)
        self._apply_row_checkbox_mode()
        self._apply_min_margin_highlighting()

    def display_sheet(self, sheet: Any) -> None:
        table = getattr(self, "calculationsTable", None)
        if not isinstance(table, QTableWidget):
            table = getattr(self, "retrade_calculations_table", None)
        if not isinstance(table, QTableWidget):
            return

        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.setUpdatesEnabled(False)
        try:
            table.clear()

            header_cells, visible_data_rows, max_col = (
                self._worksheet_to_visible_retrade_rows(sheet)
            )
            self._retrade_calculations_row_numbers = [
                excel_row for excel_row, _row_cells in visible_data_rows
            ]

            table.setColumnCount(max_col)
            table.setRowCount(len(visible_data_rows))
            if max_col == 0:
                return

            headers: list[str] = []
            for col_index in range(max_col):
                cell_payload = (
                    header_cells[col_index]
                    if col_index < len(header_cells)
                    else {"value": None, "currency": None}
                )
                value = (
                    cell_payload.get("value")
                    if isinstance(cell_payload, dict)
                    else cell_payload
                )
                headers.append("" if value is None else str(value))
            table.setHorizontalHeaderLabels(headers)

            lowered_headers = [header.lower() for header in headers]
            price_columns: set[int] = set()
            rating_columns: set[int] = set()
            for index, header_text in enumerate(lowered_headers):
                if (
                    "руб" in header_text
                    or "итого" in header_text
                    or "цена" in header_text
                ):
                    price_columns.add(index)
                if "рейтинг" in header_text:
                    rating_columns.add(index)

            column_currencies = self._retrade_column_currencies(
                headers,
                [row_cells for _excel_row, row_cells in visible_data_rows],
            )
            for row_index, (excel_row, row_cells) in enumerate(visible_data_rows):
                for col_index in range(max_col):
                    cell_payload = (
                        dict(row_cells[col_index])
                        if col_index < len(row_cells)
                        and isinstance(row_cells[col_index], dict)
                        else {"value": None, "currency": None}
                    )
                    value = cell_payload.get("value")
                    cell_payload["excel_row"] = excel_row
                    text = self._format_calculations_display_value(
                        value,
                        col_index=col_index,
                        header=headers[col_index] if col_index < len(headers) else "",
                        price_columns=price_columns,
                        rating_columns=rating_columns,
                        currency=(
                            cell_payload.get("currency")
                            or (
                                column_currencies[col_index]
                                if col_index < len(column_currencies)
                                else None
                            )
                        ),
                    )
                    if not cell_payload.get("currency") and col_index < len(column_currencies):
                        cell_payload["currency"] = column_currencies[col_index]
                    item = QTableWidgetItem(text)
                    if self._parse_retrade_numeric_value(value) is not None:
                        item.setTextAlignment(
                            int(
                                Qt.AlignmentFlag.AlignRight
                                | Qt.AlignmentFlag.AlignVCenter
                            )
                        )
                    item.setData(Qt.ItemDataRole.UserRole, cell_payload)
                    table.setItem(row_index, col_index, item)
        finally:
            table.setUpdatesEnabled(True)

        resize_table_to_contents(table)
        self._apply_min_margin_highlighting()

    def _open_retrade_calculations(self) -> None:
        current_link = str(
            getattr(self, "current_retrade_calculations_drive_link", "") or ""
        ).strip()
        link, accepted = QInputDialog.getText(
            self,
            "Открыть расчеты",
            "Ссылка на файл расчетов в Google Drive:",
            text=current_link,
        )
        if not accepted:
            return
        link = str(link or "").strip()
        if not link:
            QMessageBox.warning(self, "Ошибка", "Укажите ссылку на расчеты Google Drive")
            return

        try:
            download_result = GoogleDriveService().download_excel(link)
        except Exception as exc:
            error_text = f"Не удалось скачать расчеты с Google Drive: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return

        file_path = str(download_result.local_path)

        workbook = None
        try:
            workbook = self._load_retrade_calculations_workbook(file_path)
            worksheet = workbook.worksheets[0]
            cells_data = self._worksheet_to_retrade_calculations_cells_data(worksheet)
            self._warn_if_retrade_calculations_formulas_unresolved(
                file_path,
                worksheet,
            )
        except Exception as exc:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            error_text = f"Не удалось прочитать Excel файл: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return

        self.current_retrade_calculations_drive_file_id = download_result.file_id
        self.current_retrade_calculations_drive_link = download_result.web_view_link
        self.current_retrade_calculations_drive_name = download_result.name
        self._replace_retrade_calculations_workbook(workbook)
        self.calculations_file_path = file_path
        parsed = self._parse_retrade_calculations(cells_data)
        headers = parsed["headers"]
        rows = parsed["rows"]
        total_without_vat = parsed.get("total_without_vat")
        total_without_vat_currency = parsed.get("total_without_vat_currency")
        totals = parsed.get("totals", {})
        totals_currency = parsed.get("totals_currency", {})
        self._fill_retrade_calculations_view(
            headers,
            rows,
            total_without_vat=total_without_vat,
            total_without_vat_currency=total_without_vat_currency,
            totals=totals,
            totals_currency=totals_currency,
        )
        self._populate_retrade_sheets_list()
        self._open_retrade_calculations_tab()

        self._log_calc("файл загружен")
        self._log_calc(f"google drive: {download_result.web_view_link}")
        self._log_calc(f"заголовков: {len(headers)}")
        self._log_calc(f"строк данных: {len(rows)}")
        self._log_calc(f"сумма товаров: {totals.get('price', 0)}")
        self._log_calc(f"логистика: {totals.get('logistic', 0)}")
        self._log_calc(f"таможня: {totals.get('customs', 0)}")

    def _get_retrade_source_table(self) -> Any:
        table = getattr(self, "retradingTable", None)
        if table is not None:
            return table

        table = getattr(self, "retrade_table", None)
        if table is not None:
            return table

        ui = getattr(self, "ui", None)
        if ui is not None:
            table = getattr(ui, "retradingTable", None)
            if table is not None:
                return table
            return getattr(ui, "retrade_table", None)
        return None

    @staticmethod
    def _text_from_table_item(item: Any) -> str:
        if item is None:
            return ""
        text_getter = getattr(item, "text", None)
        if callable(text_getter):
            return str(text_getter() or "")
        return str(item)

    @classmethod
    def _find_retrade_column_index(
        cls,
        table: Any,
        *,
        fallback: int,
        required_parts: tuple[str, ...],
    ) -> int:
        try:
            column_count = int(table.columnCount())
        except Exception:
            return fallback

        header_getter = getattr(table, "horizontalHeaderItem", None)
        if callable(header_getter):
            for column_index in range(column_count):
                header_text = cls._text_from_table_item(
                    header_getter(column_index)
                ).casefold()
                if all(part in header_text for part in required_parts):
                    return column_index

        return fallback

    @classmethod
    def _find_rating_column_index(cls, table: Any) -> int:
        return cls._find_retrade_column_index(
            table,
            fallback=cls.RATING_COLUMN_INDEX,
            required_parts=("рейтинг",),
        )

    @classmethod
    def _find_best_price_column_index(cls, table: Any) -> int:
        return cls._find_retrade_column_index(
            table,
            fallback=cls.BEST_PRICE_COLUMN_INDEX,
            required_parts=("лучш", "цен"),
        )

    @classmethod
    def _parse_rating_value(cls, value: Any) -> float | None:
        return cls._parse_retrade_numeric_value(value)

    @classmethod
    def _parse_best_price_value(cls, value: Any) -> float | None:
        text = "" if value is None else str(value).replace("\xa0", " ").strip()
        if not text:
            return None

        _, raw_text = Tool.parsePrice(text)
        normalized = (
            str(raw_text or "")
            .replace("\xa0", " ")
            .replace(" ", "")
            .replace(",", ".")
        )
        match = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", normalized)
        if match is None:
            return None

        try:
            return float(match.group(0))
        except Exception:
            return None

    @staticmethod
    def _format_excel_formula_number(value: float) -> str:
        return f"{float(value):.12g}"

    @classmethod
    def _worksheet_header_row_index(cls, worksheet: Any) -> int | None:
        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=False),
            start=1,
        ):
            row_data = [{"value": cell.value, "currency": None} for cell in row]
            if cls._is_retrade_calculations_row_present(row_data):
                return row_number
        return None

    @classmethod
    def _find_worksheet_column_by_header(
        cls,
        worksheet: Any,
        predicate: Any,
    ) -> int | None:
        if worksheet.max_row < 1:
            return None

        header_row_index = cls._worksheet_header_row_index(worksheet)
        if header_row_index is None:
            return None

        for header_cell in worksheet[header_row_index]:
            header_text = cls._normalize_table_header(header_cell.value)
            if predicate(header_text):
                return int(header_cell.column)
        return None

    @staticmethod
    def _enable_workbook_formula_recalculation(workbook: Any) -> None:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    @classmethod
    def _write_realization_price_formulas_to_sheet(
        cls,
        worksheet: Any,
        row_indices: list[int],
        *,
        indices_are_excel_rows: bool = False,
    ) -> dict[int, str]:
        realization_price_col = cls._find_worksheet_column_by_header(
            worksheet,
            lambda header: (
                "реализац" in header
                and "цена" in header
                and "заед" in header
                and "безндс" in header
            ),
        )
        if realization_price_col is None:
            raise ValueError('Не найдена колонка "Цена реализации за ед. без НДС"')

        header_row_index = cls._worksheet_header_row_index(worksheet) or 1
        formulas: dict[int, str] = {}

        for row_index_raw in row_indices:
            try:
                row_index = int(row_index_raw)
            except Exception:
                continue
            if row_index < 0:
                continue

            excel_row = (
                row_index
                if indices_are_excel_rows
                else header_row_index + 1 + row_index
            )
            if excel_row <= header_row_index:
                continue
            formula = f"=ROUND(J{excel_row}*S{excel_row}, 2)"
            target_cell = worksheet.cell(
                row=excel_row,
                column=realization_price_col,
            )
            source_price_cell = worksheet.cell(row=excel_row, column=10)
            if cls._is_zero_retrade_price(source_price_cell.value):
                continue
            currency = (
                cls._detect_currency(target_cell.value, target_cell.number_format)
                or cls._detect_currency(
                    source_price_cell.value,
                    source_price_cell.number_format,
                )
            )
            target_cell.value = formula
            target_cell.number_format = cls._currency_format(
                cls._currency_symbol(currency) or currency
            )
            formulas[row_index] = formula

        return formulas

    @classmethod
    def _extract_retrade_ratings_and_best_prices(
        cls,
        table: Any,
    ) -> tuple[list[float | None], list[float | None]]:
        rating_column_index = cls._find_rating_column_index(table)
        best_price_column_index = cls._find_best_price_column_index(table)
        ratings: list[float | None] = []
        best_prices: list[float | None] = []

        try:
            row_count = int(table.rowCount())
        except Exception:
            return ratings, best_prices

        for row in range(row_count):
            rating_item = table.item(row, rating_column_index)
            best_item = table.item(row, best_price_column_index)

            ratings.append(
                cls._parse_rating_value(cls._text_from_table_item(rating_item))
            )
            best_prices.append(
                cls._parse_best_price_value(cls._text_from_table_item(best_item))
            )

        return ratings, best_prices

    @classmethod
    def _extract_retrade_best_prices(cls, table: Any) -> list[float | None]:
        _, best_prices = cls._extract_retrade_ratings_and_best_prices(table)
        return best_prices

    @staticmethod
    def _next_retrade_sheet_title(workbook: Any) -> str:
        existing_titles = {
            str(title or "").strip().casefold()
            for title in getattr(workbook, "sheetnames", [])
        }
        index = 1
        while True:
            title = f"Переторжка {index}"
            if title.casefold() not in existing_titles:
                return title
            index += 1

    @classmethod
    def _is_zero_retrade_price(cls, value: Any) -> bool:
        numeric_value = cls._parse_retrade_number_or_none(value)
        return numeric_value is not None and abs(numeric_value) <= 1e-12

    @classmethod
    def _is_zero_retrade_price_cell(
        cls,
        worksheet: Any,
        row: int,
        column: int,
        *,
        values_worksheet: Any | None = None,
    ) -> bool:
        if values_worksheet is not None:
            try:
                value = values_worksheet.cell(row=row, column=column).value
            except Exception:
                value = None
            if cls._is_zero_retrade_price(value):
                return True

        try:
            value = worksheet.cell(row=row, column=column).value
        except Exception:
            return False
        return cls._is_zero_retrade_price(value)

    @classmethod
    def _last_retrade_content_column(cls, worksheet: Any) -> int:
        last_column = 0
        for row in worksheet.iter_rows(values_only=False):
            row_values = [
                {
                    "value": cell.value,
                    "currency": cls._detect_currency(cell.value, cell.number_format),
                }
                for cell in row
            ]
            for col_index, cell_data in enumerate(row_values, start=1):
                if cls._is_retrade_calculations_cell_present(cell_data):
                    last_column = max(last_column, col_index)
        return last_column or int(getattr(worksheet, "max_column", 0) or 0)

    @classmethod
    def _write_best_prices_to_calculations_file(
        cls,
        file_path: str,
        best_prices: list[float | None],
        ratings: list[float | None] | None = None,
        *,
        min_margin: float | None = None,
        delta_percent: float | None = None,
    ) -> str:
        workbook = load_workbook(file_path)
        workbook_values = None
        try:
            try:
                workbook_values = load_workbook(file_path, data_only=True, read_only=True)
                sheet_values = workbook_values.worksheets[0]
            except Exception:
                sheet_values = None

            sheet_original = workbook.worksheets[0]
            sheet_copy = workbook.copy_worksheet(sheet_original)
            sheet_copy.title = cls._next_retrade_sheet_title(workbook)

            original_max_col = cls._last_retrade_content_column(sheet_copy)
            real_rating_col = original_max_col + 1
            best_price_col = original_max_col + 2
            formula_col = original_max_col + 3
            corrected_rating_col = original_max_col + 4
            best_price_letter = get_column_letter(best_price_col)
            formula_letter = get_column_letter(formula_col)
            corrected_rating_letter = get_column_letter(corrected_rating_col)
            realization_price_col = cls._find_worksheet_column_by_header(
                sheet_copy,
                lambda header: (
                    "реализац" in header
                    and "цена" in header
                    and "заед" in header
                    and "безндс" in header
                ),
            )
            if realization_price_col is None:
                raise ValueError(
                    'Не найдена колонка "Цена реализации за ед. без НДС"'
                )
            min_margin_value = (
                cls.RETRADE_MIN_MARGIN_DEFAULT if min_margin is None else min_margin
            )
            delta_percent_value = (
                cls.RETRADE_DELTA_PERCENT_DEFAULT
                if delta_percent is None
                else delta_percent
            )
            min_margin_text = cls._format_excel_formula_number(min_margin_value)
            delta_text = cls._format_excel_formula_number(delta_percent_value / 100)
            header_row_index = cls._worksheet_header_row_index(sheet_copy) or 1

            sheet_copy.cell(row=header_row_index, column=real_rating_col).value = (
                "Рейтинг (таблица)"
            )
            sheet_copy.cell(row=header_row_index, column=best_price_col).value = (
                "Лучшая цена за ед."
            )
            sheet_copy.cell(row=header_row_index, column=formula_col).value = "Рейтинг"
            sheet_copy.cell(row=header_row_index, column=corrected_rating_col).value = (
                "Скорректированный рейтинг"
            )
            for column_letter in (
                get_column_letter(formula_col),
                best_price_letter,
                get_column_letter(real_rating_col),
            ):
                sheet_copy.column_dimensions[column_letter].width = 18
            sheet_copy.column_dimensions[
                get_column_letter(corrected_rating_col)
            ].width = 26

            start_row = header_row_index + 1
            rating_values = ratings or []
            for index, best_price in enumerate(best_prices):
                excel_row = start_row + index
                if cls._is_zero_retrade_price_cell(
                    sheet_copy,
                    excel_row,
                    10,
                    values_worksheet=sheet_values,
                ):
                    for column in (
                        real_rating_col,
                        best_price_col,
                        formula_col,
                        corrected_rating_col,
                    ):
                        sheet_copy.cell(row=excel_row, column=column).value = None
                    continue

                rating = rating_values[index] if index < len(rating_values) else None
                if rating is not None:
                    sheet_copy.cell(row=excel_row, column=real_rating_col).value = round(
                        float(rating),
                        2,
                    )

                if best_price is not None:
                    sheet_copy.cell(row=excel_row, column=best_price_col).value = best_price

                formula = f"=ROUND({best_price_letter}{excel_row}/J{excel_row}, 2)"
                sheet_copy.cell(row=excel_row, column=formula_col).value = formula
                corrected_formula = (
                    f"=ROUND(IF({formula_letter}{excel_row}-{delta_text}<"
                    f"{min_margin_text},"
                    f"{min_margin_text},"
                    f"{formula_letter}{excel_row}-{delta_text}), 2)"
                )
                sheet_copy.cell(
                    row=excel_row,
                    column=corrected_rating_col,
                ).value = corrected_formula
                sheet_copy.cell(
                    row=excel_row,
                    column=realization_price_col,
                ).value = (
                    f"=ROUND(J{excel_row}*{corrected_rating_letter}{excel_row}, 2)"
                )

            cls._enable_workbook_formula_recalculation(workbook)
            workbook.save(file_path)
            return sheet_copy.title
        finally:
            if workbook_values is not None:
                workbook_values.close()
            workbook.close()

    def _on_generate_retrade_calculation_clicked(self, *_args: Any) -> None:
        self.generate_retrade_calculation(refresh_from_site=True)

    def _start_retrade_export_for_generation(self) -> bool:
        attached_context = self._get_attached_retrade_export_context()
        if not attached_context:
            Tool.write_log(
                "Формирование расчета без реэкспорта: текущая заявка не закреплена"
            )
            return False

        self._generate_retrade_after_export = True
        try:
            self._start_export_worker(
                trade_id=int(attached_context["trade_id"]),
                lot_id=int(attached_context["lot_id"]),
                bid_id=int(attached_context["bid_id"]),
                is_retrade=True,
                retrade_context=attached_context,
            )
            return True
        except Exception as exc:
            self._generate_retrade_after_export = False
            QMessageBox.warning(self, "Ошибка экспорта заявки", str(exc))
            return True

    def generate_retrade_calculation(self, *, refresh_from_site: bool = False) -> None:
        calculations_file_path = str(
            getattr(self, "calculations_file_path", "") or ""
        ).strip()
        if not calculations_file_path:
            QMessageBox.warning(self, "Ошибка", "Файл расчетов не выбран")
            return

        if refresh_from_site and self._start_retrade_export_for_generation():
            return

        table = self._get_retrade_source_table()
        if table is None:
            QMessageBox.warning(self, "Ошибка", "Таблица Переторжка не найдена")
            return

        selected_rows = self.get_selected_rows()

        try:
            ratings, best_prices = self._extract_retrade_ratings_and_best_prices(table)
            sheet_title = self._write_best_prices_to_calculations_file(
                calculations_file_path,
                best_prices,
                ratings=ratings,
                min_margin=self.get_min_margin(),
                delta_percent=self.get_delta_percent(),
            )
        except Exception as exc:
            error_text = f"Не удалось обновить расчет: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return

        recalc_error = ""
        recalc_skipped = False
        try:
            recalc_skipped = not force_excel_recalc(calculations_file_path)
        except Exception as exc:
            recalc_error = (
                "Формулы записаны, но Excel не удалось автоматически пересчитать: "
                f"{exc}"
            )
            Tool.write_log(recalc_error)

        drive_error = ""
        drive_file_id = str(
            getattr(self, "current_retrade_calculations_drive_file_id", "") or ""
        ).strip()
        if drive_file_id:
            try:
                upload_result = GoogleDriveService().update_excel(
                    drive_file_id,
                    calculations_file_path,
                )
                self.current_retrade_calculations_drive_file_id = upload_result.file_id
                self.current_retrade_calculations_drive_link = upload_result.web_view_link
                self.current_retrade_calculations_drive_name = upload_result.name
                self._set_retrade_calculations_loaded_status(True)
                self._log_calc(
                    f"расчет сохранен на Google Drive: {upload_result.web_view_link}"
                )
            except Exception as exc:
                drive_error = (
                    "Расчет обновлен локально, но не удалось сохранить его "
                    f"на Google Drive: {exc}"
                )
                Tool.write_log(drive_error)

        view_error = ""
        view_reloaded = self._reload_retrade_calculations_view(
            selected_sheet_name=sheet_title,
        )
        if not view_reloaded:
            view_error = (
                "Расчет обновлен, но актуальный лист не удалось открыть "
                "в интерфейсе"
            )
            Tool.write_log(view_error)

        self._log_calc(f"расчет обновлен: {calculations_file_path}")
        self._log_calc(f"лист: {sheet_title}")
        self._log_calc(f"выбрано строк расчетов: {len(selected_rows)}")
        status_bar_getter = getattr(self, "statusBar", None)
        status_bar = status_bar_getter() if callable(status_bar_getter) else None
        if status_bar is not None:
            if drive_error:
                status_message = "Расчет обновлен, но не сохранен на Google Drive"
            elif recalc_error:
                status_message = "Расчет обновлен, но Excel не пересчитан"
            elif recalc_skipped:
                status_message = (
                    "Расчет обновлен, Excel пересчитает формулы при открытии"
                )
            else:
                status_message = "Расчет успешно обновлен и пересчитан"
            if drive_file_id and not drive_error:
                status_message = f"{status_message}; сохранен на Google Drive"
            if view_reloaded:
                status_message = f"{status_message}; открыт лист {sheet_title}"
            status_bar.showMessage(status_message, 5_000)

        warnings = [
            message
            for message in (recalc_error, drive_error, view_error)
            if message
        ]
        if warnings:
            QMessageBox.warning(
                self,
                "Готово с предупреждением",
                "\n\n".join(warnings),
            )
        elif recalc_skipped:
            message = "Расчет обновлен. Формулы пересчитаются при открытии файла в Excel."
            if drive_file_id:
                message = f"{message}\nФайл сохранен на Google Drive."
            if view_reloaded:
                message = f"{message}\nОткрыт лист: {sheet_title}."
            QMessageBox.information(
                self,
                "Готово",
                message,
            )
        else:
            message = "Расчет успешно обновлен и пересчитан"
            if drive_file_id:
                message = f"{message}\nФайл сохранен на Google Drive."
            if view_reloaded:
                message = f"{message}\nОткрыт лист: {sheet_title}."
            QMessageBox.information(
                self,
                "Готово",
                message,
            )

    @staticmethod
    def _load_retrade_calculations_cells_data(file_path: str) -> list[list[dict[str, Any]]]:
        workbook_values = ExportMixin._load_retrade_calculations_workbook(file_path)
        worksheet_values = workbook_values.active

        try:
            data = ExportMixin._worksheet_to_retrade_calculations_cells_data(
                worksheet_values
            )
            ExportMixin._warn_if_retrade_calculations_formulas_unresolved(
                file_path,
                worksheet_values,
            )
            return data
        finally:
            try:
                workbook_values.close()
            except Exception:
                pass

    @staticmethod
    def _detect_currency(value: Any, number_format: Any) -> str | None:
        number_format_code = CurrencyService.normalize_currency_code(number_format)
        if number_format_code:
            return number_format_code

        value_code = CurrencyService.normalize_currency_code(value)
        return value_code or None

    def _clear_retrade_calculations_view(self) -> None:
        sheets_list = getattr(self, "sheetsList", None)
        if isinstance(sheets_list, QListWidget):
            sheets_list.clear()

        table = getattr(self, "retrade_calculations_table", None)
        if isinstance(table, QTableWidget):
            table.clear()
            table.setRowCount(0)
            table.setColumnCount(0)

        for label_name in ("sum_label", "total_label", "profit_label"):
            label = getattr(self, label_name, None)
            if isinstance(label, QLabel):
                label.setText("-")

        self.retrade_calculations_data = {
            "headers": [],
            "rows": [],
            "total_without_vat": None,
            "total_without_vat_currency": None,
            "totals": {
                "price": 0.0,
                "logistic": 0.0,
                "customs": 0.0,
            },
            "totals_currency": {
                "price": None,
                "logistic": None,
                "customs": None,
            },
        }
        total_without_vat_label = getattr(self, "total_without_vat_label", None)
        if isinstance(total_without_vat_label, QLabel):
            total_without_vat_label.setText("")
            total_without_vat_label.setVisible(False)
        price_total_label = getattr(self, "price_total_label", None)
        if isinstance(price_total_label, QLabel):
            price_total_label.setText("")
            price_total_label.setVisible(False)
        self._set_retrade_calculations_loaded_status(False)

    @staticmethod
    def _normalize_retrade_calculations_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        return str(value).strip()

    @classmethod
    def _parse_retrade_numeric_value(cls, value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            candidate = text.replace(" ", "").replace(",", ".")
        else:
            candidate = value
        try:
            return float(candidate)
        except Exception:
            return None

    @staticmethod
    def parse_number(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, str):
            value = (
                value.replace("\xa0", " ")
                .replace(" ", "")
                .replace("₽", "")
                .replace("руб", "")
                .replace("RUB", "")
                .replace("rub", "")
                .replace("USD", "")
                .replace("usd", "")
                .replace("EUR", "")
                .replace("eur", "")
                .replace("CNY", "")
                .replace("cny", "")
                .replace("JPY", "")
                .replace("jpy", "")
                .replace("KZT", "")
                .replace("kzt", "")
                .replace("$", "")
                .replace("€", "")
                .replace("¥", "")
                .replace("₸", "")
                .replace(",", ".")
            )
        try:
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _is_empty_number_value(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            text = (
                value.replace("\xa0", " ")
                .replace(" ", "")
                .replace("₽", "")
                .replace("руб", "")
                .replace("RUB", "")
                .replace("rub", "")
                .replace("USD", "")
                .replace("usd", "")
                .replace("EUR", "")
                .replace("eur", "")
                .replace("CNY", "")
                .replace("cny", "")
                .replace("JPY", "")
                .replace("jpy", "")
                .replace("KZT", "")
                .replace("kzt", "")
                .replace("$", "")
                .replace("€", "")
                .replace("¥", "")
                .replace("₸", "")
                .strip()
            )
            return not text
        return False

    @staticmethod
    def _normalize_table_header(value: Any) -> str:
        text = str(value or "").strip().casefold().replace("ё", "е")
        return re.sub(r"[^a-zа-я0-9]+", "", text)

    @staticmethod
    def _is_numeric_table_value(value: Any) -> bool:
        if value is None or isinstance(value, bool):
            return False
        if not isinstance(value, (int, float)):
            return False
        try:
            return not pd.isna(value)
        except Exception:
            return True

    @classmethod
    def _format_number_ru(cls, value: Any) -> str:
        numeric_value = cls._parse_retrade_numeric_value(value)
        if numeric_value is None:
            return cls._normalize_retrade_calculations_cell(value)

        formatted = f"{numeric_value:,.2f}"
        return formatted.replace(",", " ").replace(".", ",")

    @classmethod
    def _format_retrade_calculations_cell_for_display(cls, cell: Any) -> str:
        if isinstance(cell, dict):
            raw_value = cell.get("value")
            currency = cell.get("currency")
        else:
            raw_value = cell
            currency = None

        formatted_value = cls._format_number_ru(raw_value)
        if not formatted_value:
            return ""

        if cls._parse_retrade_numeric_value(raw_value) is None:
            return formatted_value

        currency_symbol = cls._currency_display_symbol(currency)
        if currency_symbol:
            return f"{formatted_value} {currency_symbol}"
        return formatted_value

    @classmethod
    def _is_retrade_calculations_value_present(cls, value: Any) -> bool:
        return bool(cls._normalize_retrade_calculations_cell(value))

    @classmethod
    def _is_retrade_calculations_cell_present(cls, cell_data: Any) -> bool:
        if isinstance(cell_data, dict):
            value = cell_data.get("value")
            return cls._is_retrade_calculations_value_present(value)
        return cls._is_retrade_calculations_value_present(cell_data)

    @classmethod
    def _is_retrade_calculations_row_present(cls, row_data: Any) -> bool:
        return any(
            cls._is_retrade_calculations_cell_present(cell)
            for cell in row_data or []
        )

    @classmethod
    def _non_empty_retrade_column_indices(
        cls,
        rows: list[list[Any]],
    ) -> list[int]:
        max_col = max((len(row or []) for row in rows or []), default=0)
        indices: list[int] = []
        for col_index in range(max_col):
            if any(
                col_index < len(row)
                and cls._is_retrade_calculations_cell_present(row[col_index])
                for row in rows or []
            ):
                indices.append(col_index)
        return indices

    @staticmethod
    def _filter_retrade_row_by_indices(
        row: list[dict[str, Any]],
        indices: list[int],
    ) -> list[dict[str, Any]]:
        empty_cell = {"value": None, "currency": None}
        return [
            row[index] if index < len(row) else dict(empty_cell)
            for index in indices
        ]

    @classmethod
    def _cell_payload_currency_or_none(cls, cell: Any) -> str | None:
        if isinstance(cell, dict):
            explicit_currency = CurrencyService.normalize_currency_code(
                cell.get("currency")
            )
            if explicit_currency:
                return explicit_currency
            return cls._detect_currency(cell.get("value"), None)
        return cls._detect_currency(cell, None)

    @classmethod
    def _column_currency_from_cells(
        cls,
        header: Any,
        rows: list[list[Any]],
        col_index: int,
    ) -> str | None:
        header_currency = cls._detect_currency(header, None)
        if header_currency:
            return header_currency

        for row in rows:
            if col_index >= len(row):
                continue
            cell_currency = cls._cell_payload_currency_or_none(row[col_index])
            if cell_currency:
                return cell_currency
        return None

    @classmethod
    def _retrade_column_currencies(
        cls,
        headers: list[Any],
        rows: list[list[Any]],
    ) -> list[str | None]:
        column_count = max(
            [len(headers), *(len(row or []) for row in rows or [])],
            default=0,
        )
        return [
            cls._column_currency_from_cells(
                headers[index] if index < len(headers) else "",
                rows,
                index,
            )
            for index in range(column_count)
        ]

    @classmethod
    def _apply_retrade_column_currencies(
        cls,
        headers: list[Any],
        rows: list[list[dict[str, Any]]],
    ) -> None:
        column_currencies = cls._retrade_column_currencies(headers, rows)
        for row in rows:
            for col_index, currency in enumerate(column_currencies):
                if not currency or col_index >= len(row):
                    continue
                cell = row[col_index]
                if not isinstance(cell, dict):
                    continue
                if not CurrencyService.normalize_currency_code(cell.get("currency")):
                    cell["currency"] = currency

    @classmethod
    def _is_retrade_position_cell(cls, value: Any) -> bool:
        if value is None or isinstance(value, bool):
            return False

        if isinstance(value, (int, float)):
            try:
                if pd.isna(value):
                    return False
            except Exception:
                pass
            return True

        if isinstance(value, str):
            return value.strip().isdigit()

        return False

    @staticmethod
    def _is_service_retrade_column_header(header: str) -> bool:
        lowered = header.lower()
        return "сумма" in lowered or "итого" in lowered or "прибыль" in lowered

    @staticmethod
    def _is_price_per_unit_header(header: Any) -> bool:
        return isinstance(header, str) and "цена за ед" in header.lower()

    @staticmethod
    def _find_retrade_column(headers: list[Any], keywords: list[str]) -> int | None:
        for index, header in enumerate(headers):
            if not isinstance(header, str):
                continue
            lowered_header = header.lower()
            if any(keyword in lowered_header for keyword in keywords):
                return index
        return None

    @classmethod
    def _find_update_positions_column(
        cls,
        headers: list[Any],
        kind: str,
    ) -> int | None:
        for index, header in enumerate(headers):
            normalized = cls._normalize_table_header(header)
            if not normalized:
                continue

            if kind == "source_price":
                if (
                    "ценазаедбезндс" in normalized
                    or (
                        "цена" in normalized
                        and "заед" in normalized
                        and "безндс" in normalized
                        and "реализац" not in normalized
                        and "предлага" not in normalized
                    )
                ):
                    return index
            elif kind == "corrected_rating":
                if "скоррект" in normalized and "рейтинг" in normalized:
                    return index
            elif kind == "sale_price":
                if "реализац" in normalized and "цена" in normalized and "заед" in normalized:
                    return index
            elif kind == "proposal_price":
                if "предлага" in normalized and "цена" in normalized and "заед" in normalized:
                    return index
        return None

    @classmethod
    def _get_update_positions_columns(
        cls,
        headers: list[Any],
        rows: list[list[Any]] | None = None,
        *,
        column_offset: int = 0,
    ) -> dict[str, int | None]:
        row_width = max((len(row or []) for row in (rows or [])), default=0)
        column_count = max(len(headers), row_width)
        source_price_col = (
            column_offset + 9 if column_offset + 9 < column_count else None
        )
        corrected_rating_col = (
            column_offset + 18 if column_offset + 18 < column_count else None
        )
        return {
            "source_price": (
                source_price_col
                if source_price_col is not None
                else cls._find_update_positions_column(headers, "source_price")
            ),
            "corrected_rating": (
                corrected_rating_col
                if corrected_rating_col is not None
                else cls._find_update_positions_column(headers, "corrected_rating")
            ),
            "sale_price": cls._find_update_positions_column(headers, "sale_price"),
        }

    @staticmethod
    def _cell_payload_value(cell: Any) -> Any:
        if isinstance(cell, dict):
            return cell.get("value")
        return cell

    @classmethod
    def _cell_payload_currency(cls, cell: Any) -> str | None:
        if isinstance(cell, dict):
            currency = cell.get("currency")
            if currency:
                return currency
            return cls._detect_currency(cell.get("value"), None)
        return cls._detect_currency(cell, None)

    @classmethod
    def _calculate_updated_position_prices(
        cls,
        headers: list[Any],
        rows: list[list[Any]],
        *,
        column_offset: int = 0,
    ) -> tuple[list[dict[str, Any]], int]:
        columns = cls._get_update_positions_columns(
            headers,
            rows,
            column_offset=column_offset,
        )
        missing = []
        if columns["source_price"] is None:
            missing.append("Цена за ед. без НДС")
        if columns["corrected_rating"] is None:
            missing.append("Скорректированный рейтинг")
        if columns["sale_price"] is None:
            missing.append("Цена реализации за ед. без НДС")
        if missing:
            raise ValueError("Не найдены колонки: " + ", ".join(missing))

        source_price_col = int(columns["source_price"])
        corrected_rating_col = int(columns["corrected_rating"])
        sale_price_col = int(columns["sale_price"])
        updates: list[dict[str, Any]] = []

        for row_index, row in enumerate(rows):
            row_values = list(row or [])
            price_cell = (
                row_values[source_price_col]
                if source_price_col < len(row_values)
                else None
            )
            coef_cell = (
                row_values[corrected_rating_col]
                if corrected_rating_col < len(row_values)
                else None
            )
            target_cell = (
                row_values[sale_price_col]
                if sale_price_col < len(row_values)
                else None
            )
            price_raw = cls._cell_payload_value(price_cell)
            coef_raw = cls._cell_payload_value(coef_cell)

            if cls._is_empty_number_value(price_raw) or cls._is_empty_number_value(coef_raw):
                continue
            if cls._is_zero_retrade_price(price_raw):
                continue

            new_price = round(cls.parse_number(price_raw) * cls.parse_number(coef_raw), 2)
            updates.append(
                {
                    "row": row_index,
                    "value": new_price,
                    "currency": (
                        cls._cell_payload_currency(price_cell)
                        or cls._cell_payload_currency(target_cell)
                        or cls._cell_payload_currency(coef_cell)
                    ),
                }
            )

        return updates, sale_price_col

    @classmethod
    def _formula_update_row_indices(
        cls,
        headers: list[Any],
        rows: list[list[Any]],
        *,
        column_offset: int = 0,
    ) -> list[int]:
        columns = cls._get_update_positions_columns(
            headers,
            rows,
            column_offset=column_offset,
        )
        if columns["source_price"] is None:
            return []

        source_price_col = int(columns["source_price"])
        row_indices: list[int] = []
        for row_index, row in enumerate(rows):
            row_values = list(row or [])
            price_cell = (
                row_values[source_price_col]
                if source_price_col < len(row_values)
                else None
            )
            price_raw = cls._cell_payload_value(price_cell)
            if (
                not cls._is_empty_number_value(price_raw)
                and not cls._is_zero_retrade_price(price_raw)
            ):
                row_indices.append(row_index)
        return row_indices

    @classmethod
    def _sum_retrade_column(
        cls,
        rows: list[list[dict[str, Any]]],
        col_index: int | None,
    ) -> tuple[float, str | None]:
        if col_index is None:
            return 0.0, None

        total = 0.0
        detected_currency: str | None = None
        for row in rows:
            if col_index >= len(row):
                continue
            cell = row[col_index]
            if isinstance(cell, dict):
                value = cell.get("value")
                if detected_currency is None and cell.get("currency"):
                    detected_currency = cell.get("currency")
            else:
                value = cell

            numeric_value = cls._parse_retrade_numeric_value(value)
            if numeric_value is None:
                continue
            total += numeric_value

        return total, detected_currency

    @staticmethod
    def _is_total_without_vat_marker(value: Any) -> bool:
        if not isinstance(value, str):
            return False
        return "итого без ндс" in value.lower()

    @classmethod
    def _extract_total_without_vat_from_row(
        cls,
        row_cells: list[dict[str, Any]],
    ) -> tuple[Any, str | None]:
        numeric_candidates: list[tuple[Any, str | None]] = []
        for cell in row_cells:
            if not isinstance(cell, dict):
                continue
            value = cell.get("value")
            numeric = cls._parse_retrade_numeric_value(value)
            if numeric is None:
                continue
            numeric_candidates.append((value, cell.get("currency")))

        if not numeric_candidates:
            return None, None

        selected_value, selected_currency = numeric_candidates[-1]
        return selected_value, selected_currency

    @staticmethod
    def _log_calc(message: str) -> None:
        text = f"[CALC] {message}"
        Tool.write_log(text)

    @classmethod
    def _parse_retrade_calculations(
        cls,
        cells_data: list[list[dict[str, Any]]],
    ) -> dict[str, Any]:
        headers: list[str] = []
        position_rows: list[list[dict[str, Any]]] = []
        total_without_vat: Any = None
        total_without_vat_currency: str | None = None
        header_found = False

        for raw_row in cells_data:
            row_data = list(raw_row or [])
            if not cls._is_retrade_calculations_row_present(row_data):
                continue

            if not header_found:
                headers = [
                    cls._normalize_retrade_calculations_cell(
                        cell.get("value") if isinstance(cell, dict) else cell
                    )
                    for cell in row_data
                ]
                while headers and not headers[-1]:
                    headers.pop()
                header_found = True
                continue

            normalized_row: list[dict[str, Any]] = []
            for cell in row_data:
                if isinstance(cell, dict):
                    normalized_row.append(
                        {
                            "value": cell.get("value"),
                            "currency": cell.get("currency"),
                        }
                    )
                else:
                    normalized_row.append({"value": cell, "currency": None})

            while normalized_row and not cls._is_retrade_calculations_cell_present(normalized_row[-1]):
                normalized_row.pop()
            if not normalized_row:
                continue

            if any(
                cls._is_total_without_vat_marker(cell.get("value"))
                for cell in normalized_row
                if isinstance(cell, dict)
            ):
                extracted_value, extracted_currency = cls._extract_total_without_vat_from_row(
                    normalized_row
                )
                if extracted_value is not None:
                    total_without_vat = extracted_value
                    total_without_vat_currency = extracted_currency
                continue

            first_cell_value = (
                normalized_row[0].get("value")
                if normalized_row and isinstance(normalized_row[0], dict)
                else None
            )
            if not cls._is_retrade_position_cell(first_cell_value):
                continue

            position_rows.append(normalized_row)

        filtered_headers: list[str] = []
        filtered_indices: list[int] = []
        max_filtered_width = max(
            [len(headers), *(len(row or []) for row in position_rows)],
            default=0,
        )
        for index in range(max_filtered_width):
            header = headers[index] if index < len(headers) else ""
            header = "" if header is None else str(header)
            if cls._is_service_retrade_column_header(header):
                continue
            column_has_values = bool(header) or any(
                index < len(row)
                and cls._is_retrade_calculations_cell_present(row[index])
                for row in position_rows
            )
            if not column_has_values:
                continue
            filtered_headers.append(header)
            filtered_indices.append(index)

        filtered_rows: list[list[dict[str, Any]]] = []
        for row in position_rows:
            new_row: list[dict[str, Any]] = []
            for index in filtered_indices:
                if index < len(row):
                    new_row.append(row[index])
                else:
                    new_row.append({"value": None, "currency": None})

            while new_row and not cls._is_retrade_calculations_cell_present(new_row[-1]):
                new_row.pop()
            if not new_row:
                continue
            filtered_rows.append(new_row)

        cls._apply_retrade_column_currencies(filtered_headers, filtered_rows)

        price_col_index = cls._find_retrade_column(filtered_headers, ["цена за ед"])
        logistic_col_index = cls._find_retrade_column(
            filtered_headers,
            ["логист", "доставк", "транспорт"],
        )
        customs_col_index = cls._find_retrade_column(filtered_headers, ["тамож", "пошлин"])

        if price_col_index is None:
            cls._log_calc("Ошибка: столбец 'Цена за ед. без НДС' не найден")
        if logistic_col_index is None:
            cls._log_calc("Ошибка: столбец 'Логистика' не найден")
        if customs_col_index is None:
            cls._log_calc("Ошибка: столбец 'Таможня' не найден")

        price_total, price_total_currency = cls._sum_retrade_column(filtered_rows, price_col_index)
        logistic_total, logistic_total_currency = cls._sum_retrade_column(
            filtered_rows,
            logistic_col_index,
        )
        customs_total, customs_total_currency = cls._sum_retrade_column(
            filtered_rows,
            customs_col_index,
        )

        return {
            "headers": filtered_headers,
            "rows": filtered_rows,
            "total_without_vat": total_without_vat,
            "total_without_vat_currency": total_without_vat_currency,
            "totals": {
                "price": price_total,
                "logistic": logistic_total,
                "customs": customs_total,
            },
            "totals_currency": {
                "price": price_total_currency,
                "logistic": logistic_total_currency,
                "customs": customs_total_currency,
            },
        }

    def _fill_retrade_calculations_view(
        self,
        headers: list[str],
        rows: list[list[dict[str, Any]]],
        *,
        total_without_vat: Any = None,
        total_without_vat_currency: str | None = None,
        totals: dict[str, float] | None = None,
        totals_currency: dict[str, str | None] | None = None,
    ) -> None:
        self._clear_retrade_calculations_view()
        self._apply_retrade_column_currencies(headers, rows)

        table = getattr(self, "retrade_calculations_table", None)
        if isinstance(table, QTableWidget):
            self._configure_excel_like_table(table)
            table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            table.setAlternatingRowColors(True)
            column_count = max(len(headers), max((len(row) for row in rows), default=0))
            table.setRowCount(len(rows))
            table.setColumnCount(column_count)
            if column_count > 0:
                header_labels = list(headers[:column_count])
                if len(header_labels) < column_count:
                    header_labels.extend(
                        f"Колонка {index + 1}"
                        for index in range(len(header_labels), column_count)
                    )
                table.setHorizontalHeaderLabels(header_labels)
            for row_index, row_values in enumerate(rows):
                for col_index in range(column_count):
                    cell_payload = (
                        row_values[col_index]
                        if col_index < len(row_values)
                        else {"value": None, "currency": None}
                    )
                    if isinstance(cell_payload, dict):
                        currency = cell_payload.get("currency")
                    else:
                        currency = None

                    display_value = self._format_retrade_calculations_cell_for_display(cell_payload)
                    item = QTableWidgetItem(display_value)
                    item.setData(Qt.ItemDataRole.UserRole, cell_payload)
                    raw_value = cell_payload.get("value") if isinstance(cell_payload, dict) else cell_payload
                    if self._is_numeric_table_value(raw_value):
                        item.setTextAlignment(
                            int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        )
                    else:
                        item.setTextAlignment(
                            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        )
                    if currency:
                        item.setToolTip(f"Валюта: {currency}")
                    table.setItem(row_index, col_index, item)
            resize_table_to_contents(table)
            self._apply_min_margin_highlighting()

        totals_data = totals or {}
        totals_currency_data = totals_currency or {}
        default_currency = total_without_vat_currency
        price_currency = totals_currency_data.get("price") or default_currency
        logistic_currency = totals_currency_data.get("logistic") or default_currency
        customs_currency = totals_currency_data.get("customs") or default_currency

        price_display = self._format_retrade_calculations_cell_for_display(
            {
                "value": totals_data.get("price", 0.0),
                "currency": price_currency,
            }
        )
        logistic_display = self._format_retrade_calculations_cell_for_display(
            {
                "value": totals_data.get("logistic", 0.0),
                "currency": logistic_currency,
            }
        )
        customs_display = self._format_retrade_calculations_cell_for_display(
            {
                "value": totals_data.get("customs", 0.0),
                "currency": customs_currency,
            }
        )

        sum_label = getattr(self, "sum_label", None)
        if isinstance(sum_label, QLabel):
            sum_label.setText(price_display or "0,00")
        total_label = getattr(self, "total_label", None)
        if isinstance(total_label, QLabel):
            total_label.setText(logistic_display or "0,00")
        profit_label = getattr(self, "profit_label", None)
        if isinstance(profit_label, QLabel):
            profit_label.setText(customs_display or "0,00")

        total_without_vat_label = getattr(self, "total_without_vat_label", None)
        if isinstance(total_without_vat_label, QLabel):
            if total_without_vat is None:
                total_without_vat_label.setText("")
                total_without_vat_label.setVisible(False)
            else:
                formatted_total = self._format_retrade_calculations_cell_for_display(
                    {
                        "value": total_without_vat,
                        "currency": total_without_vat_currency,
                    }
                )
                if formatted_total:
                    total_without_vat_label.setText(f"Итого без НДС: {formatted_total}")
                    total_without_vat_label.setVisible(True)
                else:
                    total_without_vat_label.setText("")
                    total_without_vat_label.setVisible(False)

        price_total_label = getattr(self, "price_total_label", None)
        if isinstance(price_total_label, QLabel):
            price_total_label.setText("")
            price_total_label.setVisible(False)

        self.retrade_calculations_data = {
            "headers": list(headers),
            "rows": [list(row) for row in rows],
            "total_without_vat": total_without_vat,
            "total_without_vat_currency": total_without_vat_currency,
            "totals": {
                "price": totals_data.get("price", 0.0),
                "logistic": totals_data.get("logistic", 0.0),
                "customs": totals_data.get("customs", 0.0),
            },
            "totals_currency": {
                "price": price_currency,
                "logistic": logistic_currency,
                "customs": customs_currency,
            },
        }
        self._set_retrade_calculations_loaded_status(bool(headers or rows))

    @classmethod
    def _table_headers(cls, table: Any) -> list[str]:
        try:
            column_count = int(table.columnCount())
        except Exception:
            return []

        headers: list[str] = []
        header_getter = getattr(table, "horizontalHeaderItem", None)
        for column_index in range(column_count):
            header_item = header_getter(column_index) if callable(header_getter) else None
            headers.append(cls._text_from_table_item(header_item))
        return headers

    @staticmethod
    def _table_item_payload_value(item: Any) -> Any:
        if item is None:
            return None

        data_getter = getattr(item, "data", None)
        if callable(data_getter):
            try:
                payload = data_getter(Qt.ItemDataRole.UserRole)
            except Exception:
                payload = None
            if isinstance(payload, dict):
                return payload.get("value")
            if payload is not None:
                return payload

        text_getter = getattr(item, "text", None)
        if callable(text_getter):
            return text_getter()
        return None

    @staticmethod
    def _table_item_payload_currency(item: Any) -> str | None:
        if item is None:
            return None
        data_getter = getattr(item, "data", None)
        if not callable(data_getter):
            return None
        try:
            payload = data_getter(Qt.ItemDataRole.UserRole)
        except Exception:
            return None
        if isinstance(payload, dict):
            return payload.get("currency")
        return None

    @staticmethod
    def _table_item_payload_excel_row(item: Any) -> int | None:
        if item is None:
            return None
        data_getter = getattr(item, "data", None)
        if not callable(data_getter):
            return None
        try:
            payload = data_getter(Qt.ItemDataRole.UserRole)
        except Exception:
            return None
        if not isinstance(payload, dict):
            return None
        try:
            excel_row = int(payload.get("excel_row"))
        except (TypeError, ValueError):
            return None
        return excel_row if excel_row > 0 else None

    @classmethod
    def _table_payload_rows(cls, table: Any) -> list[list[dict[str, Any]]]:
        try:
            row_count = int(table.rowCount())
            column_count = int(table.columnCount())
        except Exception:
            return []

        rows: list[list[dict[str, Any]]] = []
        for row_index in range(row_count):
            row: list[dict[str, Any]] = []
            for column_index in range(column_count):
                item = table.item(row_index, column_index)
                cell_payload = {
                    "value": cls._table_item_payload_value(item),
                    "currency": cls._table_item_payload_currency(item),
                }
                excel_row = cls._table_item_payload_excel_row(item)
                if excel_row is not None:
                    cell_payload["excel_row"] = excel_row
                row.append(cell_payload)
            rows.append(row)
        return rows

    def _table_rows_to_excel_rows(
        self,
        table: Any,
        row_indices: list[int],
    ) -> dict[int, int]:
        try:
            column_count = int(table.columnCount())
        except Exception:
            column_count = 0

        row_numbers = getattr(self, "_retrade_calculations_row_numbers", None)
        result: dict[int, int] = {}
        for row_index in row_indices:
            try:
                normalized_row = int(row_index)
            except (TypeError, ValueError):
                continue
            if normalized_row < 0:
                continue

            excel_row: int | None = None
            for column_index in range(column_count):
                item = table.item(normalized_row, column_index)
                excel_row = self._table_item_payload_excel_row(item)
                if excel_row is not None:
                    break

            if excel_row is None and isinstance(row_numbers, list):
                if normalized_row < len(row_numbers):
                    try:
                        excel_row = int(row_numbers[normalized_row])
                    except (TypeError, ValueError):
                        excel_row = None

            if excel_row is not None and excel_row > 0:
                result[normalized_row] = excel_row
        return result

    def _set_table_numeric_item(
        self,
        table: Any,
        row: int,
        column: int,
        value: float,
        *,
        currency: str | None = None,
        show_currency: bool = False,
        formula: str | None = None,
    ) -> None:
        item = table.item(row, column)
        payload = {"value": value, "currency": currency}
        excel_row = self._table_item_payload_excel_row(item)
        if excel_row is not None:
            payload["excel_row"] = excel_row
        if formula:
            payload["formula"] = formula
        if show_currency:
            text = self._format_retrade_calculations_cell_for_display(payload)
        else:
            text = self._format_number_ru(value)

        if item is None:
            item = QTableWidgetItem(text)
            table.setItem(row, column, item)
        else:
            item.setText(text)

        item.setData(Qt.ItemDataRole.UserRole, payload)
        item.setTextAlignment(
            int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        )
        tooltip_parts = []
        if currency:
            tooltip_parts.append(f"Валюта: {currency}")
        if formula:
            tooltip_parts.append(f"Формула: {formula}")
        item.setToolTip("\n".join(tooltip_parts))

    def _update_retrade_calculations_data_price(
        self,
        row: int,
        column: int,
        value: float,
        currency: str | None,
        formula: str | None = None,
    ) -> None:
        data = getattr(self, "retrade_calculations_data", None)
        if not isinstance(data, dict):
            return
        rows = data.get("rows")
        if not isinstance(rows, list) or row >= len(rows):
            return
        row_values = rows[row]
        if not isinstance(row_values, list):
            return
        while len(row_values) <= column:
            row_values.append({"value": None, "currency": None})
        cell = row_values[column]
        if not isinstance(cell, dict):
            row_values[column] = {"value": value, "currency": currency}
            return
        cell["value"] = value
        if currency:
            cell["currency"] = currency
        if formula:
            cell["formula"] = formula

    def _write_update_position_formulas_to_current_calculations_file(
        self,
        row_indices: list[int],
        *,
        indices_are_excel_rows: bool = False,
    ) -> dict[int, str]:
        file_path = str(getattr(self, "calculations_file_path", "") or "").strip()
        if not file_path:
            raise ValueError("Файл расчетов не выбран")

        workbook = load_workbook(file_path, data_only=False)
        try:
            sheet_name = str(
                getattr(self, "current_calculations_sheet_name", "") or ""
            )
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.worksheets[0]

            formulas = self._write_realization_price_formulas_to_sheet(
                worksheet,
                row_indices,
                indices_are_excel_rows=indices_are_excel_rows,
            )
            self._enable_workbook_formula_recalculation(workbook)
            workbook.save(file_path)
            return formulas
        finally:
            workbook.close()

    def update_retrade_positions(self) -> None:
        calculations_table = getattr(self, "retrade_calculations_table", None)
        if not isinstance(calculations_table, QTableWidget):
            QMessageBox.warning(self, "Ошибка", "Таблица расчетов не найдена")
            return

        main_table = self._get_retrade_source_table()
        if not isinstance(main_table, QTableWidget):
            QMessageBox.warning(self, "Ошибка", "Основная таблица не найдена")
            return

        calculations_headers = self._table_headers(calculations_table)
        calculations_rows = self._table_payload_rows(calculations_table)
        column_offset = self.get_calculations_column_offset()
        main_headers = self._table_headers(main_table)
        main_price_col = self._find_update_positions_column(
            main_headers,
            "proposal_price",
        )
        if main_price_col is None:
            QMessageBox.warning(
                self,
                "Ошибка",
                'Не найдена колонка "Предлагаемая цена за ед." в основной таблице',
            )
            return

        try:
            updates, sale_price_col = self._calculate_updated_position_prices(
                calculations_headers,
                calculations_rows,
                column_offset=column_offset,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Ошибка", str(exc))
            return

        formula_row_indices = self._formula_update_row_indices(
            calculations_headers,
            calculations_rows,
            column_offset=column_offset,
        )
        if not formula_row_indices:
            QMessageBox.information(
                self,
                "Обновить цены",
                "Нет строк с заполненной исходной ценой",
            )
            return

        excel_rows_by_table_row = self._table_rows_to_excel_rows(
            calculations_table,
            formula_row_indices,
        )
        use_excel_rows = len(excel_rows_by_table_row) == len(formula_row_indices)
        formula_targets = (
            [excel_rows_by_table_row[row_index] for row_index in formula_row_indices]
            if use_excel_rows
            else formula_row_indices
        )

        try:
            formulas_by_row = (
                self._write_update_position_formulas_to_current_calculations_file(
                    formula_targets,
                    indices_are_excel_rows=use_excel_rows,
                )
            )
        except Exception as exc:
            error_text = f"Не удалось обновить формулы в файле расчетов: {exc}"
            Tool.write_log(error_text)
            QMessageBox.warning(self, "Ошибка", error_text)
            return

        updated_count = 0
        main_row_count = main_table.rowCount()
        for update in updates:
            row_index = int(update["row"])
            if row_index >= main_row_count:
                continue

            new_price = float(update["value"])
            currency = update.get("currency")
            formula_key = (
                excel_rows_by_table_row.get(row_index)
                if use_excel_rows
                else row_index
            )
            formula = formulas_by_row.get(formula_key)
            self._set_table_numeric_item(
                calculations_table,
                row_index,
                sale_price_col,
                new_price,
                currency=currency,
                show_currency=True,
                formula=formula,
            )
            self._set_table_numeric_item(
                main_table,
                row_index,
                int(main_price_col),
                new_price,
            )
            if isinstance(main_table, QTableWidget):
                self._recalculate_retrade_main_table_row(main_table, row_index)
            self._update_retrade_calculations_data_price(
                row_index,
                sale_price_col,
                new_price,
                currency,
                formula=formula,
            )
            updated_count += 1

        resize_table_to_contents(calculations_table)
        resize_table_to_contents(main_table)

        status_bar_getter = getattr(self, "statusBar", None)
        status_bar = status_bar_getter() if callable(status_bar_getter) else None
        if status_bar is not None:
            status_bar.showMessage(
                (
                    f"Обновлено цен: {updated_count}; "
                    f"формулы сохранены: {len(formulas_by_row)}"
                ),
                5_000,
            )
        if updated_count == 0:
            QMessageBox.information(
                self,
                "Обновить цены",
                (
                    "Формулы сохранены в Excel. Для отображения чисел в таблицах "
                    "нужны заполненные цена и скорректированный рейтинг."
                ),
            )

    def _set_retrade_calculations_loaded_status(self, is_loaded: bool) -> None:
        self.retrade_calculations_loaded = bool(is_loaded)
        status_label = getattr(self, "label_retrade_calculations_status", None)
        if not isinstance(status_label, QLabel):
            return

        if self.retrade_calculations_loaded:
            drive_name = str(
                getattr(self, "current_retrade_calculations_drive_name", "") or ""
            ).strip()
            drive_link = str(
                getattr(self, "current_retrade_calculations_drive_link", "") or ""
            ).strip()
            if drive_name:
                status_label.setText(f"Расчеты подключены: {drive_name}")
            else:
                status_label.setText("Расчеты подключены")
            status_label.setToolTip(drive_link)
            status_label.setStyleSheet("color: #1f8f3a; font-weight: 600;")
            return

        status_label.setText("Расчеты не подключены")
        status_label.setToolTip("")
        status_label.setStyleSheet("color: #c62828; font-weight: 600;")

    def _toggle_auto_trade_status(self) -> None:
        status_label = getattr(self, "label_auto_trade_status", None)
        if status_label is None:
            return
        is_enabled = str(status_label.text() or "").strip() == "Включено"
        if is_enabled:
            self._stop_auto_trade_timer()
            self._set_auto_trade_status(False)
            self._log_ui("Автоматическое ведение торгов: Выключено")
            return

        current_bid_id = self._get_current_retrade_bid_id()
        if current_bid_id is None:
            QMessageBox.warning(
                self,
                "Автоматическое ведение торгов",
                "Сначала экспортируйте переторжку, чтобы закрепить текущую заявку.",
            )
            return

        if not self._confirm_auto_trade_enable_if_needed():
            return

        self._set_auto_trade_status(True)
        self._log_ui("Автоматическое ведение торгов: Включено")
        current_number = str(getattr(self, "current_retrade", "") or "").strip()
        if current_number:
            self._log_auto_trade(
                f"Текущая переторжка: заявка {current_number}, bid_id={current_bid_id}"
            )
        self._start_auto_trade_timer_if_needed()

    def _set_auto_trade_status(self, is_enabled: bool) -> None:
        status_label = getattr(self, "label_auto_trade_status", None)
        if status_label is None:
            return
        if is_enabled:
            status_label.setText("Включено")
            status_label.setStyleSheet("color: #1f8f3a; font-weight: 600;")
            return
        status_label.setText("Выключено")
        status_label.setStyleSheet("color: #c62828; font-weight: 600;")

    def _confirm_auto_trade_enable(self) -> bool:
        dialog = QMessageBox(self)
        dialog.setIcon(QMessageBox.Icon.Warning)
        dialog.setWindowTitle("Подтверждение включения")
        dialog.setText(
            "Вы собираетесь включить автоматическое ведение торгов.\n\n"
            "ВНИМАНИЕ:\n"
            "Данная функция работает в автоматическом режиме и может отправлять ценовые "
            "предложения без дополнительного подтверждения.\n\n"
            "Использование функции требует ОБЯЗАТЕЛЬНОГО контроля со стороны пользователя.\n\n"
            "Рекомендуется не оставлять систему без присмотра во время работы.\n\n"
            "Продолжить?"
        )
        yes_button = dialog.addButton("Да", QMessageBox.ButtonRole.YesRole)
        cancel_button = dialog.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
        dialog.setDefaultButton(cancel_button)
        dialog.exec()
        return dialog.clickedButton() is yes_button

    def _confirm_auto_trade_enable_if_needed(self) -> bool:
        if bool(Config.settings.get("skip_auto_trade_warning", False)):
            return True
        return self._confirm_auto_trade_enable()

    def _ensure_auto_trade_timer(self) -> QTimer:
        timer = getattr(self, "_auto_trade_timer", None)
        if isinstance(timer, QTimer):
            return timer

        timer = QTimer(self)
        timer.setSingleShot(True)
        timer.timeout.connect(self._on_auto_trade_timer_timeout)
        self._auto_trade_timer = timer
        return timer

    def _stop_auto_trade_timer(self) -> None:
        timer = self._ensure_auto_trade_timer()
        if timer.isActive():
            timer.stop()

    def _start_auto_trade_timer_if_needed(self) -> None:
        self._stop_auto_trade_timer()
        if not bool(Config.settings.get("use_auto_trade_timer", False)):
            return

        default_minutes = int(Config.DEFAULT_SETTINGS.get("auto_trade_timer_minutes", 30))
        raw_minutes = Config.settings.get("auto_trade_timer_minutes", default_minutes)
        try:
            minutes = int(raw_minutes)
        except (TypeError, ValueError):
            minutes = default_minutes
        minutes = max(
            self.AUTO_TRADE_TIMER_MIN_MINUTES,
            min(self.AUTO_TRADE_TIMER_MAX_MINUTES, minutes),
        )

        timer = self._ensure_auto_trade_timer()
        timer.start(minutes * 60 * 1000)
        self._log_auto_trade(f"Таймер запущен на {minutes} мин")

    def _on_auto_trade_timer_timeout(self) -> None:
        self._set_auto_trade_status(False)
        self._log_auto_trade("Автоматические торги выключены по таймеру")

    @staticmethod
    def _log_auto_trade(message: str) -> None:
        text = f"[AUTO TRADE] {message}"
        Tool.write_log(text)

    def _ensure_export_button(self) -> None:
        if hasattr(self, "btn_export_trade") and hasattr(self, "btn_export_retrade"):
            return

        ensure_tab = getattr(self, "_ensure_platform_tab", None)
        if callable(ensure_tab):
            ensure_tab()

        web_tab = getattr(self.ui, "webTab", None)
        if web_tab is None:
            raise RuntimeError("Не найден webTab для кнопки экспорта")

        root_layout = web_tab.layout()
        header_layout: QHBoxLayout | None = None
        if root_layout is not None and root_layout.count() > 0:
            header_item = root_layout.itemAt(0)
            if header_item is not None:
                layout = header_item.layout()
                if isinstance(layout, QHBoxLayout):
                    header_layout = layout

        if header_layout is None:
            raise RuntimeError("Не удалось найти layout заголовка вкладки заявок")

        if not hasattr(self, "btn_export_trade"):
            self.btn_export_trade = QPushButton("Экспорт", web_tab)
            self.btn_export_trade.setObjectName("btn_export_trade")
            self.ui.btn_export_trade = self.btn_export_trade
            header_layout.addWidget(self.btn_export_trade)

        if not hasattr(self, "btn_export_retrade"):
            self.btn_export_retrade = QPushButton("Экспорт переторжки", web_tab)
            self.btn_export_retrade.setObjectName("btn_export_retrade")
            self.ui.btn_export_retrade = self.btn_export_retrade
            header_layout.addWidget(self.btn_export_retrade)

    def export_selected_trade(self) -> None:
        try:
            trade = self._get_selected_trade_for_submission_export()
            self._set_pending_submission_export_metadata(trade)
            lot_id = self._submission_lot_id_from_trade(trade)
            trade_id = self._submission_trade_id_from_trade(trade)
            self._start_export_worker(
                trade_id=trade_id,
                lot_id=lot_id,
                is_submission_acceptance=True,
                submission_search_text=self._submission_search_text_from_trade(trade),
            )
        except Exception as exc:
            self._on_export_error(str(exc))

    def export_trade(self, lot_id: int, trade: dict[str, Any] | None = None) -> None:
        trade_id: int | None = None
        submission_search_text = ""
        if isinstance(trade, dict):
            self._set_pending_submission_export_metadata(trade)
            trade_id = self._submission_trade_id_from_trade(trade)
            submission_search_text = self._submission_search_text_from_trade(trade)
        else:
            metadata = getattr(self, "_pending_submission_export_metadata", {})
            if isinstance(metadata, dict):
                metadata_lot_id = str(metadata.get("lot_id", "") or "").strip()
                if not metadata_lot_id or metadata_lot_id == str(lot_id).strip():
                    try:
                        trade_id = self._parse_positive_trade_id(
                            metadata.get("trade_id") or metadata.get("id")
                        )
                    except Exception:
                        trade_id = None
                    submission_search_text = self._submission_search_text_from_trade(metadata)
        self._start_export_worker(
            trade_id=trade_id,
            lot_id=lot_id,
            is_submission_acceptance=True,
            submission_search_text=submission_search_text,
        )

    def export_selected_retrade(self) -> None:
        attached_context = self._get_attached_retrade_export_context()
        if attached_context:
            try:
                self._start_export_worker(
                    trade_id=int(attached_context["trade_id"]),
                    lot_id=int(attached_context["lot_id"]),
                    bid_id=int(attached_context["bid_id"]),
                    is_retrade=True,
                    retrade_context=attached_context,
                )
            except Exception as exc:
                self._on_export_error(str(exc))
            return

        table_retrades = getattr(self, "table_retrades", None)
        if table_retrades is None:
            table_retrades = getattr(getattr(self, "ui", None), "table_retrades", None)
        if table_retrades is None:
            QMessageBox.warning(self, "Ошибка", "Таблица переторжек не найдена")
            return

        selected = table_retrades.currentRow()
        retrades = getattr(self, "retrades", [])
        if selected < 0 or not isinstance(retrades, list) or selected >= len(retrades):
            QMessageBox.warning(self, "Ошибка", "Выберите переторжку")
            return

        retr = retrades[selected]
        if not isinstance(retr, dict):
            QMessageBox.warning(self, "Ошибка", "Выберите переторжку")
            return

        try:
            trade_id = self._parse_positive_trade_id(retr.get("id"))
            lot_id = self._get_retrade_lot_id_for_export(retr)
            selected_offer = self._get_selected_retrade_offer_for_export()
            bid_id = self._parse_positive_bid_id(selected_offer.get("bid_id"))
            retrade_context = self._build_current_retrade_context(
                retrade=retr,
                offer=selected_offer,
                trade_id=trade_id,
                lot_id=lot_id,
                bid_id=bid_id,
            )
            self._start_export_worker(
                trade_id=trade_id,
                lot_id=lot_id,
                bid_id=bid_id,
                is_retrade=True,
                retrade_context=retrade_context,
            )
        except Exception as exc:
            self._on_export_error(str(exc))

    def _get_attached_retrade_export_context(self) -> dict[str, Any]:
        context = getattr(self, "current_retrade_context", {})
        if not isinstance(context, dict) or not context:
            return {}

        try:
            trade_id = self._parse_positive_trade_id(context.get("trade_id"))
            lot_id = self._parse_positive_lot_id(context.get("lot_id"))
            bid_id = self._parse_positive_bid_id(context.get("bid_id"))
        except Exception:
            return {}

        attached_context = dict(context)
        attached_context["trade_id"] = trade_id
        attached_context["lot_id"] = lot_id
        attached_context["bid_id"] = bid_id
        return attached_context

    @staticmethod
    def _parse_positive_trade_id(raw_trade_id: Any) -> int:
        try:
            trade_id = int(raw_trade_id)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Некорректный trade_id переторжки: {raw_trade_id}") from exc
        if trade_id <= 0:
            raise ValueError(f"Некорректный trade_id переторжки: {trade_id}")
        return trade_id

    @staticmethod
    def _parse_positive_lot_id(raw_lot_id: Any) -> int:
        try:
            lot_id = int(raw_lot_id)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Некорректный lot_id для экспорта переторжки: {raw_lot_id}") from exc
        if lot_id <= 0:
            raise ValueError(f"Некорректный lot_id для экспорта переторжки: {lot_id}")
        return lot_id

    def _get_retrade_lot_id_for_export(self, retrade: dict[str, Any]) -> int:
        lots = retrade.get("lots")
        if isinstance(lots, list) and lots:
            first_lot = lots[0]
            if isinstance(first_lot, dict):
                return self._parse_positive_lot_id(first_lot.get("id"))

        raise Exception("У выбранной переторжки отсутствует lot_id")

    @staticmethod
    def _parse_positive_bid_id(raw_bid_id: Any) -> int:
        try:
            bid_id = int(raw_bid_id)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Некорректный bid_id для экспорта переторжки: {raw_bid_id}") from exc
        if bid_id <= 0:
            raise ValueError(f"Некорректный bid_id для экспорта переторжки: {bid_id}")
        return bid_id

    def _get_selected_retrade_offer_for_export(self) -> dict[str, Any]:
        table_offers = getattr(self, "table_retrade_offers", None)
        if table_offers is None:
            table_offers = getattr(getattr(self, "ui", None), "table_retrade_offers", None)
        if table_offers is None:
            raise Exception("Таблица предложений переторжки не найдена")

        offers = getattr(self, "retrade_offers", [])
        if not isinstance(offers, list) or not offers:
            raise Exception("Нет предложений переторжки для экспорта")

        selected_row = table_offers.currentRow()
        if selected_row < 0 or selected_row >= len(offers):
            raise Exception("Выберите предложение переторжки")

        selected_offer = offers[selected_row]
        if not isinstance(selected_offer, dict):
            raise Exception("Выберите предложение переторжки")

        return selected_offer

    def _get_selected_retrade_bid_id_for_export(self) -> int:
        selected_offer = self._get_selected_retrade_offer_for_export()
        return self._parse_positive_bid_id(selected_offer.get("bid_id"))

    @staticmethod
    def _retrade_number_text(retrade: dict[str, Any]) -> str:
        return str(
            retrade.get("number")
            or retrade.get("registeredNumber")
            or retrade.get("id")
            or ""
        ).strip()

    @staticmethod
    def _retrade_offer_number_text(offer: dict[str, Any], bid_id: int) -> str:
        number = str(
            offer.get("number")
            or offer.get("registeredNumber")
            or offer.get("bidNumber")
            or ""
        ).strip()
        return number or str(bid_id)

    def _build_current_retrade_context(
        self,
        *,
        retrade: dict[str, Any],
        offer: dict[str, Any],
        trade_id: int,
        lot_id: int,
        bid_id: int,
    ) -> dict[str, Any]:
        offer_number = self._retrade_offer_number_text(offer, bid_id)
        retrade_number = self._retrade_number_text(retrade)
        return {
            "number": offer_number,
            "bid_number": offer_number,
            "bid_id": int(bid_id),
            "trade_id": int(trade_id),
            "lot_id": int(lot_id),
            "retrade_number": retrade_number,
            "title": str(retrade.get("title") or "").strip(),
            "status": str(
                retrade.get("status") or retrade.get("processStatus") or ""
            ).strip(),
            "bidder_title": str(offer.get("bidder_title") or "").strip(),
            "price": offer.get("price"),
        }

    def _set_current_retrade_context(self, context: dict[str, Any] | None) -> None:
        previous_identity = self._retrade_context_identity(
            getattr(self, "current_retrade_context", {})
        )
        retrade_context = dict(context) if isinstance(context, dict) else {}
        current_identity = self._retrade_context_identity(retrade_context)
        if previous_identity != current_identity:
            self.current_retrade_last_export_at = self._format_retrade_datetime(
                retrade_context.get("last_export_at")
            )
        elif retrade_context.get("last_export_at"):
            self.current_retrade_last_export_at = self._format_retrade_datetime(
                retrade_context.get("last_export_at")
            )
        self.current_retrade_context = retrade_context
        self.current_retrade = str(
            retrade_context.get("number")
            or retrade_context.get("bid_number")
            or ""
        ).strip()

        self.current_retrade_bid_id = None
        self.current_retrade_trade_id = None
        self.current_retrade_lot_id = None

        for attr_name, key, parser in (
            ("current_retrade_bid_id", "bid_id", self._parse_positive_bid_id),
            ("current_retrade_trade_id", "trade_id", self._parse_positive_trade_id),
            ("current_retrade_lot_id", "lot_id", self._parse_positive_lot_id),
        ):
            raw_value = retrade_context.get(key)
            if raw_value is None:
                continue
            try:
                setattr(self, attr_name, parser(raw_value))
            except Exception:
                Tool.write_log(
                    f"Некорректный {key} текущей переторжки: {raw_value}"
                )

        if self.current_retrade:
            Tool.write_log(
                "Текущая переторжка закреплена за заявкой "
                f"{self.current_retrade}"
            )
        self._refresh_retrade_context_labels()

    def _clear_current_retrade_context(self) -> None:
        self.current_retrade_last_export_at = ""
        self._set_current_retrade_context({})

    def get_current_retrade_context(self) -> dict[str, Any]:
        context = getattr(self, "current_retrade_context", {})
        return dict(context) if isinstance(context, dict) else {}

    def _get_selected_trade_for_submission_export(self) -> dict[str, Any]:
        table = getattr(self.ui, "tradesTable", None)
        if table is None:
            raise RuntimeError("Таблица приема заявок не найдена")

        selection_model = table.selectionModel()
        selected_rows = selection_model.selectedRows() if selection_model is not None else []
        if not selected_rows:
            raise ValueError("Выберите заявку в таблице перед экспортом")

        row = selected_rows[0].row()
        item = table.item(row, 0)
        if item is not None:
            trade_data = item.data(Qt.ItemDataRole.UserRole)
            if isinstance(trade_data, dict):
                return trade_data

        filtered_trades = getattr(self, "filtered_trades", [])
        if isinstance(filtered_trades, list) and 0 <= row < len(filtered_trades):
            trade = filtered_trades[row]
            if isinstance(trade, dict):
                return trade

        raise ValueError("Не удалось получить данные выбранной заявки")

    def _submission_lot_id_from_trade(self, trade: dict[str, Any]) -> int:
        lots = trade.get("lots")
        if isinstance(lots, list) and lots:
            first_lot = lots[0]
            if isinstance(first_lot, dict):
                return self._parse_positive_lot_id(first_lot.get("id"))
        raise ValueError("У выбранной заявки отсутствует lot_id")

    def _submission_trade_id_from_trade(self, trade: dict[str, Any]) -> int | None:
        try:
            return self._parse_positive_trade_id(trade.get("id"))
        except Exception:
            return None

    @staticmethod
    def _submission_search_text_from_trade(trade: dict[str, Any]) -> str:
        number = str(
            trade.get("registeredNumber")
            or trade.get("number")
            or ""
        ).strip()
        title = str(trade.get("title") or "").strip()
        return " ".join(part for part in (number, title) if part).strip()

    def _get_selected_submission_lot_id_for_export(self) -> int:
        return self._submission_lot_id_from_trade(
            self._get_selected_trade_for_submission_export()
        )

    @staticmethod
    def _trade_currency_text(trade: dict[str, Any]) -> str:
        currency = trade.get("currency")
        if isinstance(currency, dict):
            return str(
                currency.get("title")
                or currency.get("code")
                or currency.get("name")
                or ""
            ).strip()
        return str(
            trade.get("currency.title")
            or trade.get("currency")
            or ""
        ).strip()

    @classmethod
    def _submission_export_metadata_from_trade(cls, trade: dict[str, Any]) -> dict[str, str]:
        lot_id = ""
        lots = trade.get("lots")
        if isinstance(lots, list) and lots:
            first_lot = lots[0]
            if isinstance(first_lot, dict):
                lot_id = str(first_lot.get("id") or "").strip()

        return {
            "trade_id": str(trade.get("id") or "").strip(),
            "number": str(
                trade.get("registeredNumber")
                or trade.get("number")
                or ""
            ).strip(),
            "title": str(trade.get("title") or "").strip(),
            "currency": cls._trade_currency_text(trade),
            "lot_id": lot_id,
        }

    @staticmethod
    def _normalize_submission_context_metadata(
        submission_context: dict[str, Any] | None,
    ) -> dict[str, str]:
        context = submission_context if isinstance(submission_context, dict) else {}
        return {
            "customer": str(context.get("customer", "") or "").strip(),
            "producer": str(
                context.get("producer", "")
                or context.get("manufacturer", "")
                or ""
            ).strip(),
            "offer_validity_period": str(
                context.get("offer_validity_period", "") or ""
            ).strip(),
            "delivery_order": str(context.get("delivery_order", "") or "").strip(),
            "payment_terms": str(context.get("payment_terms", "") or "").strip(),
            "payment_condition": str(
                context.get("payment_condition", "") or ""
            ).strip(),
            "supplier_status": str(context.get("supplier_status", "") or "").strip(),
            "warranty": str(
                context.get("warranty", "")
                or context.get("guarantee", "")
                or ""
            ).strip(),
        }

    def _set_pending_submission_export_metadata(
        self,
        trade: dict[str, Any],
        *,
        submission_context: dict[str, Any] | None = None,
    ) -> None:
        metadata = (
            self._submission_export_metadata_from_trade(trade)
            if isinstance(trade, dict)
            else {}
        )
        context_metadata = self._normalize_submission_context_metadata(submission_context)
        if context_metadata["customer"]:
            metadata["customer"] = context_metadata["customer"]
        if context_metadata["producer"]:
            metadata["producer"] = context_metadata["producer"]
            metadata["manufacturer"] = context_metadata["producer"]
        if context_metadata["offer_validity_period"]:
            metadata["offer_validity_period"] = context_metadata["offer_validity_period"]
        if context_metadata["delivery_order"]:
            metadata["delivery_order"] = context_metadata["delivery_order"]
        if context_metadata["payment_terms"]:
            metadata["payment_terms"] = context_metadata["payment_terms"]
        elif context_metadata["payment_condition"]:
            metadata["payment_terms"] = context_metadata["payment_condition"]
        if context_metadata["payment_condition"]:
            metadata["payment_condition"] = context_metadata["payment_condition"]
        if context_metadata["supplier_status"]:
            metadata["supplier_status"] = context_metadata["supplier_status"]
        if context_metadata["warranty"]:
            metadata["warranty"] = context_metadata["warranty"]
        self._pending_submission_export_metadata = metadata

    def _start_export_worker(
        self,
        *,
        trade_id: int | None = None,
        lot_id: int | None = None,
        bid_id: int | None = None,
        is_retrade: bool = False,
        is_submission_acceptance: bool = False,
        submission_search_text: str = "",
        retrade_context: dict[str, Any] | None = None,
    ) -> None:
        if self._export_trade_worker is not None and self._export_trade_worker.isRunning():
            raise RuntimeError("Экспорт заявки уже выполняется")

        if trade_id is None and lot_id is None:
            raise ValueError("Не указан идентификатор для экспорта")
        if trade_id is not None and int(trade_id) <= 0:
            raise ValueError(f"Некорректный идентификатор для экспорта: {trade_id}")
        if lot_id is not None and int(lot_id) <= 0:
            raise ValueError(f"Некорректный идентификатор для экспорта: {lot_id}")
        if bid_id is not None and int(bid_id) <= 0:
            raise ValueError(f"Некорректный идентификатор для экспорта: {bid_id}")
        if is_retrade and lot_id is None:
            raise Exception("У выбранной переторжки отсутствует lot_id")
        if is_retrade and (trade_id is None or int(trade_id) <= 0):
            raise Exception("Не указан trade_id для открытия страницы переторжки")
        if is_retrade and (bid_id is None or int(bid_id) <= 0):
            raise Exception("Выберите предложение переторжки")
        if is_submission_acceptance and (lot_id is None or int(lot_id) <= 0):
            raise Exception("У выбранной заявки отсутствует lot_id")

        if is_retrade and bid_id is not None:
            self._pending_retrade_bid_id = int(bid_id)
            pending_context = (
                dict(retrade_context) if isinstance(retrade_context, dict) else {}
            )
            if not pending_context:
                pending_context = {
                    "number": str(bid_id),
                    "bid_number": str(bid_id),
                    "bid_id": int(bid_id),
                    "trade_id": int(trade_id),
                    "lot_id": int(lot_id),
                }
            self._pending_retrade_context = pending_context
            self.current_retrade_excel_path = ""
            self._set_current_retrade_context(pending_context)
        else:
            self._pending_retrade_bid_id = None
            self._pending_retrade_context = {}
        self._active_export_workflow = (
            "retrade"
            if is_retrade
            else "submission_acceptance"
            if is_submission_acceptance
            else "trade"
        )
        identifier_for_path: Any = (
            trade_id
            if trade_id is not None
            else lot_id
            if lot_id is not None
            else "unknown"
        )
        download_path = (
            self._build_submission_export_download_path(identifier_for_path)
            if is_submission_acceptance
            else self._build_export_download_path(identifier_for_path)
        )
        self._set_export_loading_state(is_loading=True)

        worker = ExportTradeWorker(
            trade_id=trade_id,
            lot_id=lot_id,
            bid_id=bid_id,
            is_retrade=is_retrade,
            is_submission_acceptance=is_submission_acceptance,
            submission_search_text=submission_search_text,
            download_path=download_path,
            parent=self,
        )
        worker.finished.connect(self._on_export_finished)
        worker.error.connect(self._on_export_error)
        self._export_trade_worker = worker
        worker.start()

    def _get_selected_trade_id_for_export(self) -> int:
        shared_getter = getattr(self, "_get_selected_trade_id", None)
        if callable(shared_getter):
            return int(shared_getter())

        table = getattr(self.ui, "tradesTable", None)
        if table is None:
            raise RuntimeError("Таблица заявок не найдена")

        selection_model = table.selectionModel()
        selected_rows = selection_model.selectedRows() if selection_model is not None else []
        if not selected_rows:
            raise ValueError("Выберите заявку в таблице перед экспортом")

        row = selected_rows[0].row()
        trade_item = table.item(row, 0)
        trade_text = trade_item.text().strip() if trade_item is not None else ""
        if not trade_text:
            raise ValueError("Не удалось получить trade_id из выбранной строки")

        try:
            return int(trade_text)
        except ValueError as exc:
            raise ValueError(f"Некорректный trade_id в таблице: {trade_text}") from exc

    @staticmethod
    def _build_export_download_path(identifier: Any) -> str:
        base_dir_raw = str(Config.config.get("pathToSaveExcel", "") or "").strip()
        base_dir = Path(base_dir_raw).expanduser() if base_dir_raw else (Path.home() / "Downloads")
        if base_dir.exists() and not base_dir.is_dir():
            raise NotADirectoryError(f"Папка для экспорта недоступна: {base_dir}")
        base_dir.mkdir(parents=True, exist_ok=True)

        identifier_text = str(identifier or "").strip()
        safe_identifier = re.sub(r"[^0-9A-Za-z_-]+", "_", identifier_text).strip("_") or "unknown"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"trade_{safe_identifier}_{timestamp}.xlsx"
        return str((base_dir / file_name).resolve())

    @staticmethod
    def _build_submission_export_download_path(identifier: Any) -> str:
        base_dir = Tool.user_data_dir("MyApp") / "temp" / "exports" / "submission"
        base_dir.mkdir(parents=True, exist_ok=True)

        identifier_text = str(identifier or "").strip()
        safe_identifier = re.sub(r"[^0-9A-Za-z_-]+", "_", identifier_text).strip("_") or "unknown"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"submission_{safe_identifier}_{timestamp}.xlsx"
        return str((base_dir / file_name).resolve())

    def _set_export_loading_state(self, *, is_loading: bool) -> None:
        if hasattr(self, "btn_export_trade"):
            self.btn_export_trade.setEnabled(not is_loading)
            self.btn_export_trade.setText("Экспорт..." if is_loading else "Экспорт")
        if hasattr(self, "btn_export_retrade"):
            self.btn_export_retrade.setEnabled(not is_loading)
            self.btn_export_retrade.setText(
                "Экспорт..." if is_loading else "Экспорт переторжки"
            )
        if hasattr(self, "btnGenerate"):
            self.btnGenerate.setEnabled(not is_loading)
            if getattr(self, "_generate_retrade_after_export", False):
                self.btnGenerate.setText(
                    "Обновление..." if is_loading else "Сформировать"
                )
        if is_loading:
            workflow = str(getattr(self, "_active_export_workflow", "") or "")
            if getattr(self, "_generate_retrade_after_export", False):
                status_message = "Обновление таблицы переторжки с сайта..."
            elif workflow == "submission_acceptance":
                status_message = "Экспорт таблицы приема заявок..."
            elif workflow == "retrade":
                status_message = "Экспорт таблицы переторжки..."
            else:
                status_message = "Экспорт таблицы..."
            self._show_export_status(status_message)

    def _show_export_status(self, message: str, timeout_ms: int = 0) -> None:
        show_status = getattr(self, "_show_status_message", None)
        if callable(show_status):
            show_status(message, timeout_ms)
            return
        status_bar_getter = getattr(self, "statusBar", None)
        status_bar = status_bar_getter() if callable(status_bar_getter) else None
        if status_bar is not None and message:
            status_bar.showMessage(message, timeout_ms)

    def _finish_export(self, status_message: str) -> None:
        self._set_export_loading_state(is_loading=False)
        worker = self._export_trade_worker
        self._export_trade_worker = None
        if worker is not None:
            delete_later = getattr(worker, "deleteLater", None)
            if callable(delete_later):
                delete_later()
        self._pending_retrade_bid_id = None
        self._pending_retrade_context = {}
        self._active_export_workflow = ""
        self._generate_retrade_after_export = False
        self._pending_submission_export_metadata = {}
        self._show_export_status(status_message, 5_000)

    @staticmethod
    def _developer_skip_table_fill_errors_enabled() -> bool:
        return bool(Config.settings.get("developer_skip_table_fill_errors", False))

    def _prepare_exported_excel_for_opening(self, file_path_text: str) -> bool:
        excel_processor = getattr(self, "excel_processor", None)
        if excel_processor is None:
            excel_processor = ExcelProcessor()
            self.excel_processor = excel_processor

        try:
            if not excel_processor.can_fill_exported_excel(file_path_text):
                Tool.write_log(
                    "Пропуск пост-обработки Excel: файл сформирован напрямую из JSON"
                )
                return True

            source_rows = self.get_table_rows()
            excel_processor.fill_exported_excel(file_path_text, source_rows)
            return True
        except RowCountMismatchError as exc:
            if self._developer_skip_table_fill_errors_enabled():
                Tool.write_log(
                    "Несовпадение строк Excel пропущено настройкой разработчика: "
                    f"{exc}"
                )
                return True

            action = self._ask_export_row_count_mismatch_action(exc)

            if action == self.EXPORT_MISMATCH_OPEN_WITHOUT_COPY:
                Tool.write_log(
                    f"Открытие Excel без копирования данных из таблицы: {file_path_text}"
                )
                return True

            if action == self.EXPORT_MISMATCH_OPEN_AND_COPY:
                Tool.write_log(
                    "Частичное копирование данных в Excel при несовпадении строк: "
                    f"Excel={exc.excel_rows}, Таблица={exc.source_rows}"
                )
                excel_processor.fill_exported_excel(
                    file_path_text,
                    source_rows,
                    strict_row_count=False,
                )
                return True

            Tool.write_log(
                f"Открытие Excel отменено после несовпадения строк: {file_path_text}"
            )
            return False
        except Exception as exc:
            if self._developer_skip_table_fill_errors_enabled():
                Tool.write_log(
                    "Ошибка пост-обработки Excel пропущена настройкой разработчика: "
                    f"{exc}"
                )
                return True
            raise

    def _ask_export_row_count_mismatch_action(
        self,
        mismatch: RowCountMismatchError,
    ) -> str:
        message_box = QMessageBox(self)
        icon_enum = getattr(QMessageBox, "Icon", None)
        warning_icon = (
            getattr(icon_enum, "Warning", None)
            if icon_enum is not None
            else getattr(QMessageBox, "Warning", None)
        )
        if warning_icon is not None:
            message_box.setIcon(warning_icon)

        message_box.setWindowTitle("Несовпадение строк")
        message_box.setText(str(mismatch))
        message_box.setInformativeText(
            "Можно открыть скачанный Excel без копирования данных из таблицы. "
            "При копировании будут заполнены только строки, которые есть и в Excel, "
            "и в таблице."
        )

        role_enum = getattr(QMessageBox, "ButtonRole", QMessageBox)
        accept_role = getattr(role_enum, "AcceptRole", 0)
        action_role = getattr(role_enum, "ActionRole", accept_role)
        reject_role = getattr(role_enum, "RejectRole", accept_role)
        open_button = message_box.addButton(
            "Открыть без копирования",
            accept_role,
        )
        copy_button = message_box.addButton(
            "Открыть и копировать",
            action_role,
        )
        cancel_button = message_box.addButton(
            "Отмена",
            reject_role,
        )
        message_box.setDefaultButton(open_button)
        message_box.setEscapeButton(cancel_button)
        message_box.exec()

        clicked_button = message_box.clickedButton()
        if clicked_button == open_button:
            return self.EXPORT_MISMATCH_OPEN_WITHOUT_COPY
        if clicked_button == copy_button:
            return self.EXPORT_MISMATCH_OPEN_AND_COPY
        return self.EXPORT_MISMATCH_CANCEL

    def get_table_rows(
        self,
        *,
        default_manufacturer: str = "",
        default_supplier_status: str = "",
        default_warranty: str = "",
    ) -> list[dict]:
        table = getattr(getattr(self, "ui", None), "KpTable", None)
        if table is None:
            return []

        rows: list[dict] = []
        row_count = table.rowCount()
        column_count = table.columnCount()
        headers = [
            table.horizontalHeaderItem(column).text()
            if table.horizontalHeaderItem(column) is not None
            else ""
            for column in range(column_count)
        ]
        normalized_headers = [self._normalize_table_header(header) for header in headers]

        def find_column(kind: str) -> int | None:
            for column, normalized in enumerate(normalized_headers):
                if not normalized:
                    continue
                if kind == "number" and normalized in {"№", "n", "номер"}:
                    return column
                if kind == "name" and "наимен" in normalized:
                    return column
                if kind == "sku" and ("каталож" in normalized or "артикул" in normalized):
                    return column
                if kind == "unit" and ("едизм" in normalized or "единицаизмер" in normalized):
                    return column
                if kind == "qty" and ("колво" in normalized or "количество" in normalized):
                    return column
                if kind == "manufacturer" and "производ" in normalized:
                    return column
                if kind == "supplier_status" and "статус" in normalized and "постав" in normalized:
                    return column
                if kind == "warranty" and "гарант" in normalized:
                    return column
                if kind == "sale_price" and (
                    "ценареализациизаедбезндс" in normalized
                    or (
                        "ценареализациизаед" in normalized
                        and "ндс" in normalized
                    )
                ):
                    return column
                if kind == "sale_total" and (
                    "итогореализациибезндс" in normalized
                    or "суммареализациибезндс" in normalized
                ):
                    return column
            return None

        def safe_column(header_kind: str, fallback: int) -> int | None:
            detected = find_column(header_kind)
            if detected is not None:
                return detected
            return fallback if fallback < column_count else None

        number_col = safe_column("number", 0)
        name_col = safe_column("name", 1)
        sku_col = safe_column("sku", 2)
        unit_col = safe_column("unit", 3)
        qty_col = safe_column("qty", 4)
        base_price_col = 5 if 5 < column_count else None
        base_total_col = 6 if 6 < column_count else None
        sale_price_col = find_column("sale_price")
        sale_total_col = find_column("sale_total")
        final_price_col = (
            sale_price_col
            if sale_price_col is not None
            else 10
            if 10 < column_count
            else base_price_col
        )
        final_total_col = (
            sale_total_col
            if sale_total_col is not None
            else 11
            if 11 < column_count
            else base_total_col
        )
        delivery_col = 13 if 13 < column_count else None
        supplier_delivery_col = 14 if 14 < column_count else None
        manufacturer_col = find_column("manufacturer")
        supplier_status_col = find_column("supplier_status")
        warranty_col = find_column("warranty")
        default_manufacturer_text = str(default_manufacturer or "").strip()
        default_supplier_status_text = str(default_supplier_status or "").strip()
        default_warranty_text = str(default_warranty or "").strip()
        table_data = getattr(self, "tableData", {})
        table_data_currencies = (
            table_data.get("currency", [])
            if isinstance(table_data, dict)
            else []
        )
        table_currency = (
            CurrencyService.detect_currency_from_values(table_data_currencies)
            or CurrencyService.detect_currency_from_values(headers)
        )

        def table_data_currency(row: int) -> str:
            if isinstance(table_data_currencies, (list, tuple)):
                if 0 <= row < len(table_data_currencies):
                    return CurrencyService.detect_currency_from_value(
                        table_data_currencies[row]
                    )
                return ""
            return CurrencyService.detect_currency_from_value(table_data_currencies)

        def cell_text(row: int, column: int | None) -> str:
            if column is None or column < 0 or column >= column_count:
                return ""
            item = table.item(row, column)
            if item is None:
                return ""
            value = self._table_item_edit_value(item)
            if value is None:
                return ""
            return str(value).strip()

        def first_text(row: int, *columns: int | None) -> str:
            for column in columns:
                text = cell_text(row, column)
                if text:
                    return text
            return ""

        def money_value(text: str) -> float | str:
            parsed = self._parse_retrade_number_or_none(text)
            if parsed is not None:
                return parsed
            return text

        def is_zero_amount(text: str) -> bool:
            parsed = self._parse_retrade_number_or_none(text)
            return parsed is not None and abs(parsed) < 1e-9

        def skip_submission_defaults(
            source_price_text: str,
            price_text: str,
            total_text: str,
        ) -> bool:
            return is_zero_amount(source_price_text) or is_zero_amount(price_text) or (
                not price_text and is_zero_amount(total_text)
            )

        for row_index in range(row_count):
            source_price_text = first_text(row_index, base_price_col)
            source_total_text = first_text(row_index, base_total_col)
            price_text = first_text(row_index, final_price_col)
            total_text = first_text(row_index, final_total_col)
            row_currency = (
                CurrencyService.detect_currency_from_values(
                    (
                        price_text,
                        total_text,
                        source_price_text,
                        source_total_text,
                    )
                )
                or table_data_currency(row_index)
                or table_currency
            )
            delivery_time = first_text(row_index, delivery_col, supplier_delivery_col)
            manufacturer = cell_text(row_index, manufacturer_col)
            if not manufacturer:
                manufacturer = default_manufacturer_text
            if skip_submission_defaults(source_price_text, price_text, total_text):
                supplier_status = ""
                warranty = ""
            else:
                supplier_status = cell_text(row_index, supplier_status_col)
                if not supplier_status:
                    supplier_status = default_supplier_status_text
                warranty = cell_text(row_index, warranty_col)
                if not warranty:
                    warranty = default_warranty_text

            rows.append(
                {
                    "number": cell_text(row_index, number_col),
                    "name": cell_text(row_index, name_col),
                    "sku": cell_text(row_index, sku_col),
                    "unit": cell_text(row_index, unit_col),
                    "qty": cell_text(row_index, qty_col),
                    "price": money_value(price_text) if price_text else "",
                    "total": money_value(total_text) if total_text else "",
                    "delivery_time": delivery_time,
                    "supplier_delivery_time": cell_text(row_index, supplier_delivery_col),
                    "manufacturer": manufacturer,
                    "tech_characteristics": manufacturer,
                    "technical_characteristics": manufacturer,
                    "supplier_status": supplier_status,
                    "warranty": warranty,
                    "guarantee": warranty,
                    "currency": row_currency,
                }
            )

        return rows

    @staticmethod
    def _log_ui(message: str) -> None:
        text = f"[UI] {message}"
        Tool.write_log(text)

    def update_retrade_table(self, df: pd.DataFrame) -> None:
        table = getattr(self, "retrade_table", None)
        if table is None:
            raise RuntimeError("Таблица Переторжка не найдена")
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame()

        self._log_ui("Обновление таблицы Переторжка")

        self._configure_excel_like_table(table)
        table.clear()
        table.setRowCount(len(df))
        table.setColumnCount(len(df.columns))
        table.setHorizontalHeaderLabels(df.columns.tolist())

        for row_index in range(len(df)):
            for col_index in range(len(df.columns)):
                cell_value = df.iloc[row_index, col_index]
                value = "" if pd.isna(cell_value) else str(cell_value)
                item = QTableWidgetItem(value)
                if self._is_numeric_table_value(cell_value):
                    item.setTextAlignment(
                        int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    )
                else:
                    item.setTextAlignment(
                        int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    )
                table.setItem(row_index, col_index, item)

        resize_table_to_contents(table)
        self._log_ui(f"Строк: {len(df)}")

    def _activate_retrade_tab(self) -> None:
        tabs = getattr(self, "tabWidget", None)
        if tabs is None:
            tabs = getattr(getattr(self, "ui", None), "tabWidget", None)
        retrade_tab = getattr(self, "retrade_tab", None)
        if isinstance(tabs, QTabWidget) and retrade_tab is not None:
            index = tabs.indexOf(retrade_tab)
            if index >= 0:
                tabs.setCurrentIndex(index)
        inner_tabs = getattr(self, "retrade_inner_tabs", None)
        if isinstance(inner_tabs, QTabWidget):
            inner_tabs.setCurrentIndex(self.RETRADE_INNER_TAB_MAIN)

    def _activate_submission_acceptance_tab(self) -> None:
        tabs = getattr(self, "tabWidget", None)
        if tabs is None:
            tabs = getattr(getattr(self, "ui", None), "tabWidget", None)
        web_tab = getattr(getattr(self, "ui", None), "webTab", None)
        if isinstance(tabs, QTabWidget) and web_tab is not None:
            index = tabs.indexOf(web_tab)
            if index >= 0:
                tabs.setCurrentIndex(index)

    def _ensure_submission_acceptance_export_table(self) -> QTableWidget:
        table = getattr(getattr(self, "ui", None), "submission_acceptance_export_table", None)
        if isinstance(table, QTableWidget):
            return table

        web_tab = getattr(getattr(self, "ui", None), "webTab", None)
        if web_tab is None:
            raise RuntimeError("Вкладка 'Загрузка с ЭТП' не найдена")

        root_layout = web_tab.layout()
        if root_layout is None:
            raise RuntimeError("Не найден layout вкладки 'Загрузка с ЭТП'")

        label = QLabel("Экспорт приема заявок", web_tab)
        label.setObjectName("submissionAcceptanceExportLabel")

        table = QTableWidget(web_tab)
        table.setObjectName("submission_acceptance_export_table")
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)
        table.verticalHeader().setVisible(False)
        configure_table_autosize(table)

        trades_table = getattr(getattr(self, "ui", None), "tradesTable", None)
        insert_index = root_layout.indexOf(trades_table) if trades_table is not None else -1
        if insert_index >= 0:
            root_layout.insertWidget(insert_index + 1, label)
            root_layout.insertWidget(insert_index + 2, table)
        else:
            root_layout.addWidget(label)
            root_layout.addWidget(table)

        self.submission_acceptance_export_label = label
        self.submission_acceptance_export_table = table
        self.ui.submission_acceptance_export_label = label
        self.ui.submission_acceptance_export_table = table
        return table

    @staticmethod
    def _dataframe_to_display_text(value: Any) -> str:
        if pd.isna(value):
            return ""
        return str(value)

    def update_submission_acceptance_export_table(self, dataframe: pd.DataFrame) -> None:
        table = self._ensure_submission_acceptance_export_table()
        frame = dataframe if isinstance(dataframe, pd.DataFrame) else pd.DataFrame()
        headers = [str(column) for column in frame.columns.tolist()]

        table.setUpdatesEnabled(False)
        try:
            table.clear()
            table.setRowCount(len(frame))
            table.setColumnCount(len(headers))
            if headers:
                table.setHorizontalHeaderLabels(headers)
            for row_index, row_values in enumerate(frame.itertuples(index=False, name=None)):
                for column_index, value in enumerate(row_values):
                    table.setItem(
                        row_index,
                        column_index,
                        QTableWidgetItem(self._dataframe_to_display_text(value)),
                    )
        finally:
            table.setUpdatesEnabled(True)

        resize_table_to_contents(table)
        label = getattr(self, "submission_acceptance_export_label", None)
        if isinstance(label, QLabel):
            label.show()
        table.show()

    def _prepare_submission_export_for_loading(self, file_path_text: str) -> None:
        excel_processor = getattr(self, "excel_processor", None)
        if excel_processor is None:
            excel_processor = ExcelProcessor()
            self.excel_processor = excel_processor

        try:
            metadata = getattr(self, "_pending_submission_export_metadata", {})
            default_manufacturer = ""
            if isinstance(metadata, dict):
                default_manufacturer = str(
                    metadata.get("producer", "")
                    or metadata.get("manufacturer", "")
                    or ""
                ).strip()
                default_supplier_status = str(
                    metadata.get("supplier_status", "") or ""
                ).strip()
                default_warranty = str(
                    metadata.get("warranty", "")
                    or metadata.get("guarantee", "")
                    or ""
                ).strip()
            else:
                default_supplier_status = ""
                default_warranty = ""
            source_rows = self.get_table_rows(
                default_manufacturer=default_manufacturer,
                default_supplier_status=default_supplier_status,
                default_warranty=default_warranty,
            )
            source_currency = CurrencyService.detect_currency_from_values(source_rows)
            if source_currency:
                if not isinstance(metadata, dict):
                    metadata = {}
                metadata["currency"] = source_currency
                self._pending_submission_export_metadata = metadata
            if not source_rows:
                Tool.write_log(
                    "Заполнение файла приема заявок пропущено: Подготовка КП пуста"
                )
                return

            if not excel_processor.can_fill_exported_excel(file_path_text):
                Tool.write_log(
                    "Заполнение файла приема заявок пропущено: не найдены подходящие колонки"
                )
                return

            excel_processor.fill_exported_excel(
                file_path_text,
                source_rows,
                strict_row_count=False,
            )
            Tool.write_log(
                "Файл приема заявок заполнен данными из Полной таблицы: "
                f"{file_path_text}"
            )
        except Exception as exc:
            if self._developer_skip_table_fill_errors_enabled():
                Tool.write_log(
                    "Ошибка заполнения файла приема заявок пропущена настройкой "
                    f"разработчика: {exc}"
                )
                return
            raise

    def _on_submission_acceptance_export_finished(self, file_path_text: str) -> None:
        if not file_path_text:
            QMessageBox.warning(self, "Экспорт приема заявок", "Файл не был скачан")
            self._finish_export("Экспорт не выполнен")
            return

        try:
            export_path = Path(file_path_text).expanduser()
            if not export_path.exists() or not export_path.is_file():
                raise FileNotFoundError(f"Excel файл не найден: {export_path}")

            self.current_submission_acceptance_excel_path = str(export_path.resolve())
            self._prepare_submission_export_for_loading(str(export_path))
            load_submission_export = getattr(self, "load_submission_export_file", None)
            if not callable(load_submission_export):
                raise RuntimeError("Метод загрузки во вкладку Подача заявки не найден")
            rows_count = int(load_submission_export(str(export_path)))
            apply_submission_metadata = getattr(
                self,
                "apply_submission_export_metadata",
                None,
            )
            if callable(apply_submission_metadata):
                apply_submission_metadata(
                    getattr(self, "_pending_submission_export_metadata", {})
                )
        except Exception as exc:
            error_text = str(exc or "Ошибка обработки Excel")
            Tool.write_log(f"Ошибка загрузки Excel во вкладку Подача заявки: {error_text}")
            QMessageBox.critical(self, "Ошибка", error_text)
            set_pipeline_error_status = getattr(self, "_set_pipeline_error_status", None)
            if callable(set_pipeline_error_status):
                set_pipeline_error_status()
            self._finish_export("Ошибка обработки Excel")
            return

        Tool.write_log(f"Экспорт приема заявок завершен: {file_path_text}")
        QMessageBox.information(
            self,
            "Экспорт приема заявок",
            f"Файл успешно экспортирован и загружен во вкладку Подача заявки:\n"
            f"{file_path_text}\n\n"
            f"Позиций: {rows_count}",
        )
        self._finish_export("Excel приема заявок экспортирован")

    def _on_export_finished(self, file_path: str) -> None:
        file_path_text = str(file_path or "").strip()
        workflow = str(getattr(self, "_active_export_workflow", "") or "")
        generate_after_export = bool(
            getattr(self, "_generate_retrade_after_export", False)
        )

        if workflow == "submission_acceptance":
            self._on_submission_acceptance_export_finished(file_path_text)
            return

        if not file_path_text:
            Tool.write_log("Экспорт переторжки пропущен: пользователь не участвует")
            QMessageBox.information(
                self,
                "Экспорт заявки",
                "Экспорт пропущен: нет участия в выбранной переторжке",
            )
            self._finish_export("Экспорт пропущен")
            return

        if file_path_text:
            try:
                if not self._prepare_exported_excel_for_opening(file_path_text):
                    self._finish_export("Открытие Excel отменено")
                    return

                export_path = Path(file_path_text).expanduser()
                if not export_path.exists() or not export_path.is_file():
                    raise FileNotFoundError(f"Excel файл не найден: {export_path}")

                self.current_retrade_excel_path = str(export_path.resolve())
                pending_context = getattr(self, "_pending_retrade_context", {})
                if isinstance(pending_context, dict) and pending_context:
                    self._set_current_retrade_context(pending_context)
                else:
                    pending_bid_id = getattr(self, "_pending_retrade_bid_id", None)
                    self.current_retrade_bid_id = (
                        int(pending_bid_id)
                        if pending_bid_id is not None
                        else None
                    )

                try:
                    dataframe = pd.read_excel(export_path)
                except ValueError as exc:
                    if "No columns to parse from file" in str(exc):
                        dataframe = pd.DataFrame()
                    else:
                        raise

                self._log_ui("Excel загружен")
                update_method = getattr(getattr(self, "ui", None), "update_retrade_table", None)
                if callable(update_method):
                    update_method(dataframe)
                else:
                    self.update_retrade_table(dataframe)
                if workflow == "retrade":
                    self._mark_current_retrade_table_exported_now()
                self._activate_retrade_tab()
            except Exception as exc:
                error_text = str(exc or "Ошибка обработки Excel")
                Tool.write_log(f"Ошибка пост-обработки Excel: {error_text}")
                QMessageBox.critical(self, "Ошибка", error_text)
                set_pipeline_error_status = getattr(self, "_set_pipeline_error_status", None)
                if callable(set_pipeline_error_status):
                    set_pipeline_error_status()
                self._finish_export("Ошибка пост-обработки Excel")
                return

        if workflow == "retrade" and generate_after_export:
            Tool.write_log(
                "Свежий экспорт переторжки загружен, продолжаю формирование расчета"
            )
            self._finish_export("Excel переторжки обновлен")
            self.generate_retrade_calculation(refresh_from_site=False)
            return

        info_text = (
            f"Файл успешно экспортирован:\n{file_path_text}"
            if file_path_text
            else "Файл успешно экспортирован"
        )
        Tool.write_log(f"Экспорт заявки завершен: {file_path_text}")
        QMessageBox.information(self, "Экспорт заявки", info_text)
        self._finish_export("Excel файл экспортирован")

    def _on_export_error(self, message: str) -> None:
        error_text = str(message or "Неизвестная ошибка")
        Tool.write_log(f"Ошибка экспорта заявки: {error_text}")
        QMessageBox.warning(self, "Ошибка экспорта заявки", error_text)
        self._finish_export("Ошибка экспорта заявки")
