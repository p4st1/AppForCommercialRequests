import unittest
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

if Workbook is not None:
    from create import createExcelFile
else:
    createExcelFile = None


@unittest.skipIf(Workbook is None, "openpyxl is not installed")
class CreateExcelLinkTests(unittest.TestCase):
    def test_append_remote_links_leave_gap_before_urls(self):
        workbook = Workbook()
        worksheet = workbook.active

        createExcelFile._append_docx_remote_link(
            worksheet,
            12,
            "https://drive.google.com/file/d/docx/view",
        )
        createExcelFile._append_calculations_remote_link(
            worksheet,
            13,
            "https://drive.google.com/file/d/xlsx/view",
        )

        self.assertEqual(worksheet["D12"].value, "Ссылка на КП DOCX")
        self.assertEqual(
            worksheet["E12"].value,
            "https://drive.google.com/file/d/docx/view",
        )
        self.assertEqual(
            worksheet["E12"].hyperlink.target,
            "https://drive.google.com/file/d/docx/view",
        )
        self.assertEqual(worksheet["D13"].value, "Ссылка на расчеты")
        self.assertEqual(
            worksheet["E13"].hyperlink.target,
            "https://drive.google.com/file/d/xlsx/view",
        )

    def test_append_calculations_link_aligns_legacy_docx_row(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "links.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet["A12"] = "Ссылка на КП DOCX"
            worksheet["B12"] = "https://drive.google.com/file/d/docx/view"
            worksheet["B12"].hyperlink = "https://drive.google.com/file/d/docx/view"
            workbook.save(file_path)

            createExcelFile.append_calculations_remote_link_to_file(
                str(file_path),
                "https://drive.google.com/file/d/xlsx/view",
            )

            updated = load_workbook(file_path)
            try:
                worksheet = updated.active
                self.assertIsNone(worksheet["A12"].value)
                self.assertIsNone(worksheet["B12"].value)
                self.assertEqual(worksheet["D12"].value, "Ссылка на КП DOCX")
                self.assertEqual(
                    worksheet["E12"].hyperlink.target,
                    "https://drive.google.com/file/d/docx/view",
                )
                self.assertEqual(worksheet["D13"].value, "Ссылка на расчеты")
                self.assertEqual(
                    worksheet["E13"].hyperlink.target,
                    "https://drive.google.com/file/d/xlsx/view",
                )
            finally:
                updated.close()


if __name__ == "__main__":
    unittest.main()
