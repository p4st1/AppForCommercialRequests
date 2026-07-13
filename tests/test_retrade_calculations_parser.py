import sys
import tempfile
import unittest
from pathlib import Path
from types import ModuleType, SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def _ensure_pyside_stubs() -> None:
    pyside6 = sys.modules.get("PySide6")
    if pyside6 is None:
        pyside6 = ModuleType("PySide6")
        sys.modules["PySide6"] = pyside6

    qtcore = sys.modules.get("PySide6.QtCore")
    if qtcore is None:
        qtcore = ModuleType("PySide6.QtCore")
        sys.modules["PySide6.QtCore"] = qtcore

    class _Signal:
        def __init__(self, *_args, **_kwargs):
            self._callbacks = []

        def connect(self, callback):
            self._callbacks.append(callback)

        def emit(self, *args, **kwargs):
            for callback in list(self._callbacks):
                callback(*args, **kwargs)

    class _QThread:
        def __init__(self, *_args, **_kwargs):
            pass

        def start(self):
            run = getattr(self, "run", None)
            if callable(run):
                run()

    class _QSettings:
        _values = {}

        def __init__(self, *_args, **_kwargs):
            pass

        def value(self, key, default=None, type=None):
            value = self._values.get(key, default)
            if type is not None and value is not None:
                try:
                    return type(value)
                except Exception:
                    return default
            return value

        def setValue(self, key, value):
            self._values[key] = value

    class _QTimer:
        def __init__(self, *_args, **_kwargs):
            self.timeout = _Signal()

        def setSingleShot(self, *_args, **_kwargs):
            pass

        def setInterval(self, *_args, **_kwargs):
            pass

        def isActive(self):
            return False

        def start(self, *_args, **_kwargs):
            pass

        def stop(self):
            pass

    class _QDateTime:
        @staticmethod
        def currentDateTime():
            return _QDateTime()

        def addSecs(self, *_args, **_kwargs):
            return self

        def msecsTo(self, *_args, **_kwargs):
            return 0

        def toString(self, *_args, **_kwargs):
            return ""

    qtcore.QSettings = getattr(qtcore, "QSettings", _QSettings)
    qtcore.QDateTime = getattr(qtcore, "QDateTime", _QDateTime)
    qtcore.QThread = getattr(qtcore, "QThread", _QThread)
    qtcore.Signal = getattr(qtcore, "Signal", _Signal)
    qtcore.QTimer = getattr(qtcore, "QTimer", _QTimer)

    qt = getattr(qtcore, "Qt", type("Qt", (), {})())
    for nested_name, values in {
        "ItemDataRole": {"EditRole": 2, "UserRole": 32, "BackgroundRole": 8},
        "AlignmentFlag": {
            "AlignLeft": 1,
            "AlignRight": 2,
            "AlignVCenter": 4,
            "AlignCenter": 8,
        },
        "PenStyle": {"SolidLine": 1},
        "ItemFlag": {
            "ItemIsEditable": 1,
            "ItemIsUserCheckable": 2,
            "ItemIsEnabled": 4,
        },
        "CheckState": {"Unchecked": 0, "Checked": 2},
    }.items():
        nested = getattr(qt, nested_name, type(nested_name, (), {})())
        for attr, value in values.items():
            if not hasattr(nested, attr):
                setattr(nested, attr, value)
        setattr(qt, nested_name, nested)
    qtcore.Qt = qt
    pyside6.QtCore = qtcore

    qtgui = sys.modules.get("PySide6.QtGui")
    if qtgui is None:
        qtgui = ModuleType("PySide6.QtGui")
        sys.modules["PySide6.QtGui"] = qtgui

    class _QAction:
        def __init__(self, *_args, **_kwargs):
            self.triggered = _Signal()

    class _QColor:
        def __init__(self, *_args, **_kwargs):
            pass

        def isValid(self):
            return True

    qtgui.QAction = getattr(qtgui, "QAction", _QAction)
    qtgui.QColor = getattr(qtgui, "QColor", _QColor)
    pyside6.QtGui = qtgui

    qtuitools = sys.modules.get("PySide6.QtUiTools")
    if qtuitools is None:
        qtuitools = ModuleType("PySide6.QtUiTools")
        sys.modules["PySide6.QtUiTools"] = qtuitools
    if not hasattr(qtuitools, "loadUiType"):
        qtuitools.loadUiType = lambda *_args, **_kwargs: (object, object)

    class _QUiLoader:
        def load(self, *_args, **_kwargs):
            return None

    qtuitools.QUiLoader = getattr(qtuitools, "QUiLoader", _QUiLoader)
    pyside6.QtUiTools = qtuitools

    qtwidgets = sys.modules.get("PySide6.QtWidgets")
    if qtwidgets is None:
        qtwidgets = ModuleType("PySide6.QtWidgets")
        sys.modules["PySide6.QtWidgets"] = qtwidgets

    class _Widget:
        def __init__(self, *_args, **_kwargs):
            pass

    class _QFileDialog:
        @staticmethod
        def getOpenFileName(*_args, **_kwargs):
            return ("", "")

    class _QInputDialog:
        @staticmethod
        def getText(*_args, **_kwargs):
            return ("", False)

    class _QDialog(_Widget):
        class DialogCode:
            Accepted = 1

    class _QDialogButtonBox(_Widget):
        class StandardButton:
            Ok = 1
            Cancel = 2

    class _QMessageBox:
        @staticmethod
        def warning(*_args, **_kwargs):
            return 0

        @staticmethod
        def critical(*_args, **_kwargs):
            return 0

        @staticmethod
        def information(*_args, **_kwargs):
            return 0

    class _QAbstractItemView:
        SelectionBehavior = type("SelectionBehavior", (), {"SelectRows": 1})
        EditTrigger = type("EditTrigger", (), {"NoEditTriggers": 0})

    class _QHeaderView:
        ResizeMode = type(
            "ResizeMode",
            (),
            {"Interactive": 1, "ResizeToContents": 2, "Stretch": 3},
        )

    class _QTableWidgetItem:
        def __init__(self, text=""):
            self._text = str(text)
            self._data = {}
            self._flags = 0

        def text(self):
            return self._text

        def setText(self, text):
            self._text = str(text)

        def data(self, role):
            return self._data.get(role)

        def setData(self, role, value):
            self._data[role] = value

        def flags(self):
            return self._flags

        def setFlags(self, flags):
            self._flags = flags

        def setTextAlignment(self, *_args, **_kwargs):
            pass

        def setCheckState(self, *_args, **_kwargs):
            pass

        def setBackground(self, *_args, **_kwargs):
            pass

    for name, value in {
        "QAbstractItemView": _QAbstractItemView,
        "QCheckBox": _Widget,
        "QDateTimeEdit": _Widget,
        "QDialog": _QDialog,
        "QDialogButtonBox": _QDialogButtonBox,
        "QDoubleSpinBox": _Widget,
        "QFileDialog": _QFileDialog,
        "QFormLayout": _Widget,
        "QInputDialog": _QInputDialog,
        "QHeaderView": _QHeaderView,
        "QHBoxLayout": _Widget,
        "QLabel": _Widget,
        "QListWidget": _Widget,
        "QListWidgetItem": _Widget,
        "QMessageBox": _QMessageBox,
        "QPushButton": _Widget,
        "QSpinBox": _Widget,
        "QTableWidget": _Widget,
        "QTableWidgetItem": _QTableWidgetItem,
        "QTabWidget": _Widget,
        "QVBoxLayout": _Widget,
        "QWidget": _Widget,
    }.items():
        if not hasattr(qtwidgets, name):
            setattr(qtwidgets, name, value)
    pyside6.QtWidgets = qtwidgets


_ensure_pyside_stubs()

try:
    import requests  # noqa: F401
except ModuleNotFoundError:
    sys.modules["requests"] = ModuleType("requests")

import ui_mixins.export_mixin as export_mixin_module
from ui_mixins.export_mixin import ExportMixin


class _FakeItem:
    def __init__(self, text):
        self._text = text

    def text(self):
        return str(self._text)


class _FakeRetradeTable:
    def __init__(self, headers, rows):
        self._headers = [_FakeItem(header) for header in headers]
        self._rows = [
            [None if value is None else _FakeItem(value) for value in row]
            for row in rows
        ]

    def columnCount(self):
        return len(self._headers)

    def rowCount(self):
        return len(self._rows)

    def horizontalHeaderItem(self, column):
        if column < 0 or column >= len(self._headers):
            return None
        return self._headers[column]

    def item(self, row, column):
        if row < 0 or row >= len(self._rows):
            return None
        row_values = self._rows[row]
        if column < 0 or column >= len(row_values):
            return None
        return row_values[column]


class _FakeEditableItem:
    def __init__(self, value):
        self._value = value

    def data(self, _role):
        return self._value

    def text(self):
        return "" if self._value is None else str(self._value)


class _FakeEditableTable:
    def __init__(self, rows, headers=None):
        self._headers = [_FakeItem(header) for header in (headers or [])]
        self._rows = [
            [None if value is None else _FakeEditableItem(value) for value in row]
            for row in rows
        ]

    def rowCount(self):
        return len(self._rows)

    def columnCount(self):
        return max([len(self._headers), *(len(row) for row in self._rows)], default=0)

    def horizontalHeaderItem(self, column):
        if column < 0 or column >= len(self._headers):
            return None
        return self._headers[column]

    def item(self, row, column):
        if row < 0 or row >= len(self._rows):
            return None
        row_values = self._rows[row]
        if column < 0 or column >= len(row_values):
            return None
        return row_values[column]


class RetradeCalculationsParserTests(unittest.TestCase):
    def test_parse_uses_first_non_empty_row_as_headers(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": "", "currency": None}, {"value": None, "currency": None}],
                [{"value": "№", "currency": None}, {"value": "Цена", "currency": None}],
                [{"value": 1, "currency": None}, {"value": 1000, "currency": "RUB"}],
                [{"value": 2, "currency": None}, {"value": 5, "currency": None}],
            ]
        )
        self.assertEqual(
            parsed["headers"],
            ["№", "Цена"],
        )
        self.assertEqual(
            parsed["rows"],
            [
                [
                    {"value": 1, "currency": None},
                    {"value": 1000, "currency": "RUB"},
                ],
                [
                    {"value": 2, "currency": None},
                    {"value": 5, "currency": "RUB"},
                ],
            ],
        )
        self.assertIsNone(parsed["total_without_vat"])
        self.assertIsNone(parsed["total_without_vat_currency"])
        self.assertEqual(parsed["totals"], {"price": 0.0, "logistic": 0.0, "customs": 0.0})
        self.assertEqual(
            parsed["totals_currency"],
            {"price": None, "logistic": None, "customs": None},
        )

    def test_parse_skips_fully_empty_rows_and_keeps_all_data_rows(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": "Код", "currency": None}, {"value": "Значение", "currency": None}],
                [{"value": "", "currency": None}, {"value": None, "currency": None}],
                [{"value": 2, "currency": None}, {"value": 20, "currency": "USD"}],
                [{"value": 3, "currency": None}, {"value": "30 руб", "currency": "RUB"}],
            ]
        )
        self.assertEqual(parsed["headers"], ["Код", "Значение"])
        self.assertEqual(
            parsed["rows"],
            [
                [
                    {"value": 2, "currency": None},
                    {"value": 20, "currency": "USD"},
                ],
                [
                    {"value": 3, "currency": None},
                    {"value": "30 руб", "currency": "RUB"},
                ],
            ],
        )
        self.assertIsNone(parsed["total_without_vat"])
        self.assertIsNone(parsed["total_without_vat_currency"])
        self.assertEqual(parsed["totals"], {"price": 0.0, "logistic": 0.0, "customs": 0.0})
        self.assertEqual(
            parsed["totals_currency"],
            {"price": None, "logistic": None, "customs": None},
        )

    def test_parse_skips_rows_that_only_have_currency_formatting(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [
                    {"value": None, "currency": "RUB"},
                    {"value": "", "currency": "RUB"},
                ],
                [
                    {"value": "№", "currency": None},
                    {"value": "Цена", "currency": None},
                ],
                [
                    {"value": None, "currency": "RUB"},
                    {"value": None, "currency": "RUB"},
                ],
                [{"value": 1, "currency": None}, {"value": 100, "currency": "RUB"}],
            ]
        )

        self.assertEqual(parsed["headers"], ["№", "Цена"])
        self.assertEqual(
            parsed["rows"],
            [[{"value": 1, "currency": None}, {"value": 100, "currency": "RUB"}]],
        )

    def test_parse_propagates_currency_within_column(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": "№", "currency": None}, {"value": "Цена", "currency": None}],
                [{"value": 1, "currency": None}, {"value": 100, "currency": "CNY"}],
                [{"value": 2, "currency": None}, {"value": 200, "currency": None}],
            ]
        )

        self.assertEqual(parsed["rows"][1][1]["currency"], "CNY")

    def test_worksheet_visible_rows_skip_empty_rows_before_and_after_headers(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([None, None])
        worksheet.append(["№", "Цена"])
        worksheet.append([None, None])
        worksheet.append([1, 100])
        worksheet["B1"].number_format = '#,##0.00 "₽"'
        worksheet["B3"].number_format = '#,##0.00 "₽"'

        headers, rows, column_count = ExportMixin._worksheet_to_visible_retrade_rows(
            worksheet
        )

        self.assertEqual(column_count, 2)
        self.assertEqual([cell["value"] for cell in headers], ["№", "Цена"])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 4)
        self.assertEqual([cell["value"] for cell in rows[0][1]], [1, 100])

    def test_worksheet_visible_rows_skip_empty_internal_columns(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["№", "Срок поставщика", None, None, "Рейтинг"])
        worksheet.append([1, "30 дней", None, None, 3])
        worksheet.cell(row=1, column=3).number_format = '#,##0.00 "₽"'
        worksheet.cell(row=2, column=4).number_format = '#,##0.00 "₽"'

        headers, rows, column_count = ExportMixin._worksheet_to_visible_retrade_rows(
            worksheet
        )

        self.assertEqual(column_count, 3)
        self.assertEqual(
            [cell["value"] for cell in headers],
            ["№", "Срок поставщика", "Рейтинг"],
        )
        self.assertEqual([cell["value"] for cell in rows[0][1]], [1, "30 дней", 3])

    def test_parse_returns_empty_structure_for_empty_file(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": None, "currency": None}],
                [{"value": "", "currency": None}],
            ]
        )
        self.assertEqual(
            parsed,
            {
                "headers": [],
                "rows": [],
                "total_without_vat": None,
                "total_without_vat_currency": None,
                "totals": {"price": 0.0, "logistic": 0.0, "customs": 0.0},
                "totals_currency": {"price": None, "logistic": None, "customs": None},
            },
        )

    def test_parse_extracts_total_without_vat_and_removes_service_row(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": "Наименование", "currency": None}, {"value": "Сумма", "currency": None}],
                [{"value": 1, "currency": None}, {"value": 1250000, "currency": "RUB"}],
                [{"value": "Итого без НДС", "currency": None}, {"value": 1250000, "currency": "RUB"}],
            ]
        )
        self.assertEqual(
            parsed["headers"],
            ["Наименование"],
        )
        self.assertEqual(
            parsed["rows"],
            [[{"value": 1, "currency": None}]],
        )
        self.assertEqual(parsed["total_without_vat"], 1250000)
        self.assertEqual(parsed["total_without_vat_currency"], "RUB")
        self.assertEqual(parsed["totals"], {"price": 0.0, "logistic": 0.0, "customs": 0.0})
        self.assertEqual(
            parsed["totals_currency"],
            {"price": None, "logistic": None, "customs": None},
        )

    def test_parse_keeps_only_position_rows_by_first_cell(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [{"value": "№", "currency": None}, {"value": "Сумма", "currency": None}],
                [{"value": 1, "currency": None}, {"value": 1000, "currency": "RUB"}],
                [{"value": "2", "currency": None}, {"value": 2000, "currency": "RUB"}],
                [{"value": "Прибыль", "currency": None}, {"value": 500, "currency": "RUB"}],
                [{"value": "Итого", "currency": None}, {"value": 3000, "currency": "RUB"}],
            ]
        )
        self.assertEqual(
            parsed["headers"],
            ["№"],
        )
        self.assertEqual(
            parsed["rows"],
            [
                [{"value": 1, "currency": None}],
                [{"value": "2", "currency": None}],
            ],
        )
        self.assertIsNone(parsed["total_without_vat"])
        self.assertIsNone(parsed["total_without_vat_currency"])
        self.assertEqual(parsed["totals"], {"price": 0.0, "logistic": 0.0, "customs": 0.0})
        self.assertEqual(
            parsed["totals_currency"],
            {"price": None, "logistic": None, "customs": None},
        )

    def test_parse_filters_service_columns_and_calculates_price_total(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [
                    {"value": "№", "currency": None},
                    {"value": "Цена за ед. без НДС", "currency": None},
                    {"value": "Сумма", "currency": None},
                    {"value": "Прибыль", "currency": None},
                ],
                [
                    {"value": 1, "currency": None},
                    {"value": 1000, "currency": "RUB"},
                    {"value": 3000, "currency": "RUB"},
                    {"value": 100, "currency": "RUB"},
                ],
                [
                    {"value": "2", "currency": None},
                    {"value": 2500.5, "currency": "RUB"},
                    {"value": 7501.5, "currency": "RUB"},
                    {"value": 250, "currency": "RUB"},
                ],
                [
                    {"value": "Итого", "currency": None},
                    {"value": None, "currency": None},
                    {"value": 10501.5, "currency": "RUB"},
                    {"value": 350, "currency": "RUB"},
                ],
            ]
        )
        self.assertEqual(parsed["headers"], ["№", "Цена за ед. без НДС"])
        self.assertEqual(
            parsed["rows"],
            [
                [
                    {"value": 1, "currency": None},
                    {"value": 1000, "currency": "RUB"},
                ],
                [
                    {"value": "2", "currency": None},
                    {"value": 2500.5, "currency": "RUB"},
                ],
            ],
        )
        self.assertEqual(parsed["totals"]["price"], 3500.5)
        self.assertEqual(parsed["totals"]["logistic"], 0.0)
        self.assertEqual(parsed["totals"]["customs"], 0.0)
        self.assertEqual(parsed["totals_currency"]["price"], "RUB")
        self.assertIsNone(parsed["totals_currency"]["logistic"])
        self.assertIsNone(parsed["totals_currency"]["customs"])

    def test_parse_calculates_logistic_and_customs_totals(self):
        parsed = ExportMixin._parse_retrade_calculations(
            [
                [
                    {"value": "№", "currency": None},
                    {"value": "Цена за ед. без НДС", "currency": None},
                    {"value": "Логистика", "currency": None},
                    {"value": "Таможня", "currency": None},
                ],
                [
                    {"value": 1, "currency": None},
                    {"value": 1000, "currency": "RUB"},
                    {"value": 100, "currency": "RUB"},
                    {"value": 50, "currency": "RUB"},
                ],
                [
                    {"value": 2, "currency": None},
                    {"value": 2000, "currency": "RUB"},
                    {"value": 200, "currency": "RUB"},
                    {"value": 80, "currency": "RUB"},
                ],
            ]
        )

        self.assertEqual(parsed["totals"], {"price": 3000.0, "logistic": 300.0, "customs": 130.0})
        self.assertEqual(
            parsed["totals_currency"],
            {"price": "RUB", "logistic": "RUB", "customs": "RUB"},
        )

    def test_detect_currency_from_number_format(self):
        self.assertEqual(ExportMixin._detect_currency(1000, '#,##0.00 "₽"'), "RUB")
        self.assertEqual(ExportMixin._detect_currency(1000, '#,##0.00 "$"'), "USD")
        self.assertEqual(ExportMixin._detect_currency(1000, '#,##0.00 "EUR"'), "EUR")
        self.assertEqual(ExportMixin._detect_currency(1000, '#,##0.00 "¥"'), "CNY")

    def test_detect_currency_from_string_fallback(self):
        self.assertEqual(ExportMixin._detect_currency("1000 руб", "General"), "RUB")
        self.assertEqual(ExportMixin._detect_currency("1 000,00 ₽", "General"), "RUB")
        self.assertEqual(ExportMixin._detect_currency("Total $100", "General"), "USD")
        self.assertEqual(ExportMixin._detect_currency("Amount 50 eur", "General"), "EUR")
        self.assertEqual(ExportMixin._detect_currency("Amount 50 CNY", "General"), "CNY")
        self.assertEqual(ExportMixin._detect_currency("Amount 50 CYN", "General"), "CNY")

    def test_format_number_ru(self):
        self.assertEqual(ExportMixin._format_number_ru(100000), "100 000,00")
        self.assertEqual(ExportMixin._format_number_ru(1500.5), "1 500,50")
        self.assertEqual(ExportMixin._format_number_ru(100), "100,00")

    def test_format_cell_displays_currency_suffix(self):
        self.assertEqual(
            ExportMixin._format_retrade_calculations_cell_for_display(
                {"value": 100000, "currency": "RUB"}
            ),
            "100 000,00 ₽",
        )
        self.assertEqual(
            ExportMixin._format_retrade_calculations_cell_for_display(
                {"value": 5000, "currency": "USD"}
            ),
            "5 000,00 $",
        )
        self.assertEqual(
            ExportMixin._format_retrade_calculations_cell_for_display(
                {"value": 1234.5, "currency": "EUR"}
            ),
            "1 234,50 €",
        )
        self.assertEqual(
            ExportMixin._format_retrade_calculations_cell_for_display(
                {"value": 1234.5, "currency": "CNY"}
            ),
            "1 234,50 ¥",
        )

    def test_format_calculations_display_value_uses_cell_currency_for_prices(self):
        mixin = ExportMixin()

        self.assertEqual(
            mixin._format_calculations_display_value(
                1234.5,
                col_index=1,
                header="Цена за ед.",
                price_columns={1},
                rating_columns=set(),
                currency="CNY",
            ),
            "1 234,50 ¥",
        )

    def test_generated_retrade_rounding_columns_detects_right_appended_block(self):
        self.assertEqual(
            ExportMixin._generated_retrade_rounding_columns(
                [
                    "Наименование",
                    "Рейтинг",
                    "Рейтинг ЭТП",
                    "Лучшая цена за ед.",
                    "Разница",
                    "Наценка",
                ]
            ),
            {2, 3, 4, 5},
        )
        self.assertEqual(
            ExportMixin._generated_retrade_rounding_columns(
                [
                    "Наименование",
                    "Рейтинг (таблица)",
                    "Лучшая цена за ед.",
                    "Рейтинг",
                    "Скорректированный рейтинг",
                ]
            ),
            {1, 2, 3, 4},
        )
        self.assertEqual(
            ExportMixin._generated_retrade_rounding_columns(
                [
                    "Рейтинг",
                    "Лучшая цена за ед.",
                    "Наименование",
                    "Разница",
                    "Наценка",
                ]
            ),
            set(),
        )

    def test_rounding_checkbox_applies_only_to_generated_right_columns(self):
        mixin = ExportMixin()
        mixin.is_retrade_rounding_enabled = lambda: False

        self.assertEqual(
            mixin._format_calculations_display_value(
                "1,2345",
                col_index=1,
                header="Рейтинг",
                price_columns=set(),
                rating_columns={1, 4},
                rounding_columns={3, 4, 5, 6},
            ),
            "1.23",
        )
        self.assertEqual(
            mixin._format_calculations_display_value(
                "1,2345",
                col_index=4,
                header="Рейтинг",
                price_columns=set(),
                rating_columns={1, 4},
                rounding_columns={3, 4, 5, 6},
            ),
            "1.2345",
        )
        self.assertEqual(
            mixin._format_calculations_display_value(
                827753.714,
                col_index=2,
                header="Цена за ед.",
                price_columns={2, 3},
                rating_columns=set(),
                rounding_columns={3, 4, 5, 6},
            ),
            "827 753,71",
        )
        self.assertEqual(
            mixin._format_calculations_display_value(
                827753.714,
                col_index=3,
                header="Лучшая цена за ед.",
                price_columns={2, 3},
                rating_columns=set(),
                rounding_columns={3, 4, 5, 6},
            ),
            "827753,714",
        )

    def test_format_cell_returns_non_numeric_as_is(self):
        self.assertEqual(
            ExportMixin._format_retrade_calculations_cell_for_display(
                {"value": "N/A", "currency": "RUB"}
            ),
            "N/A",
        )

    def test_parse_number_accepts_currency_and_decimal_comma(self):
        self.assertEqual(ExportMixin.parse_number("30 033,85 ₽"), 30033.85)
        self.assertEqual(ExportMixin.parse_number("30 033,85 CYN"), 30033.85)
        self.assertEqual(ExportMixin.parse_number("1,25"), 1.25)
        self.assertEqual(ExportMixin.parse_number(None), 0.0)

    def test_write_retrade_table_to_worksheet_preserves_formulas_and_blanks(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["A", "B", "C"])
        worksheet.cell(row=2, column=1).value = "old"
        worksheet.cell(row=2, column=2).value = "=SUM(A2:A3)"
        worksheet.cell(row=2, column=3).value = "keep"
        worksheet.cell(row=3, column=1).value = "old text"
        worksheet.cell(row=3, column=2).value = "keep blank"

        table = _FakeEditableTable(
            [
                [10, 20, ""],
                ["new text", None, "=A1"],
            ]
        )

        mixin = ExportMixin()
        written_count = mixin._write_retrade_table_to_worksheet(
            worksheet,
            table,
        )

        self.assertEqual(written_count, 2)
        self.assertEqual(worksheet.cell(row=2, column=1).value, 10.0)
        self.assertEqual(worksheet.cell(row=2, column=2).value, "=SUM(A2:A3)")
        self.assertEqual(worksheet.cell(row=2, column=3).value, "keep")
        self.assertEqual(worksheet.cell(row=3, column=1).value, "new text")
        self.assertEqual(worksheet.cell(row=3, column=2).value, "keep blank")
        self.assertIsNone(worksheet.cell(row=3, column=3).value)

    def test_write_retrade_table_to_worksheet_saves_money_text_as_number(self):
        workbook = Workbook()
        worksheet = workbook.active
        headers = ["Предлагаемая цена за ед.", "Кол-во", "Сумма, CNY"]
        worksheet.append(headers)
        table = _FakeEditableTable(
            [["100 000,00", "2", "200 000,00"]],
            headers=headers,
        )

        mixin = ExportMixin()
        written_count = mixin._write_retrade_table_to_worksheet(
            worksheet,
            table,
        )

        self.assertEqual(written_count, 3)
        self.assertEqual(worksheet.cell(row=2, column=1).value, 100000.0)
        self.assertEqual(worksheet.cell(row=2, column=2).value, 2.0)
        self.assertEqual(worksheet.cell(row=2, column=3).value, 200000.0)
        self.assertEqual(
            worksheet.cell(row=2, column=1).number_format,
            mixin._currency_format("¥"),
        )
        self.assertEqual(worksheet.cell(row=2, column=2).number_format, "General")
        self.assertEqual(
            worksheet.cell(row=2, column=3).number_format,
            mixin._currency_format("¥"),
        )

    def test_write_retrade_table_to_worksheet_formats_money_formulas(self):
        workbook = Workbook()
        worksheet = workbook.active
        headers = ["Предлагаемая цена за ед.", "Кол-во", "Сумма, USD"]
        worksheet.append(headers)
        worksheet.cell(row=2, column=1).value = "=A3"
        worksheet.cell(row=2, column=3).value = "=A2*B2"
        table = _FakeEditableTable(
            [["100 000,00", "2", "200 000,00"]],
            headers=headers,
        )

        mixin = ExportMixin()
        written_count = mixin._write_retrade_table_to_worksheet(
            worksheet,
            table,
        )

        self.assertEqual(written_count, 1)
        self.assertEqual(worksheet.cell(row=2, column=1).value, "=A3")
        self.assertEqual(worksheet.cell(row=2, column=3).value, "=A2*B2")
        self.assertEqual(
            worksheet.cell(row=2, column=1).number_format,
            mixin._currency_format("$"),
        )
        self.assertEqual(
            worksheet.cell(row=2, column=3).number_format,
            mixin._currency_format("$"),
        )

    def test_collect_missing_retrade_required_cells(self):
        headers = [
            "Наименование",
            "Срок поставки",
            "Производитель",
            "Технические характеристики",
            "Условия гарантий качества",
        ]
        table = _FakeEditableTable(
            [
                ["Позиция 1", "30 дней", "Atlas Copco", "Описание", "6 мес"],
                ["Позиция 2", "", "Atlas Copco", "", "6 мес"],
                ["", "", "", "", ""],
            ],
            headers=headers,
        )

        missing = ExportMixin._collect_missing_retrade_required_cells(table)

        self.assertEqual(
            missing,
            [
                {"row": 1, "column": 1, "label": "Срок поставки"},
                {"row": 1, "column": 3, "label": "Технические характеристики"},
            ],
        )

    def test_retrade_main_row_has_position_ignores_default_manual_fields(self):
        headers = [
            "Наименование",
            "Срок поставки",
            "Производитель",
            "Условия гарантий качества",
        ]
        table = _FakeEditableTable(
            [
                ["Позиция 1", "", "", "6 мес"],
                ["", "", "", "6 мес"],
            ],
            headers=headers,
        )
        columns = ExportMixin._get_retrade_main_columns(headers)
        position_columns = ExportMixin._retrade_main_position_columns(
            headers,
            columns,
        )

        self.assertTrue(
            ExportMixin._retrade_main_row_has_position(table, 0, position_columns)
        )
        self.assertFalse(
            ExportMixin._retrade_main_row_has_position(table, 1, position_columns)
        )

    def test_get_table_rows_uses_sale_price_and_default_producer(self):
        headers = [
            "№",
            "Наименование",
            "Каталожный номер",
            "Ед. изм.",
            "Кол-во",
            "Цена за ед. без НДС",
            "Итого без НДС",
            "Логистика",
            "Таможня",
            "Цена за ед",
            "Цена реализации за ед. без НДС",
            "Итого реализации без НДС",
            "Итого реализации с НДС",
            "Срок поставки",
            "Срок поставщика",
        ]
        table = _FakeEditableTable(
            [["1", "Насос", "P-1", "шт", 2, 100, 200, "", "", "", 150, 300, 360, "30 дней", "20 дней"]],
            headers=headers,
        )
        window = ExportMixin()
        window.ui = SimpleNamespace(KpTable=table)
        window.tableData = {"currency": ["¥"]}

        rows = window.get_table_rows(default_manufacturer="Завод")

        self.assertEqual(rows[0]["name"], "Насос")
        self.assertEqual(rows[0]["price"], 150.0)
        self.assertEqual(rows[0]["total"], 300.0)
        self.assertEqual(rows[0]["currency"], "CNY")
        self.assertEqual(rows[0]["manufacturer"], "Завод")
        self.assertEqual(rows[0]["technical_characteristics"], "Завод")

    def test_get_table_rows_applies_submission_defaults(self):
        headers = [
            "№",
            "Наименование",
            "Каталожный номер",
            "Ед. изм.",
            "Кол-во",
            "Цена за ед. без НДС",
            "Итого без НДС",
            "Логистика",
            "Таможня",
            "Цена за ед",
            "Цена реализации за ед. без НДС",
            "Итого реализации без НДС",
            "Итого реализации с НДС",
            "Срок поставки",
            "Срок поставщика",
        ]
        table = _FakeEditableTable(
            [["1", "Насос", "P-1", "шт", 2, 100, 200, "", "", "", 150, 300, 360, "30 дней", "20 дней"]],
            headers=headers,
        )
        window = ExportMixin()
        window.ui = SimpleNamespace(KpTable=table)

        rows = window.get_table_rows(
            default_supplier_status="Посредник",
            default_warranty="12 мес.",
        )

        self.assertEqual(rows[0]["supplier_status"], "Посредник")
        self.assertEqual(rows[0]["warranty"], "12 мес.")

    def test_get_table_rows_leaves_status_and_warranty_empty_for_zero_rows(self):
        headers = [
            "№",
            "Наименование",
            "Каталожный номер",
            "Ед. изм.",
            "Кол-во",
            "Цена за ед. без НДС",
            "Итого без НДС",
            "Логистика",
            "Таможня",
            "Цена за ед",
            "Цена реализации за ед. без НДС",
            "Итого реализации без НДС",
            "Итого реализации с НДС",
            "Срок поставки",
            "Срок поставщика",
        ]
        table = _FakeEditableTable(
            [
                [
                    "1",
                    "Насос",
                    "P-1",
                    "шт",
                    32,
                    0,
                    0,
                    "",
                    "",
                    "",
                    "",
                    "0,00",
                    "0,00",
                    "30 дней",
                    "20 дней",
                ],
                [
                    "2",
                    "Клапан",
                    "V-1",
                    "шт",
                    1,
                    "0,00",
                    "0,00",
                    "",
                    "",
                    "",
                    "#ДЕЛ/0!",
                    "#ДЕЛ/0!",
                    "#ДЕЛ/0!",
                    "30 дней",
                    "20 дней",
                ],
            ],
            headers=headers,
        )
        window = ExportMixin()
        window.ui = SimpleNamespace(KpTable=table)

        rows = window.get_table_rows(
            default_supplier_status="Посредник",
            default_warranty="12 мес.",
        )

        self.assertEqual(rows[0]["price"], "")
        self.assertEqual(rows[0]["total"], 0.0)
        self.assertEqual(rows[0]["supplier_status"], "")
        self.assertEqual(rows[0]["warranty"], "")
        self.assertEqual(rows[0]["guarantee"], "")
        self.assertEqual(rows[1]["supplier_status"], "")
        self.assertEqual(rows[1]["warranty"], "")
        self.assertEqual(rows[1]["guarantee"], "")

    def test_prepare_submission_export_for_loading_keeps_table_currency_metadata(self):
        class _FakeExcelProcessor:
            def __init__(self):
                self.filled_rows = None

            def can_fill_exported_excel(self, _file_path):
                return True

            def fill_exported_excel(self, _file_path, source_rows, **_kwargs):
                self.filled_rows = source_rows

        processor = _FakeExcelProcessor()
        window = ExportMixin()
        window.excel_processor = processor
        window._pending_submission_export_metadata = {"number": "REQ-1"}
        window.get_table_rows = lambda **_kwargs: [
            {"name": "Насос", "price": "150 ¥", "currency": "CNY"}
        ]

        window._prepare_submission_export_for_loading("submission.xlsx")

        self.assertEqual(window._pending_submission_export_metadata["currency"], "CNY")
        self.assertEqual(processor.filled_rows[0]["currency"], "CNY")

    def test_format_missing_retrade_required_cells_message(self):
        message = ExportMixin._format_missing_retrade_required_cells_message(
            [
                {"row": 1, "column": 1, "label": "Срок поставки"},
                {"row": 1, "column": 2, "label": "Производитель"},
            ]
        )

        self.assertIn("строка 2: Срок поставки", message)
        self.assertIn("строка 2: Производитель", message)
        self.assertIn("Пропустить", message)

    def test_recalculate_retrade_main_row_values_updates_total(self):
        headers = ["Предлагаемая цена за ед.", "Кол-во", "Сумма"]
        columns = ExportMixin._get_retrade_main_columns(headers)

        row = ExportMixin._recalculate_retrade_main_row_values(
            ["100 000,00", "2", None],
            columns,
        )

        self.assertEqual(row, ["100 000,00", "2", 200000.0])
        self.assertEqual(ExportMixin.format_money(row[2]), "200 000,00")

    def test_format_rating_always_uses_two_decimals(self):
        mixin = ExportMixin()

        self.assertEqual(mixin.format_rating("1,234"), "1.23")
        self.assertEqual(
            mixin._format_calculations_display_value(
                "1,2",
                col_index=3,
                header="Рейтинг",
                price_columns=set(),
                rating_columns={3},
            ),
            "1.20",
        )

    def test_calculate_updated_position_prices_uses_markup_column(self):
        updates, sale_price_col = ExportMixin._calculate_updated_position_prices(
            [
                "№",
                "Цена за ед. без НДС",
                "Наценка",
                "Цена реализации за ед. без НДС",
            ],
            [
                [
                    {"value": 1, "currency": None},
                    {"value": "30 033,85 ₽", "currency": "RUB"},
                    {"value": "1,25", "currency": None},
                    {"value": None, "currency": "RUB"},
                ],
                [
                    {"value": 2, "currency": None},
                    {"value": "", "currency": None},
                    {"value": "1,10", "currency": None},
                    {"value": None, "currency": None},
                ],
                [
                    {"value": 3, "currency": None},
                    {"value": "0,00 ₽", "currency": "RUB"},
                    {"value": "1,50", "currency": None},
                    {"value": None, "currency": "RUB"},
                ],
            ],
        )

        self.assertEqual(sale_price_col, 3)
        self.assertEqual(
            updates,
            [{"row": 0, "value": 37542.31, "currency": "RUB"}],
        )

    def test_rounding_helpers_use_requested_digits_or_skip_rounding(self):
        self.assertEqual(
            ExportMixin._excel_formula("A1/B1", rounding_digits=0),
            "=ROUND(A1/B1, 0)",
        )
        self.assertEqual(
            ExportMixin._excel_formula("A1/B1", rounding_digits=None),
            "=A1/B1",
        )
        self.assertEqual(
            ExportMixin._truncate_excel_formula("A1/B1", 2),
            "=TRUNC(A1/B1, 2)",
        )
        self.assertEqual(ExportMixin._excel_trunc(1.239, 2), 1.23)
        self.assertEqual(
            ExportMixin._round_generated_value(1.236, rounding_digits=1),
            1.2,
        )
        self.assertEqual(
            ExportMixin._round_generated_value(1.236, rounding_digits=None),
            1.236,
        )

    def test_calculate_updated_position_prices_prefers_excel_j_and_s(self):
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[5] = "Цена за ед. без НДС"
        headers[9] = "Колонка J"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[18] = "Скорректированный рейтинг"
        row = [{"value": None, "currency": None} for _ in headers]
        row[5] = {"value": 999, "currency": "RUB"}
        row[9] = {"value": 100, "currency": "RUB"}
        row[18] = {"value": "1,25", "currency": None}

        updates, sale_price_col = ExportMixin._calculate_updated_position_prices(
            headers,
            [row],
        )

        self.assertEqual(sale_price_col, 10)
        self.assertEqual(updates, [{"row": 0, "value": 125.0, "currency": "RUB"}])

    def test_calculate_updated_position_prices_uses_excel_columns_with_offset(self):
        headers = [""] + [f"Колонка {index}" for index in range(1, 20)]
        headers[6] = "Цена за ед. без НДС"
        headers[11] = "Цена реализации за ед. без НДС"
        headers[19] = "Скорректированный рейтинг"
        row = [{"value": None, "currency": None} for _ in headers]
        row[6] = {"value": 999, "currency": "RUB"}
        row[10] = {"value": 100, "currency": "RUB"}
        row[19] = {"value": "1,25", "currency": None}

        updates, sale_price_col = ExportMixin._calculate_updated_position_prices(
            headers,
            [row],
            column_offset=1,
        )

        self.assertEqual(sale_price_col, 11)
        self.assertEqual(updates, [{"row": 0, "value": 125.0, "currency": "RUB"}])

    def test_calculate_updated_position_prices_detects_currency_from_display_text(self):
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[9] = "Цена за ед. без НДС"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[18] = "Скорректированный рейтинг"
        row = [{"value": None, "currency": None} for _ in headers]
        row[9] = {"value": "100,00 ₽", "currency": None}
        row[18] = {"value": "1,25", "currency": None}

        updates, sale_price_col = ExportMixin._calculate_updated_position_prices(
            headers,
            [row],
        )

        self.assertEqual(sale_price_col, 10)
        self.assertEqual(updates, [{"row": 0, "value": 125.0, "currency": "RUB"}])

    def test_calculate_updated_position_prices_requires_columns(self):
        with self.assertRaises(ValueError) as context:
            ExportMixin._calculate_updated_position_prices(["№"], [])

        message = str(context.exception)
        self.assertIn("Цена за ед. без НДС", message)
        self.assertIn("Наценка", message)

    def test_formula_update_row_indices_do_not_require_rating_value(self):
        row_indices = ExportMixin._formula_update_row_indices(
            [
                "№",
                "Цена за ед. без НДС",
                "Скорректированный рейтинг",
                "Цена реализации за ед. без НДС",
            ],
            [
                [
                    {"value": 1, "currency": None},
                    {"value": 81, "currency": "RUB"},
                    {"value": None, "currency": None},
                    {"value": None, "currency": "RUB"},
                ],
                [
                    {"value": 2, "currency": None},
                    {"value": "", "currency": None},
                    {"value": None, "currency": None},
                    {"value": None, "currency": "RUB"},
                ],
                [
                    {"value": 3, "currency": None},
                    {"value": 0, "currency": "RUB"},
                    {"value": None, "currency": None},
                    {"value": None, "currency": "RUB"},
                ],
            ],
        )

        self.assertEqual(row_indices, [0])

    def test_write_realization_price_formulas_to_sheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[5] = "Цена за ед. без НДС"
        headers[9] = "Колонка J"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[18] = "Скорректированный рейтинг"
        worksheet.append(headers)
        worksheet.append([None for _ in headers])
        worksheet.append([None for _ in headers])
        worksheet.append([None for _ in headers])
        worksheet.cell(row=2, column=10).number_format = '#,##0.00 "₽"'
        worksheet.cell(row=3, column=10).value = 0
        worksheet.cell(row=3, column=11).value = "=OLD"

        formulas = ExportMixin._write_realization_price_formulas_to_sheet(
            worksheet,
            [0, 1, 2],
        )

        self.assertEqual(
            formulas,
            {
                0: "=ROUND(J2*S2, 2)",
                2: "=ROUND(J4*S4, 2)",
            },
        )
        self.assertEqual(worksheet.cell(row=2, column=11).value, "=ROUND(J2*S2, 2)")
        self.assertEqual(
            worksheet.cell(row=2, column=11).number_format,
            ExportMixin._currency_format("₽"),
        )
        self.assertEqual(worksheet.cell(row=3, column=11).value, "=OLD")
        self.assertEqual(worksheet.cell(row=4, column=11).value, "=ROUND(J4*S4, 2)")
        self.assertEqual(
            worksheet.cell(row=4, column=11).number_format,
            ExportMixin._currency_format(None),
        )

    def test_is_zero_retrade_price_cell_uses_cached_value_sheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=2, column=10).value = "=A2"

        values_workbook = Workbook()
        values_worksheet = values_workbook.active
        values_worksheet.cell(row=2, column=10).value = 0

        self.assertTrue(
            ExportMixin._is_zero_retrade_price_cell(
                worksheet,
                2,
                10,
                values_worksheet=values_worksheet,
            )
        )

    def test_write_realization_price_formulas_to_sheet_uses_first_non_empty_header_row(
        self,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[9] = "Цена за ед. без НДС"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[18] = "Скорректированный рейтинг"
        worksheet.append([None for _ in headers])
        worksheet.append([None for _ in headers])
        worksheet.append(headers)
        worksheet.append([None for _ in headers])

        formulas = ExportMixin._write_realization_price_formulas_to_sheet(
            worksheet,
            [0],
        )

        self.assertEqual(formulas, {0: "=ROUND(J4*S4, 2)"})
        self.assertEqual(
            worksheet.cell(row=4, column=11).value,
            "=ROUND(J4*S4, 2)",
        )

    def test_write_realization_price_formulas_to_sheet_accepts_excel_row_indices(self):
        workbook = Workbook()
        worksheet = workbook.active
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[9] = "Цена за ед. без НДС"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[18] = "Скорректированный рейтинг"
        worksheet.append([None for _ in headers])
        worksheet.append(headers)
        worksheet.append([None for _ in headers])
        worksheet.append([None for _ in headers])

        formulas = ExportMixin._write_realization_price_formulas_to_sheet(
            worksheet,
            [4],
            indices_are_excel_rows=True,
        )

        self.assertEqual(formulas, {4: "=ROUND(J4*S4, 2)"})
        self.assertEqual(
            worksheet.cell(row=4, column=11).value,
            "=ROUND(J4*S4, 2)",
        )

    def test_write_update_position_formulas_to_current_file_saves_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "calc.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Рассчеты"
            headers = [f"Колонка {index}" for index in range(1, 20)]
            headers[9] = "Цена за ед. без НДС"
            headers[10] = "Цена реализации за ед. без НДС"
            headers[18] = "Скорректированный рейтинг"
            worksheet.append(headers)
            worksheet.append([None for _ in headers])
            worksheet.append([None for _ in headers])
            workbook.save(file_path)

            mixin = ExportMixin()
            mixin.calculations_file_path = str(file_path)
            mixin.current_calculations_sheet_name = "Рассчеты"

            formulas = mixin._write_update_position_formulas_to_current_calculations_file(
                [0, 1]
            )

            self.assertEqual(
                formulas,
                {
                    0: "=ROUND(J2*S2, 2)",
                    1: "=ROUND(J3*S3, 2)",
                },
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook["Рассчеты"]
                self.assertEqual(
                    result_sheet.cell(row=2, column=11).value,
                    "=ROUND(J2*S2, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=11).value,
                    "=ROUND(J3*S3, 2)",
                )
                self.assertEqual(result_workbook.calculation.calcMode, "auto")
                self.assertTrue(result_workbook.calculation.fullCalcOnLoad)
                self.assertTrue(result_workbook.calculation.forceFullCalc)
            finally:
                result_workbook.close()

    def test_extract_best_prices_uses_header_and_parses_currency(self):
        table = _FakeRetradeTable(
            ["Наименование", "Лучшая цена за ед."],
            [
                ["Двигатель", "102 188,50 ¥"],
                ["Насос", ""],
                ["Клапан", "нет цены"],
            ],
        )

        self.assertEqual(
            ExportMixin._extract_retrade_best_prices(table),
            [102188.5, None, None],
        )

    def test_extract_ratings_and_best_prices_uses_headers(self):
        table = _FakeRetradeTable(
            ["Наименование", "Рейтинг", "Лучшая цена за ед."],
            [
                ["Двигатель", "1,25", "102 188,50 ₽"],
                ["Насос", "", ""],
                ["Клапан", "нет рейтинга", "нет цены"],
            ],
        )

        self.assertEqual(
            ExportMixin._extract_retrade_ratings_and_best_prices(table),
            ([1.25, None, None], [102188.5, None, None]),
        )

    def test_extract_retrade_proposal_prices_uses_main_table_price_column(self):
        table = _FakeEditableTable(
            [["Насос", "777,77 ₽", "2"], ["Клапан", None, "3"]],
            headers=["Наименование", "Предлагаемая цена за ед.", "Кол-во"],
        )

        self.assertEqual(
            ExportMixin._extract_retrade_proposal_prices(table),
            [
                {"value": 777.77, "currency": "₽"},
                {"value": None, "currency": None},
            ],
        )

    def test_next_retrade_sheet_title_uses_next_number(self):
        workbook = Workbook()
        workbook.active.title = "Sheet"
        workbook.create_sheet("Переторжка 1")
        workbook.create_sheet("Переторжка 2")

        self.assertEqual(
            ExportMixin._next_retrade_sheet_title(workbook),
            "Переторжка 3",
        )

    def test_populate_retrade_sheets_list_selects_requested_sheet(self):
        class _FakeListWidgetItem:
            def __init__(self, text):
                self._text = str(text)

            def text(self):
                return self._text

            def setBackground(self, *_args, **_kwargs):
                pass

        class _FakeSheetsList:
            def __init__(self):
                self.items = []
                self.current_row = -1
                self.signals_blocked = False

            def blockSignals(self, value):
                self.signals_blocked = bool(value)

            def clear(self):
                self.items.clear()
                self.current_row = -1

            def addItem(self, item):
                self.items.append(item)

            def count(self):
                return len(self.items)

            def setCurrentRow(self, row):
                self.current_row = int(row)

            def currentItem(self):
                if self.current_row < 0 or self.current_row >= len(self.items):
                    return None
                return self.items[self.current_row]

        original_list_widget = export_mixin_module.QListWidget
        original_list_widget_item = export_mixin_module.QListWidgetItem
        original_color = export_mixin_module.QColor
        export_mixin_module.QListWidget = _FakeSheetsList
        export_mixin_module.QListWidgetItem = _FakeListWidgetItem
        export_mixin_module.QColor = lambda *_args, **_kwargs: object()
        try:
            mixin = ExportMixin()
            mixin.workbook = SimpleNamespace(
                sheetnames=["Рассчеты", "Переторжка 1", "Переторжка 2"]
            )
            mixin.sheetsList = _FakeSheetsList()
            selected_sheets = []
            mixin.on_sheet_selected = selected_sheets.append

            mixin._populate_retrade_sheets_list(
                selected_sheet_name="Переторжка 2"
            )

            self.assertEqual(mixin.sheetsList.current_row, 2)
            self.assertEqual(selected_sheets, ["Переторжка 2"])
            self.assertFalse(mixin.sheetsList.signals_blocked)
        finally:
            export_mixin_module.QListWidget = original_list_widget
            export_mixin_module.QListWidgetItem = original_list_widget_item
            export_mixin_module.QColor = original_color

    def test_load_retrade_calculations_workbook_evaluates_uncached_formulas(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Рассчеты"
        headers = [f"Колонка {index}" for index in range(1, 20)]
        headers[9] = "Цена за ед. без НДС"
        headers[10] = "Цена реализации за ед. без НДС"
        headers[16] = "Лучшая цена за ед."
        headers[17] = "Рейтинг"
        headers[18] = "Скорректированный рейтинг"
        worksheet.append(headers)
        worksheet.cell(row=2, column=10).value = 100
        worksheet.cell(row=2, column=17).value = 125
        worksheet.cell(row=2, column=18).value = "=TRUNC(Q2/J2, 2)"
        worksheet.cell(row=2, column=19).value = (
            "=TRUNC(IF(R2-0.02<1.15,1.15,R2-0.02), 2)"
        )
        worksheet.cell(row=2, column=11).value = "=ROUND(J2*S2, 2)"
        worksheet.cell(row=3, column=10).value = "=SUM(J2:K2)"

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "calc.xlsx"
            workbook.save(file_path)
            workbook.close()

            loaded_workbook = ExportMixin._load_retrade_calculations_workbook(
                str(file_path)
            )
            try:
                loaded_sheet = loaded_workbook["Рассчеты"]
                self.assertEqual(loaded_sheet.cell(row=2, column=18).value, 1.25)
                self.assertEqual(loaded_sheet.cell(row=2, column=19).value, 1.23)
                self.assertEqual(loaded_sheet.cell(row=2, column=11).value, 123.0)
                self.assertEqual(loaded_sheet.cell(row=3, column=10).value, 223.0)
            finally:
                loaded_workbook.close()

    def test_write_best_prices_to_calculations_file_creates_updated_sheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        first_row = [None for _ in headers]
        first_row[0] = "Двигатель"
        first_row[9] = 100
        second_row = [None for _ in headers]
        second_row[0] = "Насос"
        second_row[9] = 200
        worksheet.append(headers)
        worksheet.append(first_row)
        worksheet.append(second_row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [102188.5, None, 3000.0],
                ratings=[1.25, None, 2.0],
                min_margin=1.15,
                delta_percent=2,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                self.assertEqual(sheet_title, "Переторжка 1")
                self.assertIn(sheet_title, result_workbook.sheetnames)
                result_sheet = result_workbook[sheet_title]
                real_rating_col_index = 16
                best_price_col_index = 17
                formula_col_index = 18
                corrected_rating_col_index = 19
                realization_price_col_index = 11

                self.assertEqual(
                    result_sheet.cell(row=1, column=real_rating_col_index).value,
                    "Рейтинг ЭТП",
                )
                self.assertEqual(
                    result_sheet.cell(row=1, column=best_price_col_index).value,
                    "Лучшая цена за ед.",
                )
                self.assertEqual(
                    result_sheet.cell(row=1, column=formula_col_index).value,
                    "Разница",
                )
                self.assertEqual(
                    result_sheet.cell(
                        row=1,
                        column=corrected_rating_col_index,
                    ).value,
                    "Наценка",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=best_price_col_index).value,
                    102188.5,
                )
                self.assertIsNone(
                    result_sheet.cell(row=3, column=best_price_col_index).value
                )
                self.assertEqual(
                    result_sheet.cell(row=4, column=best_price_col_index).value,
                    3000,
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=real_rating_col_index).value,
                    1.25,
                )
                self.assertIsNone(
                    result_sheet.cell(row=3, column=real_rating_col_index).value
                )
                self.assertEqual(
                    result_sheet.cell(row=4, column=real_rating_col_index).value,
                    2,
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=formula_col_index).value,
                    "=TRUNC(Q2/J2, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=formula_col_index).value,
                    "=TRUNC(Q3/J3, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=4, column=formula_col_index).value,
                    "=TRUNC(Q4/J4, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=corrected_rating_col_index).value,
                    "=TRUNC(IF(R2-0.02<1.15,1.15,R2-0.02), 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=corrected_rating_col_index).value,
                    "=TRUNC(IF(R3-0.02<1.15,1.15,R3-0.02), 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=4, column=corrected_rating_col_index).value,
                    "=TRUNC(IF(R4-0.02<1.15,1.15,R4-0.02), 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=realization_price_col_index).value,
                    "=ROUND(J2*S2, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=realization_price_col_index).value,
                    "=ROUND(J3*S3, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=4, column=realization_price_col_index).value,
                    "=ROUND(J4*S4, 2)",
                )
                self.assertEqual(
                    result_sheet.column_dimensions[get_column_letter(formula_col_index)].width,
                    18,
                )
                self.assertEqual(
                    result_sheet.column_dimensions[
                        get_column_letter(corrected_rating_col_index)
                    ].width,
                    26,
                )
                self.assertEqual(
                    result_sheet.column_dimensions[get_column_letter(best_price_col_index)].width,
                    18,
                )
                self.assertEqual(
                    result_sheet.column_dimensions[get_column_letter(real_rating_col_index)].width,
                    18,
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_to_calculations_file_can_skip_rounding(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        row = [None for _ in headers]
        row[9] = 100
        worksheet.append(headers)
        worksheet.append(row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [125.6789],
                ratings=[1.2345],
                min_margin=1.15,
                delta_percent=2,
                rounding_digits=None,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                self.assertEqual(result_sheet.cell(row=2, column=16).value, 1.2345)
                self.assertEqual(result_sheet.cell(row=2, column=17).value, 125.6789)
                self.assertEqual(
                    result_sheet.cell(row=2, column=18).value,
                    "=TRUNC(Q2/J2, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=19).value,
                    "=TRUNC(IF(R2-0.02<1.15,1.15,R2-0.02), 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=11).value,
                    "=ROUND(J2*S2, 2)",
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_can_disable_difference_markup_truncation(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        row = [None for _ in headers]
        row[9] = 100
        worksheet.append(headers)
        worksheet.append(row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [125.6789],
                ratings=[1.2345],
                min_margin=1.15,
                delta_percent=2,
                rounding_digits=3,
                truncate_difference_markup=False,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                self.assertEqual(
                    result_sheet.cell(row=2, column=18).value,
                    "=ROUND(Q2/J2, 3)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=19).value,
                    "=ROUND(IF(R2-0.02<1.15,1.15,R2-0.02), 3)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=11).value,
                    "=ROUND(J2*S2, 2)",
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_can_keep_rating_one_rows_unchanged(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        first_row = [None for _ in headers]
        first_row[9] = 100
        first_row[10] = "=OLD"
        second_row = [None for _ in headers]
        second_row[9] = 200
        second_row[10] = "=OLD2"
        worksheet.append(headers)
        worksheet.append(first_row)
        worksheet.append(second_row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [125, 300],
                ratings=[1, 1.5],
                min_margin=1.15,
                delta_percent=2,
                skip_rating_one_positions=True,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                real_rating_col_index = 16
                best_price_col_index = 17
                formula_col_index = 18
                corrected_rating_col_index = 19
                realization_price_col_index = 11

                self.assertEqual(
                    result_sheet.cell(row=2, column=real_rating_col_index).value,
                    1,
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=best_price_col_index).value
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=formula_col_index).value
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=corrected_rating_col_index).value
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=realization_price_col_index).value,
                    "=OLD",
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=best_price_col_index).value,
                    300,
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=realization_price_col_index).value,
                    "=ROUND(J3*S3, 2)",
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_uses_table_price_for_rating_one_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        first_row = [None for _ in headers]
        first_row[9] = 100
        first_row[10] = 111
        second_row = [None for _ in headers]
        second_row[9] = 200
        second_row[10] = 222
        worksheet.append(headers)
        worksheet.append(first_row)
        worksheet.append(second_row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [125, 300],
                ratings=[1, 1.5],
                min_margin=1.15,
                delta_percent=2,
                skip_rating_one_positions=True,
                rating_one_prices=[
                    {"value": "555,55 ₽", "currency": "RUB"},
                    {"value": 777, "currency": "RUB"},
                ],
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                real_rating_col_index = 16
                best_price_col_index = 17
                formula_col_index = 18
                corrected_rating_col_index = 19
                realization_price_col_index = 11

                self.assertEqual(
                    result_sheet.cell(row=2, column=real_rating_col_index).value,
                    1,
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=best_price_col_index).value
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=formula_col_index).value
                )
                self.assertIsNone(
                    result_sheet.cell(row=2, column=corrected_rating_col_index).value
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=realization_price_col_index).value,
                    555.55,
                )
                self.assertEqual(
                    result_sheet.cell(
                        row=2,
                        column=realization_price_col_index,
                    ).number_format,
                    ExportMixin._currency_format("₽"),
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=best_price_col_index).value,
                    300,
                )
                self.assertEqual(
                    result_sheet.cell(row=3, column=realization_price_col_index).value,
                    "=ROUND(J3*S3, 2)",
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_ignores_blank_formatted_columns(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        worksheet.append(headers)
        worksheet.append([1, None, None, None, None, None, None, None, None, 100, None])
        worksheet.cell(row=1, column=26).number_format = '#,##0.00 "₽"'
        worksheet.cell(row=2, column=26).number_format = '#,##0.00 "₽"'

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [125],
                ratings=[1.25],
                min_margin=1.15,
                delta_percent=2,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                self.assertEqual(result_sheet.max_column, 26)
                self.assertEqual(
                    result_sheet.cell(row=1, column=16).value,
                    "Рейтинг ЭТП",
                )
                self.assertEqual(
                    result_sheet.cell(row=1, column=19).value,
                    "Наценка",
                )
                self.assertIsNone(result_sheet.cell(row=1, column=20).value)
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)

    def test_write_best_prices_to_calculations_file_skips_zero_source_price(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расчет"
        headers = [f"Колонка {index}" for index in range(1, 16)]
        headers[9] = "Цена за ед."
        headers[10] = "Цена реализации за ед. без НДС"
        first_row = [None for _ in headers]
        first_row[9] = 100
        zero_price_row = [None for _ in headers]
        zero_price_row[9] = 0
        zero_price_row[10] = "=OLD"
        worksheet.append(headers)
        worksheet.append(first_row)
        worksheet.append(zero_price_row)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        file_path = temp_file.name

        try:
            workbook.save(file_path)
            workbook.close()

            sheet_title = ExportMixin._write_best_prices_to_calculations_file(
                file_path,
                [1000, 500],
                ratings=[1.236, 1.236],
                min_margin=1.15,
                delta_percent=2,
            )

            result_workbook = load_workbook(file_path, data_only=False)
            try:
                result_sheet = result_workbook[sheet_title]
                real_rating_col_index = 16
                best_price_col_index = 17
                formula_col_index = 18
                corrected_rating_col_index = 19
                realization_price_col_index = 11

                self.assertEqual(
                    result_sheet.cell(row=2, column=real_rating_col_index).value,
                    1.24,
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=best_price_col_index).value,
                    1000,
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=formula_col_index).value,
                    "=TRUNC(Q2/J2, 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=corrected_rating_col_index).value,
                    "=TRUNC(IF(R2-0.02<1.15,1.15,R2-0.02), 2)",
                )
                self.assertEqual(
                    result_sheet.cell(row=2, column=realization_price_col_index).value,
                    "=ROUND(J2*S2, 2)",
                )
                for column in (
                    real_rating_col_index,
                    best_price_col_index,
                    formula_col_index,
                    corrected_rating_col_index,
                ):
                    self.assertIsNone(result_sheet.cell(row=3, column=column).value)
                self.assertEqual(
                    result_sheet.cell(row=3, column=realization_price_col_index).value,
                    "=OLD",
                )
            finally:
                result_workbook.close()
        finally:
            workbook.close()
            Path(file_path).unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
